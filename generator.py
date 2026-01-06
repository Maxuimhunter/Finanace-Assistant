# app.py
import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.chart import LineChart, PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import io
import re
import tempfile
from PIL import Image
import matplotlib.pyplot as plt
import seaborn as sns
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import os
import re
import json
from io import BytesIO, StringIO
import subprocess
import sys
from PIL import Image
import plotly.graph_objects as go
import plotly.express as px
from fpdf import FPDF
import base64
import io
from pathlib import Path
import re

# Set page config
st.set_page_config(
    page_title="Life & Budget Dashboard",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for styling
st.markdown("""
<style>
    .main {
        background-color: #f8f9fa;
    }
    .stButton>button {
        background-color: #6f42c1;
        color: white;
        border-radius: 8px;
        padding: 0.5rem 1rem;
    }
    .stButton>button:hover {
        background-color: #5a32a3;
        color: white;
    }
    .stTextInput>div>div>input {
        border: 1px solid #6f42c1;
        border-radius: 4px;
    }
    .css-1d391kg {
        padding-top: 3rem;
    }
</style>
""", unsafe_allow_html=True)

def create_welcome_guide(sheet):
    """Create a visually appealing Welcome Guide sheet."""
    # Set background color and hide gridlines
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = '6f42c1' # Purple

    # Set column widths for a better layout
    sheet.column_dimensions['A'].width = 4
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 60
    sheet.column_dimensions['D'].width = 4

    # Main Title
    sheet.merge_cells('B2:C2')
    title_cell = sheet['B2']
    title_cell.value = "🚀 Welcome to Your All-in-One Life Dashboard"
    title_cell.font = Font(size=24, bold=True, color='FFFFFF', name='Calibri')
    title_cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[2].height = 40

    # Subtitle
    sheet.merge_cells('B4:C4')
    subtitle_cell = sheet['B4']
    subtitle_cell.value = "Track finances, build habits, and organize your life with ease."
    subtitle_cell.font = Font(size=14, italic=True, color='595959', name='Calibri')
    subtitle_cell.alignment = Alignment(horizontal='center')
    sheet.row_dimensions[4].height = 25

    # Section: Core Features
    sheet.merge_cells('B6:C6')
    features_title = sheet['B6']
    features_title.value = "Core Features"
    features_title.font = Font(size=16, bold=True, color='6f42c1', name='Calibri')
    features_title.alignment = Alignment(horizontal='left')
    sheet.row_dimensions[6].height = 30

    features = [
        ("📊 Dashboard", "A high-level overview of your financial health and key life metrics."),
        ("💰 Financial Trackers", "Log income, expenses, savings, and investments in dedicated tabs."),
        ("💚 Health & Wellness", "Monitor weight, build habits, and focus on self-care."),
        ("📅 Life Organization", "Plan meals, manage cleaning schedules, and organize your time."),
        ("🤖 AI Insights", "Upload your sheet to get personalized financial recommendations.")
    ]

    thin_border = Border(left=Side(style='thin', color='DDDDDD'), 
                       right=Side(style='thin', color='DDDDDD'), 
                       top=Side(style='thin', color='DDDDDD'), 
                       bottom=Side(style='thin', color='DDDDDD'))

    for i, (title, desc) in enumerate(features, start=7):
        sheet.row_dimensions[i].height = 25
        # Feature Title
        title_cell = sheet.cell(row=i, column=2, value=title)
        title_cell.font = Font(size=12, bold=True, name='Calibri')
        title_cell.alignment = Alignment(vertical='center')
        # Feature Description
        desc_cell = sheet.cell(row=i, column=3, value=desc)
        desc_cell.font = Font(size=12, name='Calibri')
        desc_cell.alignment = Alignment(vertical='center', wrap_text=True)

    # Section: Getting Started
    start_row = len(features) + 9
    sheet.merge_cells(f'B{start_row}:C{start_row}')
    getting_started_title = sheet[f'B{start_row}']
    getting_started_title.value = "How to Get Started"
    getting_started_title.font = Font(size=16, bold=True, color='6f42c1', name='Calibri')
    sheet.row_dimensions[start_row].height = 30

    instructions = [
        ("1.", "Explore each tab to see what's available."),
        ("2.", "Fill in your data in the respective trackers (income, expenses, etc.)."),
        ("3.", "The 'Dashboard' and 'Charts' tabs will update automatically."),
        ("4.", "For AI analysis, save the file and upload it in the web app.")
    ]

    for i, (num, text) in enumerate(instructions, start=start_row + 1):
        sheet.row_dimensions[i].height = 25
        sheet.cell(row=i, column=2, value=num).font = Font(size=12, bold=True, name='Calibri')
        sheet.cell(row=i, column=3, value=text).font = Font(size=12, name='Calibri')

    # Footer
    footer_row = start_row + len(instructions) + 3
    sheet.merge_cells(f'B{footer_row}:C{footer_row}')
    footer_cell = sheet[f'B{footer_row}']
    footer_cell.value = "Happy tracking! ✨"
    footer_cell.font = Font(size=12, italic=True, color='595959', name='Calibri')
    footer_cell.alignment = Alignment(horizontal='center')


def worksheet_to_dataframe(sheet):
    """Converts an openpyxl worksheet to a pandas DataFrame.
    
    This function is designed to be robust and work with various Excel formats.
    It will:
    1. Find the first row that looks like a header
    2. Read all subsequent rows as data
    3. Clean up empty rows and columns
    4. Preserve all data (no aggressive filtering)
    """
    try:
        # Get all rows with values
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            return pd.DataFrame()

        # 1. Find the header row by looking for common header keywords
        header_row_index = -1
        for i, row in enumerate(all_rows):
            if not any(cell is None for cell in row[:5]):  # First 5 columns shouldn't be empty in a header
                header_candidates = [str(cell).lower() for cell in row if cell is not None]
                header_keywords = ['date', 'amount', 'source', 'description', 'category', 'type', 'name', 'item']
                if any(any(kw in str(cell).lower() for kw in header_keywords) for cell in row if cell):
                    header_row_index = i
                    break
        
        # If no header found, try to use the first non-empty row as header
        if header_row_index == -1:
            for i, row in enumerate(all_rows):
                if any(cell is not None for cell in row):
                    header_row_index = i
                    break
            
            if header_row_index == -1:
                return pd.DataFrame()  # No data found

        # 2. Get header and data
        header = [str(cell) if cell is not None else f'Column_{i+1}' 
                 for i, cell in enumerate(all_rows[header_row_index])]
        data = all_rows[header_row_index + 1:]

        # 3. Create DataFrame with proper handling of empty cells
        df = pd.DataFrame(data, columns=header)
        
        # 4. Clean the DataFrame
        # Drop completely empty rows and columns
        df = df.dropna(how='all')
        df = df.dropna(how='all', axis=1)
        
        # Convert all columns to string and strip whitespace
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.strip()
        
        # Remove completely empty rows again after type conversion
        df = df[df.astype(str).ne('None').any(axis=1)]
        df = df[df.astype(str).ne('').any(axis=1)]
        
        # Reset index after all filtering
        df.reset_index(drop=True, inplace=True)
        
        return df
        
    except Exception as e:
        print(f"Error processing sheet {sheet.title}: {str(e)}")
        return pd.DataFrame()

def is_merged_cell(sheet, row, col):
    """Check if a cell is part of a merged range."""
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and 
            merged_range.min_col <= col <= merged_range.max_col):
            return True
    return False

def create_pdf_report(month=None, sections=None):
    """Create a professional PDF report with charts and financial analysis."""
    try:
        # Set up matplotlib for better charts
        plt.style.use('default')
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Get styles
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#6f42c1'),
            alignment=TA_CENTER
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            textColor=colors.HexColor('#8b5cf6'),
            alignment=TA_CENTER,
            fontStyle='Italic'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor('#6f42c1')
        )
        
        # Content
        story = []
        
        # Title
        if month is None:
            current_month = datetime.date.today().strftime('%B %Y')
        else:
            current_month = f"{month} {datetime.date.today().year}"
        
        story.append(Paragraph(f"📊 Financial Report - {current_month}", title_style))
        story.append(Paragraph("Comprehensive financial analysis and insights", subtitle_style))
        story.append(Spacer(1, 20))
        
        # Sample data for demonstration
        weekly_income = [2500, 2800, 2400, 2600]
        weekly_expenses = [1800, 2000, 1700, 1900]
        weekly_savings = [700, 800, 700, 700]
        weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4']
        
        # 1. Key Metrics Section
        story.append(Paragraph("💰 Key Financial Metrics", heading_style))
        
        metrics_data = [
            ['Metric', 'Amount', 'Status'],
            ['Total Income', f'£{sum(weekly_income):,.0f}', '📈'],
            ['Total Expenses', f'£{sum(weekly_expenses):,.0f}', '💸'],
            ['Net Savings', f'£{sum(weekly_savings):,.0f}', '💰'],
            ['Savings Rate', f'{(sum(weekly_savings)/sum(weekly_income)*100):.1f}%', '🎯']
        ]
        
        metrics_table = Table(metrics_data, colWidths=[2.5*inch, 2*inch, 1*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6f42c1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0'))
        ]))
        
        story.append(metrics_table)
        story.append(Spacer(1, 20))
        
        # 2. Financial Summary Table (instead of charts to avoid image issues)
        story.append(Paragraph("📈 Weekly Financial Performance", heading_style))
        
        weekly_data = [['Week', 'Income', 'Expenses', 'Savings']]
        for i, week in enumerate(weeks):
            weekly_data.append([week, f'£{weekly_income[i]:,.0f}', f'£{weekly_expenses[i]:,.0f}', f'£{weekly_savings[i]:,.0f}'])
        
        weekly_table = Table(weekly_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        weekly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6f42c1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0'))
        ]))
        
        story.append(weekly_table)
        story.append(Spacer(1, 20))
        
        # 3. Budget Analysis Table
        story.append(Paragraph("📊 Budget vs Actual Analysis", heading_style))
        
        budget_categories = ['Housing', 'Food', 'Transport', 'Utilities', 'Entertainment', 'Other']
        budget_planned = [1100, 750, 450, 350, 250, 200]
        budget_actual = [1200, 800, 400, 300, 300, 200]
        
        budget_data = [['Category', 'Budget', 'Actual', 'Variance', 'Status']]
        for cat, planned, actual in zip(budget_categories, budget_planned, budget_actual):
            variance = actual - planned
            status = '✅ On Track' if actual <= planned else '⚠️ Over Budget'
            budget_data.append([cat, f'£{planned:,}', f'£{actual:,}', f'£{variance:,}', status])
        
        budget_table = Table(budget_data, colWidths=[1.5*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        budget_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6f42c1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0'))
        ]))
        
        story.append(budget_table)
        story.append(Spacer(1, 20))
        
        # 4. Financial Insights
        story.append(Paragraph("💡 Key Financial Insights", heading_style))
        
        monthly_income = sum(weekly_income)
        monthly_expenses = sum(weekly_expenses)
        monthly_savings = sum(weekly_savings)
        savings_rate = (monthly_savings/monthly_income*100) if monthly_income > 0 else 0
        
        insights = [
            f"• Your current savings rate is {savings_rate:.1f}%. Financial experts recommend saving at least 20% of your income.",
            f"• Total monthly expenses of £{monthly_expenses:,.0f} represent {(monthly_expenses/monthly_income*100):.1f}% of your income.",
            f"• Average weekly income is £{monthly_income/4:,.0f}. Consider ways to increase this through side hustles or career advancement.",
            f"• Housing costs account for {(1200/monthly_expenses*100):.1f}% of your expenses - ensure this stays below 30% of income.",
            f"• You have {len([x for x in budget_actual if x > budget_planned[budget_actual.index(x)]])} categories over budget. Review these areas for potential savings."
        ]
        
        for insight in insights:
            story.append(Paragraph(insight, styles['Normal']))
            story.append(Spacer(1, 6))
        
        # 5. Recommendations
        story.append(Paragraph("🎯 Personalized Recommendations", heading_style))
        
        recommendations = [
            "📈 Increase Income: Explore freelance opportunities or ask for a raise",
            "💰 Boost Savings: Set up automatic transfers to savings accounts",
            "📊 Track Expenses: Use budgeting apps to monitor spending patterns",
            "🎯 Set Goals: Establish specific financial targets for the next 6 months",
            "🔄 Review Regularly: Update your budget monthly to reflect changes"
        ]
        
        for rec in recommendations:
            story.append(Paragraph(rec, styles['Normal']))
            story.append(Spacer(1, 6))
        
        # Footer
        story.append(Spacer(1, 30))
        story.append(Paragraph(f"Generated on {datetime.date.today().strftime('%B %d, %Y')}", 
                              ParagraphStyle('Footer', parent=styles['Normal'], 
                                           fontSize=8, textColor=colors.grey, 
                                           alignment=TA_CENTER)))
        
        # Build PDF
        doc.build(story)
        
        # Get PDF bytes
        print(f"Error creating PDF: {str(e)}")
        return None

# Create enhanced charts
# create_enhanced_charts(wb, month):
# try:
#     # Create or get the Charts sheet
#     if 'Charts' in wb.sheetnames:
#         charts_sheet = wb['Charts']
#     else:
#         charts_sheet = wb.create_sheet('Charts')
# 
#     charts_sheet.sheet_view.showGridLines = False
# 
#     # Clear the sheet first
#     charts_sheet.delete_rows(1, charts_sheet.max_row or 100)
#     for col in charts_sheet.columns:
#         for cell in col:
#             cell.value = None
# 
#     # Enhanced title with gradient effect
#     charts_sheet.merge_cells('A1:M1')
#     charts_sheet['A1'] = f" {month} - Financial Analytics Dashboard"
#     charts_sheet['A1'].font = Font(size=20, bold=True, color='FFFFFF')
#     charts_sheet['A1'].fill = PatternFill(start_color='6f42c1', end_color='8b5cf6', fill_type="solid")
#     charts_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     charts_sheet.row_dimensions[1].height = 40
# 
#     # Add subtitle
#     charts_sheet.merge_cells('A2:M2')
#     charts_sheet['A2'] = "Comprehensive visual analysis of your financial performance"
#     charts_sheet['A2'].font = Font(size=12, color='8b5cf6', italic=True)
#     charts_sheet['A2'].alignment = Alignment(horizontal="center")
#     charts_sheet.row_dimensions[2].height = 25
# 
#     # Add decorative separator
#     charts_sheet.merge_cells('A3:M3')
#     charts_sheet['A3'] = "─" * 60
#     charts_sheet['A3'].font = Font(size=8, color='e0e0e0')
#     charts_sheet['A3'].alignment = Alignment(horizontal="center")
#     charts_sheet.row_dimensions[3].height = 10
# 
#     # Sample data for demonstration
#     weekly_income = [2500, 2800, 2400, 2600]
#     weekly_expenses = [1800, 2000, 1700, 1900]
#     weekly_savings = [700, 800, 700, 700]
#     weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4']
# 
#     # 1. Enhanced Key Metrics Cards (Top Section)
#     current_row = 4
#     monthly_income = sum(weekly_income)
#     monthly_expenses = sum(weekly_expenses)
#     monthly_savings = sum(weekly_savings)
# 
#     # Create professional metric cards
#     metrics = [
#         (" Total Income", f"£{monthly_income:,.0f}", "4CAF50", "A"),
#         (" Total Expenses", f"£{monthly_expenses:,.0f}", "F44336", "D"),
#         (" Net Savings", f"£{monthly_savings:,.0f}", "2196F3", "G"),
#         (" Savings Rate", f"{(monthly_savings/monthly_income*100 if monthly_income > 0 else 0):.1f}%", "9C27B0", "J"),
#         (" Avg Daily", f"£{monthly_income/30:,.0f}", "FF9800", "M")
#     ]
# 
#     for title, value, color, col in metrics:
#         # Card background with gradient effect
#         for row in range(current_row, current_row + 3):
#             for c in range(ord(col) - ord('A'), ord(col) - ord('A') + 1):
#                 cell = charts_sheet.cell(row=row, column=c + 1)
#                 cell.fill = PatternFill(start_color=f"{color}05", end_color=f"{color}10", fill_type="solid")
#                 cell.border = Border(
#                     left=Side(style='thin', color=f"{color}30"),
#                     right=Side(style='thin', color=f"{color}30"),
#                     top=Side(style='thin', color=f"{color}30"),
#                     bottom=Side(style='thin', color=f"{color}30")
#                 )
#         
#         # Title
#         charts_sheet[f'{col}{current_row}'] = title
#         charts_sheet[f'{col}{current_row}'].font = Font(size=9, bold=True, color=color)
#         charts_sheet[f'{col}{current_row}'].alignment = Alignment(horizontal="center")
#         
#         # Value
#         charts_sheet[f'{col}{current_row + 1}'] = value
#         charts_sheet[f'{col}{current_row + 1}'].font = Font(size=16, bold=True, color=color)
#         charts_sheet[f'{col}{current_row + 1}'].alignment = Alignment(horizontal="center")
# 
#     current_row += 5
# 
#     # 2. Enhanced Weekly Trend Chart (Left Side)
#     charts_sheet[f'A{current_row}'] = " Weekly Financial Performance"
#     charts_sheet[f'A{current_row}'].font = Font(bold=True, size=14, color='6f42c1')
#     current_row += 1
# 
#     # Write weekly data for line chart
#     for i, week in enumerate(weeks):
#         charts_sheet.cell(row=current_row + i, column=1, value=week)
#         charts_sheet.cell(row=current_row + i, column=2, value=weekly_income[i])
#         charts_sheet.cell(row=current_row + i, column=3, value=weekly_expenses[i])
#         charts_sheet.cell(row=current_row + i, column=4, value=weekly_savings[i])
# 
#     # Create enhanced line chart
#     line_chart = LineChart()
#     line_chart.title = "Weekly Income vs Expenses vs Savings"
#     line_chart.style = 13
#     line_chart.y_axis.title = 'Amount (£)'
#     line_chart.x_axis.title = 'Week'
#     line_chart.height = 10
#     line_chart.width = 15
# 
#     data = Reference(charts_sheet, min_col=2, min_row=current_row, max_row=current_row + 3, max_col=4)
#     line_chart.add_data(data, titles_from_data=True)
#     line_chart.set_categories(Reference(charts_sheet, min_col=1, min_row=current_row + 1, max_row=current_row + 3))
# 
#     # Enhanced styling for line chart
#     colors = ['4CAF50', 'F44336', '2196F3']
#     line_widths = [25000, 25000, 25000]
#     for i, series in enumerate(line_chart.series):
#         series.graphicalProperties.line.width = line_widths[i]
#         series.graphicalProperties.line.solidFill = colors[i]
#         series.smooth = True  # Make lines smooth
# 
#     charts_sheet.add_chart(line_chart, f'F{current_row + 5}')
# 
#     # 3. Enhanced Expense Breakdown (Right Side)
#     pie_start_row = current_row
#     charts_sheet[f'K{pie_start_row}'] = " Expense Distribution"
#     charts_sheet[f'K{pie_start_row}'].font = Font(bold=True, size=14, color='6f42c1')
# 
#     # Write expense data with better categories
#     expense_categories = [" Housing", " Food", " Transport", " Utilities", " Entertainment", " Other"]
#     expense_amounts = [1200, 800, 400, 300, 300, 200]
# 
#     for i, (cat, amt) in enumerate(zip(expense_categories, expense_amounts)):
#         charts_sheet.cell(row=pie_start_row + 1 + i, column=11, value=cat)
#         charts_sheet.cell(row=pie_start_row + 1 + i, column=12, value=amt)
# 
#     # Create enhanced pie chart
#     pie = PieChart()
#     labels = Reference(charts_sheet, min_col=11, min_row=pie_start_row + 1, max_row=pie_start_row + 6)
#     data = Reference(charts_sheet, min_col=12, min_row=pie_start_row, max_row=pie_start_row + 6)
#     pie.add_data(data, titles_from_data=True)
#     pie.set_categories(labels)
#     pie.title = "Monthly Expense Breakdown"
#     pie.height = 10
#     pie.width = 10
#     pie.legend.position = 'r'
#     pie.dataLabels = openpyxl.chart.label.DataLabelList()
#     pie.dataLabels.showPercent = True
#     pie.dataLabels.showVal = True
# 
#     # Enhanced colors for pie chart
#     colors = ['4CAF50', '8BC34A', 'FFC107', 'FF9800', 'F44336', '9C27B0']
#     for i, point in enumerate(pie.series[0].dPt):
#         point.graphicalProperties.solidFill = colors[i % len(colors)]
#         point.graphicalProperties.line.solidFill = "FFFFFF"
#         point.graphicalProperties.line.width = 10000
# 
#     charts_sheet.add_chart(pie, f'O{pie_start_row + 8}')
# 
#     # 4. Budget vs Actual Comparison (Bottom Section)
#     budget_start_row = current_row + 20
#     charts_sheet.merge_cells(f'A{budget_start_row}:M{budget_start_row}')
#     charts_sheet[f'A{budget_start_row}'] = " Budget vs Actual Analysis"
#     charts_sheet[f'A{budget_start_row}'].font = Font(bold=True, size=14, color='6f42c1')
#     charts_sheet[f'A{budget_start_row}'].alignment = Alignment(horizontal="center")
#     budget_start_row += 2
# 
#     # Enhanced budget data
#     budget_categories = [" Housing", " Food", " Transport", " Utilities", " Entertainment", " Other"]
#     budget_planned = [1100, 750, 450, 350, 250, 200]
#     budget_actual = [1200, 800, 400, 300, 300, 200]
# 
#     # Headers
#     headers = ["Category", "Budget", "Actual", "Variance", "% Used", "Status"]
#     for col, header in enumerate(headers, 1):
#         cell = charts_sheet.cell(row=budget_start_row, column=col, value=header)
#         cell.font = Font(bold=True, size=11, color='FFFFFF')
#         cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type="solid")
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.border = Border(
#             left=Side(style='thin'), 
#             right=Side(style='thin'), 
#             top=Side(style='thin'), 
#             bottom=Side(style='thin')
#         )
# 
#     # Budget data with enhanced formatting
#     for i, (cat, planned, actual) in enumerate(zip(budget_categories, budget_planned, budget_actual)):
#         row = budget_start_row + i + 1
#         variance = actual - planned
#         percent_used = (actual / planned * 100) if planned > 0 else 0
#         status = " On Track" if actual <= planned else " Over Budget"
#         
#         charts_sheet.cell(row=row, column=1, value=cat)
#         charts_sheet.cell(row=row, column=2, value=planned).number_format = '£#,##0'
#         charts_sheet.cell(row=row, column=3, value=actual).number_format = '£#,##0'
#         charts_sheet.cell(row=row, column=4, value=variance).number_format = '£#,##0'
#         charts_sheet.cell(row=row, column=5, value=percent_used).number_format = '0.0%'
#         charts_sheet.cell(row=row, column=6, value=status)
#         
#         # Color code variance
#         variance_cell = charts_sheet.cell(row=row, column=4)
#         if variance <= 0:
#             variance_cell.font = Font(color='4CAF50')
#         else:
#             variance_cell.font = Font(color='F44336')
# 
#             charts_sheet = wb.create_sheet('Charts')
#         
#         charts_sheet.sheet_view.showGridLines = False
#         
#         # Clear the sheet first
#         charts_sheet.delete_rows(1, charts_sheet.max_row or 100)
#         for col in charts_sheet.columns:
#             for cell in col:
#                 cell.value = None
# 
#         # Enhanced title with gradient effect
#         charts_sheet.merge_cells('A1:M1')
#         charts_sheet['A1'] = f" {month} - Financial Analytics Dashboard"
#         charts_sheet['A1'].font = Font(size=20, bold=True, color='FFFFFF')
#         charts_sheet['A1'].fill = PatternFill(start_color='6f42c1', end_color='8b5cf6', fill_type="solid")
#         charts_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
#         charts_sheet.row_dimensions[1].height = 40
# 
#         # Add subtitle
#         charts_sheet.merge_cells('A2:M2')
#         charts_sheet['A2'] = "Comprehensive visual analysis of your financial performance"
#         charts_sheet['A2'].font = Font(size=12, color='8b5cf6', italic=True)
#         charts_sheet['A2'].alignment = Alignment(horizontal="center")
#         charts_sheet.row_dimensions[2].height = 25
# 
#         # Add decorative separator
#         charts_sheet.merge_cells('A3:M3')
#         charts_sheet['A3'] = "─" * 60
#         charts_sheet['A3'].font = Font(size=8, color='e0e0e0')
#         charts_sheet['A3'].alignment = Alignment(horizontal="center")
#         charts_sheet.row_dimensions[3].height = 10
# 
#         # Sample data for demonstration
#         weekly_income = [2500, 2800, 2400, 2600]
#         weekly_expenses = [1800, 2000, 1700, 1900]
#         weekly_savings = [700, 800, 700, 700]
#         weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4']
# 
#         # 1. Enhanced Key Metrics Cards (Top Section)
#         current_row = 4
#         monthly_income = sum(weekly_income)
#         monthly_expenses = sum(weekly_expenses)
#         monthly_savings = sum(weekly_savings)
# 
#         # Create professional metric cards
#         metrics = [
#             (" Total Income", f"£{monthly_income:,.0f}", "4CAF50", "A"),
#             (" Total Expenses", f"£{monthly_expenses:,.0f}", "F44336", "D"),
#             (" Net Savings", f"£{monthly_savings:,.0f}", "2196F3", "G"),
#             (" Savings Rate", f"{(monthly_savings/monthly_income*100 if monthly_income > 0 else 0):.1f}%", "9C27B0", "J"),
#             (" Avg Daily", f"£{monthly_income/30:,.0f}", "FF9800", "M")
#         ]
# 
#         for title, value, color, col in metrics:
#             # Card background with gradient effect
#             for row in range(current_row, current_row + 3):
#                 for c in range(ord(col) - ord('A'), ord(col) - ord('A') + 1):
#                     cell = charts_sheet.cell(row=row, column=c + 1)
#                     cell.fill = PatternFill(start_color=f"{color}05", end_color=f"{color}10", fill_type="solid")
#                     cell.border = Border(
#                         left=Side(style='thin', color=f"{color}30"),
#                         right=Side(style='thin', color=f"{color}30"),
#                         top=Side(style='thin', color=f"{color}30"),
#                         bottom=Side(style='thin', color=f"{color}30")
#                     )
#             
#             # Title
#             charts_sheet[f'{col}{current_row}'] = title
#             charts_sheet[f'{col}{current_row}'].font = Font(size=9, bold=True, color=color)
#             charts_sheet[f'{col}{current_row}'].alignment = Alignment(horizontal="center")
#             
#             # Value
#             charts_sheet[f'{col}{current_row + 1}'] = value
#             charts_sheet[f'{col}{current_row + 1}'].font = Font(size=16, bold=True, color=color)
#             charts_sheet[f'{col}{current_row + 1}'].alignment = Alignment(horizontal="center")
# 
#         current_row += 5
# 
#         # 2. Enhanced Weekly Trend Chart (Left Side)
#         charts_sheet[f'A{current_row}'] = " Weekly Financial Performance"
#         charts_sheet[f'A{current_row}'].font = Font(bold=True, size=14, color='6f42c1')
#         current_row += 1
# 
#         # Write weekly data for line chart
#         for i, week in enumerate(weeks):
#             charts_sheet.cell(row=current_row + i, column=1, value=week)
#             charts_sheet.cell(row=current_row + i, column=2, value=weekly_income[i])
#             charts_sheet.cell(row=current_row + i, column=3, value=weekly_expenses[i])
#             charts_sheet.cell(row=current_row + i, column=4, value=weekly_savings[i])
# 
#         # Create enhanced line chart
#         line_chart = LineChart()
#         line_chart.title = "Weekly Income vs Expenses vs Savings"
#         line_chart.style = 13
#         line_chart.y_axis.title = 'Amount (£)'
#         line_chart.x_axis.title = 'Week'
#         line_chart.height = 10
#         line_chart.width = 15
# 
#         data = Reference(charts_sheet, min_col=2, min_row=current_row, max_row=current_row + 3, max_col=4)
#         line_chart.add_data(data, titles_from_data=True)
#         line_chart.set_categories(Reference(charts_sheet, min_col=1, min_row=current_row + 1, max_row=current_row + 3))
# 
#         # Enhanced styling for line chart
#         colors = ['4CAF50', 'F44336', '2196F3']
#         line_widths = [25000, 25000, 25000]
#         for i, series in enumerate(line_chart.series):
#             series.graphicalProperties.line.width = line_widths[i]
#             series.graphicalProperties.line.solidFill = colors[i]
#             series.smooth = True  # Make lines smooth
# 
#         charts_sheet.add_chart(line_chart, f'F{current_row + 5}')
# 
#         # 3. Enhanced Expense Breakdown (Right Side)
#         pie_start_row = current_row
#         charts_sheet[f'K{pie_start_row}'] = " Expense Distribution"
#         charts_sheet[f'K{pie_start_row}'].font = Font(bold=True, size=14, color='6f42c1')
# 
#         # Write expense data with better categories
#         expense_categories = [" Housing", " Food", " Transport", " Utilities", " Entertainment", " Other"]
#         expense_amounts = [1200, 800, 400, 300, 300, 200]
# 
#         for i, (cat, amt) in enumerate(zip(expense_categories, expense_amounts)):
#             charts_sheet.cell(row=pie_start_row + 1 + i, column=11, value=cat)
#             charts_sheet.cell(row=pie_start_row + 1 + i, column=12, value=amt)
# 
#         # Create enhanced pie chart
#         pie = PieChart()
#         labels = Reference(charts_sheet, min_col=11, min_row=pie_start_row + 1, max_row=pie_start_row + 6)
#         data = Reference(charts_sheet, min_col=12, min_row=pie_start_row, max_row=pie_start_row + 6)
#         pie.add_data(data, titles_from_data=True)
#         pie.set_categories(labels)
#         pie.title = "Monthly Expense Breakdown"
#         pie.height = 10
#         pie.width = 10
#         pie.legend.position = 'r'
#         pie.dataLabels = openpyxl.chart.label.DataLabelList()
#         pie.dataLabels.showPercent = True
#         pie.dataLabels.showVal = True
# 
#         # Enhanced colors for pie chart
#         colors = ['4CAF50', '8BC34A', 'FFC107', 'FF9800', 'F44336', '9C27B0']
#         for i, point in enumerate(pie.series[0].dPt):
#             point.graphicalProperties.solidFill = colors[i % len(colors)]
#             point.graphicalProperties.line.solidFill = "FFFFFF"
#             point.graphicalProperties.line.width = 10000
# 
#         charts_sheet.add_chart(pie, f'O{pie_start_row + 8}')
# 
#         # 4. Budget vs Actual Comparison (Bottom Section)
#         budget_start_row = current_row + 20
#         charts_sheet.merge_cells(f'A{budget_start_row}:M{budget_start_row}')
#         charts_sheet[f'A{budget_start_row}'] = " Budget vs Actual Analysis"
#         charts_sheet[f'A{budget_start_row}'].font = Font(bold=True, size=14, color='6f42c1')
#         charts_sheet[f'A{budget_start_row}'].alignment = Alignment(horizontal="center")
#         budget_start_row += 2
# 
#         # Enhanced budget data
#         budget_categories = [" Housing", " Food", " Transport", " Utilities", " Entertainment", " Other"]
#         budget_planned = [1100, 750, 450, 350, 250, 200]
#         budget_actual = [1200, 800, 400, 300, 300, 200]
# 
#         # Headers
#         headers = ["Category", "Budget", "Actual", "Variance", "% Used", "Status"]
#         for col, header in enumerate(headers, 1):
#             cell = charts_sheet.cell(row=budget_start_row, column=col, value=header)
#             cell.font = Font(bold=True, size=11, color='FFFFFF')
#             cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type="solid")
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#             cell.border = Border(
#                 left=Side(style='thin'), 
#                 right=Side(style='thin'), 
#                 top=Side(style='thin'), 
#                 bottom=Side(style='thin')
#             )
# 
#         # Budget data with enhanced formatting
#         for i, (cat, planned, actual) in enumerate(zip(budget_categories, budget_planned, budget_actual)):
#             row = budget_start_row + i + 1
#             variance = actual - planned
#             percent_used = (actual / planned * 100) if planned > 0 else 0
#             status = " On Track" if actual <= planned else " Over Budget"
#             
#             charts_sheet.cell(row=row, column=1, value=cat)
#             charts_sheet.cell(row=row, column=2, value=planned).number_format = '£#,##0'
#             charts_sheet.cell(row=row, column=3, value=actual).number_format = '£#,##0'
#             charts_sheet.cell(row=row, column=4, value=variance).number_format = '£#,##0'
#             charts_sheet.cell(row=row, column=5, value=percent_used).number_format = '0.0%'
#             charts_sheet.cell(row=row, column=6, value=status)
#             
#             # Color code variance
#             variance_cell = charts_sheet.cell(row=row, column=4)
#             if variance <= 0:
#                 variance_cell.font = Font(color='4CAF50')
#             else:
#                 variance_cell.font = Font(color='F44336')
# 
#             charts_sheet = wb.create_sheet('Charts')
#         
#         charts_sheet.sheet_view.showGridLines = False
#         
#         # Clear the sheet first
#         charts_sheet.delete_rows(1, charts_sheet.max_row or 100)
#         for col in charts_sheet.columns:
#             for cell in col:
#                 cell.value = None

        # Enhanced title with gradient effect
        charts_sheet.merge_cells('A1:M1')
        charts_sheet['A1'] = f"📊 {month} - Financial Analytics Dashboard"
        charts_sheet['A1'].font = Font(size=20, bold=True, color='FFFFFF')
        charts_sheet['A1'].fill = PatternFill(start_color='6f42c1', end_color='8b5cf6', fill_type="solid")
        charts_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        charts_sheet.row_dimensions[1].height = 40

        # Add subtitle
        charts_sheet.merge_cells('A2:M2')
        charts_sheet['A2'] = "Comprehensive visual analysis of your financial performance"
        charts_sheet['A2'].font = Font(size=12, color='8b5cf6', italic=True)
        charts_sheet['A2'].alignment = Alignment(horizontal="center")
        charts_sheet.row_dimensions[2].height = 25

        # Add decorative separator
        charts_sheet.merge_cells('A3:M3')
        charts_sheet['A3'] = "─" * 60
        charts_sheet['A3'].font = Font(size=8, color='e0e0e0')
        charts_sheet['A3'].alignment = Alignment(horizontal="center")
        charts_sheet.row_dimensions[3].height = 10

        # Sample data for demonstration
        weekly_income = [2500, 2800, 2400, 2600]
        weekly_expenses = [1800, 2000, 1700, 1900]
        weekly_savings = [700, 800, 700, 700]
        weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4']

        # 1. Enhanced Key Metrics Cards (Top Section)
        current_row = 4
        monthly_income = sum(weekly_income)
        monthly_expenses = sum(weekly_expenses)
        monthly_savings = sum(weekly_savings)

        # Create professional metric cards
        metrics = [
            ("💰 Total Income", f"£{monthly_income:,.0f}", "4CAF50", "A"),
            ("💸 Total Expenses", f"£{monthly_expenses:,.0f}", "F44336", "D"),
            ("📈 Net Savings", f"£{monthly_savings:,.0f}", "2196F3", "G"),
            ("🎯 Savings Rate", f"{(monthly_savings/monthly_income*100 if monthly_income > 0 else 0):.1f}%", "9C27B0", "J"),
            ("📊 Avg Daily", f"£{monthly_income/30:,.0f}", "FF9800", "M")
        ]

        for title, value, color, col in metrics:
            # Card background with gradient effect
            for row in range(current_row, current_row + 3):
                for c in range(ord(col) - ord('A'), ord(col) - ord('A') + 1):
                    cell = charts_sheet.cell(row=row, column=c + 1)
                    cell.fill = PatternFill(start_color=f"{color}05", end_color=f"{color}10", fill_type="solid")
                    cell.border = Border(
                        left=Side(style='thin', color=f"{color}30"),
                        right=Side(style='thin', color=f"{color}30"),
                        top=Side(style='thin', color=f"{color}30"),
                        bottom=Side(style='thin', color=f"{color}30")
                    )
            
            # Title
            charts_sheet[f'{col}{current_row}'] = title
            charts_sheet[f'{col}{current_row}'].font = Font(size=9, bold=True, color=color)
            charts_sheet[f'{col}{current_row}'].alignment = Alignment(horizontal="center")
            
            # Value
            charts_sheet[f'{col}{current_row + 1}'] = value
            charts_sheet[f'{col}{current_row + 1}'].font = Font(size=16, bold=True, color=color)
            charts_sheet[f'{col}{current_row + 1}'].alignment = Alignment(horizontal="center")

        current_row += 5

        # 2. Enhanced Weekly Trend Chart (Left Side)
        charts_sheet[f'A{current_row}'] = "📈 Weekly Financial Performance"
        charts_sheet[f'A{current_row}'].font = Font(bold=True, size=14, color='6f42c1')
        current_row += 1

        # Write weekly data for line chart
        for i, week in enumerate(weeks):
            charts_sheet.cell(row=current_row + i, column=1, value=week)
            charts_sheet.cell(row=current_row + i, column=2, value=weekly_income[i])
            charts_sheet.cell(row=current_row + i, column=3, value=weekly_expenses[i])
            charts_sheet.cell(row=current_row + i, column=4, value=weekly_savings[i])

        # Create enhanced line chart
        line_chart = LineChart()
        line_chart.title = "Weekly Income vs Expenses vs Savings"
        line_chart.style = 13
        line_chart.y_axis.title = 'Amount (£)'
        line_chart.x_axis.title = 'Week'
        line_chart.height = 10
        line_chart.width = 15

        data = Reference(charts_sheet, min_col=2, min_row=current_row, max_row=current_row + 3, max_col=4)
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(Reference(charts_sheet, min_col=1, min_row=current_row + 1, max_row=current_row + 3))

        # Enhanced styling for line chart
        colors = ['4CAF50', 'F44336', '2196F3']
        line_widths = [25000, 25000, 25000]
        for i, series in enumerate(line_chart.series):
            series.graphicalProperties.line.width = line_widths[i]
            series.graphicalProperties.line.solidFill = colors[i]
            series.smooth = True  # Make lines smooth

        charts_sheet.add_chart(line_chart, f'F{current_row + 5}')

        # 3. Enhanced Expense Breakdown (Right Side)
        pie_start_row = current_row
        charts_sheet[f'K{pie_start_row}'] = "🥧 Expense Distribution"
        charts_sheet[f'K{pie_start_row}'].font = Font(bold=True, size=14, color='6f42c1')

        # Write expense data with better categories
        expense_categories = ["🏠 Housing", "🍔 Food", "🚗 Transport", "💡 Utilities", "🎮 Entertainment", "📱 Other"]
        expense_amounts = [1200, 800, 400, 300, 300, 200]

        for i, (cat, amt) in enumerate(zip(expense_categories, expense_amounts)):
            charts_sheet.cell(row=pie_start_row + 1 + i, column=11, value=cat)
            charts_sheet.cell(row=pie_start_row + 1 + i, column=12, value=amt)

        # Create enhanced pie chart
        pie = PieChart()
        labels = Reference(charts_sheet, min_col=11, min_row=pie_start_row + 1, max_row=pie_start_row + 6)
        data = Reference(charts_sheet, min_col=12, min_row=pie_start_row, max_row=pie_start_row + 6)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Monthly Expense Breakdown"
        pie.height = 10
        pie.width = 10
        pie.legend.position = 'r'
        pie.dataLabels = openpyxl.chart.label.DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True

        # Enhanced colors for pie chart
        colors = ['4CAF50', '8BC34A', 'FFC107', 'FF9800', 'F44336', '9C27B0']
        for i, point in enumerate(pie.series[0].dPt):
            point.graphicalProperties.solidFill = colors[i % len(colors)]
            point.graphicalProperties.line.solidFill = "FFFFFF"
            point.graphicalProperties.line.width = 10000

        charts_sheet.add_chart(pie, f'O{pie_start_row + 8}')

        # 4. Budget vs Actual Comparison (Bottom Section)
        budget_start_row = current_row + 20
        charts_sheet.merge_cells(f'A{budget_start_row}:M{budget_start_row}')
        charts_sheet[f'A{budget_start_row}'] = "📊 Budget vs Actual Analysis"
        charts_sheet[f'A{budget_start_row}'].font = Font(bold=True, size=14, color='6f42c1')
        charts_sheet[f'A{budget_start_row}'].alignment = Alignment(horizontal="center")
        budget_start_row += 2

        # Enhanced budget data
        budget_categories = ["🏠 Housing", "🍔 Food", "🚗 Transport", "💡 Utilities", "🎮 Entertainment", "📱 Other"]
        budget_planned = [1100, 750, 450, 350, 250, 200]
        budget_actual = [1200, 800, 400, 300, 300, 200]

        # Headers
        headers = ["Category", "Budget", "Actual", "Variance", "% Used", "Status"]
        for col, header in enumerate(headers, 1):
            cell = charts_sheet.cell(row=budget_start_row, column=col, value=header)
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

        # Budget data with enhanced formatting
        for i, (cat, planned, actual) in enumerate(zip(budget_categories, budget_planned, budget_actual)):
            row = budget_start_row + i + 1
            variance = actual - planned
            percent_used = (actual / planned * 100) if planned > 0 else 0
            status = "✅ On Track" if actual <= planned else "⚠️ Over Budget"
            
            charts_sheet.cell(row=row, column=1, value=cat)
            charts_sheet.cell(row=row, column=2, value=planned).number_format = '£#,##0'
            charts_sheet.cell(row=row, column=3, value=actual).number_format = '£#,##0'
            charts_sheet.cell(row=row, column=4, value=variance).number_format = '£#,##0'
            charts_sheet.cell(row=row, column=5, value=percent_used).number_format = '0.0%'
            charts_sheet.cell(row=row, column=6, value=status)
            
            # Color code variance
            variance_cell = charts_sheet.cell(row=row, column=4)
            if variance <= 0:
                variance_cell.font = Font(color='4CAF50')
            else:
                variance_cell.font = Font(color='F44336')
            
            # Color code percent used
            percent_cell = charts_sheet.cell(row=row, column=5)
            if percent_used <= 100:
                percent_cell.font = Font(color='4CAF50')
            else:
                percent_cell.font = Font(color='F44336')

            # Add borders
            for col in range(1, 7):
                cell = charts_sheet.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin', color='e0e0e0'), 
                    right=Side(style='thin', color='e0e0e0'), 
                    top=Side(style='thin', color='e0e0e0'), 
                    bottom=Side(style='thin', color='e0e0e0')
                )

        # 6. Create 3D Pie Chart for Budget Comparison
        pie_chart_3d = PieChart()
        pie_chart_3d.title = "Budget vs Actual Comparison"
        pie_chart_3d.style = 12
        pie_chart_3d.height = 8
        pie_chart_3d.width = 12
        pie_chart_3d.legend.position = 'r'

        pie_data = Reference(charts_sheet, min_col=2, min_row=budget_start_row, max_row=budget_start_row + 6, max_col=3)
        pie_chart_3d.add_data(pie_data, titles_from_data=True)
        pie_chart_3d.set_categories(Reference(charts_sheet, min_col=1, min_row=budget_start_row + 1, max_row=budget_start_row + 6))

        # Style pie chart
        colors = ['6f42c1', '8b5cf6', '4CAF50', 'F44336', 'FF9800', '9C27B0']
        for i, point in enumerate(pie_chart_3d.series[0].dPt):
            point.graphicalProperties.solidFill = colors[i % len(colors)]

        charts_sheet.add_chart(pie_chart_3d, f'H{budget_start_row + 10}')


        # 6. Add Financial Insights Section
        insights_start_row = budget_start_row + 25
        charts_sheet.merge_cells(f'A{insights_start_row}:M{insights_start_row}')
        charts_sheet[f'A{insights_start_row}'] = "💡 Key Financial Insights"
        charts_sheet[f'A{insights_start_row}'].font = Font(bold=True, size=14, color='6f42c1')
        charts_sheet[f'A{insights_start_row}'].alignment = Alignment(horizontal="center")
        insights_start_row += 2

        insights = [
            f"• Your savings rate is {(monthly_savings/monthly_income*100 if monthly_income > 0 else 0):.1f}%, aim for 20%+",
            f"• Total monthly expenses: £{monthly_expenses:,.0f} - review categories over budget",
            f"• Average weekly income: £{monthly_income/4:,.0f} - look for opportunities to increase",
            f"• Consider reducing entertainment budget by 10% for additional savings"
        ]

        for i, insight in enumerate(insights, start=insights_start_row):
            charts_sheet[f'A{i}'] = insight
            charts_sheet[f'A{i}'].font = Font(size=10, color='666666')
            charts_sheet.row_dimensions[i].height = 20

        # Auto-size columns
        for column in charts_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            charts_sheet.column_dimensions[column_letter].width = min(adjusted_width, 25)

        return True
    except Exception as e:
        print(f"Error creating charts: {str(e)}")
        return False

        # Data
        for i, (cat, planned, actual) in enumerate(zip(budget_categories, budget_planned, budget_actual)):
            row = budget_start_row + i + 1
            variance = actual - planned
            
            charts_sheet.cell(row=row, column=1, value=cat)
            charts_sheet.cell(row=row, column=2, value=planned).number_format = '$#,##0'
            charts_sheet.cell(row=row, column=3, value=actual).number_format = '$#,##0'
            charts_sheet.cell(row=row, column=4, value=variance).number_format = '$#,##0'
            
            # Color code variance
            if variance > 0:
                charts_sheet.cell(row=row, column=4).font = Font(color='F44336')  # Red for over budget
            elif variance < 0:
                charts_sheet.cell(row=row, column=4).font = Font(color='4CAF50')  # Green for under budget

        # Create bar chart for budget comparison
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 12
        bar_chart.title = "Budget vs Actual"
        bar_chart.y_axis.title = 'Amount ($)'
        bar_chart.height = 6
        bar_chart.width = 15

        data = Reference(charts_sheet, min_col=2, min_row=budget_start_row, 
                        max_row=budget_start_row + len(budget_categories), max_col=3)
        cats = Reference(charts_sheet, min_col=1, min_row=budget_start_row + 1, 
                        max_row=budget_start_row + len(budget_categories))
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)

        # Style colors
        bar_chart.series[0].graphicalProperties.solidFill = "4CAF50"  # Green for Budget
        bar_chart.series[1].graphicalProperties.solidFill = "2196F3"  # Blue for Actual

        charts_sheet.add_chart(bar_chart, f'I{budget_start_row + len(budget_categories) + 3}')

        # Set column widths for clean look
        for col in ['A', 'B', 'C', 'D', 'E']:
            charts_sheet.column_dimensions[col].width = 12
        for col in ['F', 'G', 'H']:
            charts_sheet.column_dimensions[col].width = 10
        for col in ['I', 'J', 'K', 'L']:
            charts_sheet.column_dimensions[col].width = 15

        # Add footer
        footer_row = budget_start_row + len(budget_categories) + 12
        charts_sheet.merge_cells(f'A{footer_row}:H{footer_row}')
        charts_sheet[f'A{footer_row}'] = f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} • Financial Dashboard"
        charts_sheet[f'A{footer_row}'].font = Font(size=9, color='999999', italic=True)
        charts_sheet[f'A{footer_row}'].alignment = Alignment(horizontal="right")

    except Exception as e:
        print(f"Error creating charts: {str(e)}")

def create_chart_header(sheet, title, row, color='6f42c1'):
    """Create a styled header for charts"""
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=14, color='FFFFFF')
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[row].height = 25
    return row + 2

def create_ai_insights_placeholder(sheet):
    """Create a placeholder for AI insights"""
    sheet.merge_cells('A1:J1')
    sheet['A1'] = "🤖 AI Insights"
    sheet['A1'].font = Font(size=20, bold=True, color='FFFFFF')
    sheet['A1'].fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type="solid")
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 40
    
    sheet.merge_cells('A2:J2')
    sheet['A2'] = "AI-powered financial analysis and recommendations"
    sheet['A2'].font = Font(size=12, color='6f42c1', italic=True)
    sheet['A2'].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[2].height = 25
    
    # Add placeholder content
    placeholder_text = [
        "AI Insights will be generated here when you upload your data for analysis.",
        "",
        "Features:",
        "• Spending pattern analysis",
        "• Budget recommendations",
        "• Savings optimization suggestions",
        "• Financial health score",
        "• Predictive insights",
        "",
        "To generate insights, go to the 'AI Insights' page and upload your completed template."
    ]
    
    for i, text in enumerate(placeholder_text, start=4):
        sheet[f'A{i}'] = text
        if text.startswith("•"):
            sheet[f'A{i}'].font = Font(color='6f42c1')

def create_dashboard(sheet, month):
    """Create the dashboard sheet with enhanced financial overview"""
    # Create header with gradient effect
    sheet.merge_cells('A1:L1')
    sheet['A1'] = f"📊 {month} - Financial Dashboard"
    sheet['A1'].font = Font(size=22, bold=True, color='FFFFFF')
    sheet['A1'].fill = PatternFill(start_color='6f42c1', end_color='8b5cf6', fill_type="solid")
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 45
    
    # Add subtitle with better styling
    sheet.merge_cells('A2:L2')
    sheet['A2'] = "Your complete financial overview at a glance"
    sheet['A2'].font = Font(size=13, color='8b5cf6', italic=True)
    sheet['A2'].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[2].height = 25
    
    # Add decorative separator
    sheet.merge_cells('A3:L3')
    sheet['A3'] = "─" * 50
    sheet['A3'].font = Font(size=8, color='e0e0e0')
    sheet['A3'].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[3].height = 10
    
    # Key Metrics Cards Section
    sheet['A4'] = "📈 Key Performance Indicators"
    sheet['A4'].font = Font(bold=True, size=16, color='6f42c1')
    sheet.row_dimensions[4].height = 30
    
    # Create metric cards with better styling
    metrics_data = [
        ("Total Income", "£0.00", "📈", "4CAF50", "A"),
        ("Total Expenses", "£0.00", "💸", "F44336", "D"),
        ("Net Savings", "£0.00", "💰", "2196F3", "G"),
        ("Savings Rate", "0.0%", "📊", "9C27B0", "J")
    ]
    
    for title, value, emoji, color, col in metrics_data:
        # Card background
        for row in range(5, 8):
            for c in range(ord(col) - ord('A'), ord(col) - ord('A') + 2):
                cell = sheet.cell(row=row, column=c + 1)
                cell.fill = PatternFill(start_color=f"{color}08", end_color=f"{color}08", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color=f"{color}40"),
                    right=Side(style='thin', color=f"{color}40"),
                    top=Side(style='thin', color=f"{color}40"),
                    bottom=Side(style='thin', color=f"{color}40")
                )
        
        # Emoji
        sheet[f'{col}5'] = emoji
        sheet[f'{col}5'].font = Font(size=16)
        sheet[f'{col}5'].alignment = Alignment(horizontal="center")
        
        # Title
        sheet[f'{col}6'] = title
        sheet[f'{col}6'].font = Font(size=10, bold=True, color=color)
        sheet[f'{col}6'].alignment = Alignment(horizontal="center")
        
        # Value
        sheet[f'{col}7'] = value
        sheet[f'{col}7'].font = Font(size=14, bold=True, color=color)
        sheet[f'{col}7'].alignment = Alignment(horizontal="center")
    
    # Financial Summary Section
    sheet.row_dimensions[9].height = 20
    sheet['A10'] = "💼 Financial Summary"
    sheet['A10'].font = Font(bold=True, size=16, color='6f42c1')
    sheet.row_dimensions[10].height = 30
    
    # Enhanced headers for summary table
    headers = ["Category", "Planned", "Actual", "Difference", "% of Budget", "Status"]
    header_colors = ["6f42c1", "6f42c1", "6f42c1", "6f42c1", "6f42c1", "6f42c1"]
    
    for col, (header, color) in enumerate(zip(headers, header_colors), 1):
        cell = sheet.cell(row=11, column=col, value=header)
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Enhanced summary data with formulas
    summary_data = [
        ["Income", 5000, "='Income Tracker'!$C$10", "=C12-B12", "=IF(B12=0,0,C12/B12*100)", 
         "=IF(C12>=B12,\"✅ On Track\",\"⚠️ Review\")"],
        ["Expenses", 3200, "='Expense Tracker'!$C$47", "=C13-B13", "=IF(B13=0,0,C13/B13*100)",
         "=IF(C13<=B13,\"✅ On Budget\",\"⚠️ Over Budget\")"],
        ["Savings", 1800, "='Savings Tracker'!$D$11", "=C14-B14", "=IF(B14=0,0,C14/B14*100)",
         "=IF(C14>=B14,\"✅ Goal Met\",\"📈 Keep Going\")"],
        ["Investments", 500, "='Stock Tracker'!$F$10", "=C15-B15", "=IF(B15=0,0,C15/B15*100)",
         "=IF(C15>=B15,\"📈 Profit\",\"📉 Review\")"],
        ["Total", "=SUM(B12:B15)", "=SUM(C12:C15)", "=C16-B16", "=IF(B16=0,0,C16/B16*100)",
         "=IF(C16>=B16,\"✅ Positive\",\"⚠️ Negative\")"]
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=12):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:  # Category column
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            elif col_idx == 6:  # Status column
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center")
            else:  # Numeric columns
                cell.number_format = '£#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            
            cell.border = Border(
                left=Side(style='thin', color='e0e0e0'), 
                right=Side(style='thin', color='e0e0e0'), 
                top=Side(style='thin', color='e0e0e0'), 
                bottom=Side(style='thin', color='e0e0e0')
            )
    
    # Monthly Purchases Summary Section
    sheet.row_dimensions[17].height = 20
    sheet['A18'] = "🛍️ Monthly Purchases Analysis"
    sheet['A18'].font = Font(bold=True, size=16, color='6f42c1')
    sheet.row_dimensions[18].height = 30
    
    # Enhanced headers for monthly purchases
    mp_headers = ["Type", "Count", "Total Amount", "Average", "Monthly Impact", "Trend"]
    for col, header in enumerate(mp_headers, 1):
        cell = sheet.cell(row=19, column=col, value=header)
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color='8b5cf6', end_color='8b5cf6', fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Enhanced monthly purchases data
    mp_data = [
        ["Subscriptions", "=COUNT('Monthly Purchases'!C:C)", "=SUM('Monthly Purchases'!C:C)", 
         "=IF(B20=0,0,C20/B20)", "=C20/('Expense Tracker'!$C$47)", "📊 Stable"],
        ["One-Time", "=COUNTIF('Monthly Purchases'!D:D,\"One-Time\")", 
         "=SUMIF('Monthly Purchases'!D:D,\"One-Time\",'Monthly Purchases'!C:C)", 
         "=IF(B21=0,0,C21/B21)", "=C21/('Expense Tracker'!$C$47)", "📈 Variable"],
        ["Total", "=B20+B21", "=C20+C21", "=IF(B22=0,0,C22/B22)", "=C22/('Expense Tracker'!$C$47)", "📊 Combined"]
    ]
    
    for row_idx, row_data in enumerate(mp_data, start=20):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:  # Type column
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            elif col_idx == 6:  # Trend column
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center")
            else:  # Numeric columns
                cell.number_format = '£#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            
            cell.border = Border(
                left=Side(style='thin', color='e0e0e0'), 
                right=Side(style='thin', color='e0e0e0'), 
                top=Side(style='thin', color='e0e0e0'), 
                bottom=Side(style='thin', color='e0e0e0')
            )
    
    # Add insights section
    sheet.row_dimensions[23].height = 20
    sheet['A24'] = "💡 Financial Insights"
    sheet['A24'].font = Font(bold=True, size=16, color='6f42c1')
    sheet.row_dimensions[24].height = 30
    
    insights = [
        "• Track your spending patterns to identify areas for improvement",
        "• Aim to save at least 20% of your monthly income",
        "• Review subscriptions regularly to eliminate unnecessary expenses",
        "• Consider setting up automatic transfers for savings goals"
    ]
    
    for i, insight in enumerate(insights, start=25):
        sheet[f'A{i}'] = insight
        sheet[f'A{i}'].font = Font(size=10, color='666666')
        sheet.row_dimensions[i].height = 20
    
    # Auto-size columns
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = min(adjusted_width, 20)

def create_excel_template(month=None, sections=None):
    """
    Create an Excel template for a specific month with the specified sections.
    
    Args:
        month (str, optional): The month to generate the template for (e.g., 'Jan', 'Feb'). 
                              If None, uses current month.
        sections (list, optional): List of section names to include. If None, includes all sections.
                                  Possible values: 'dashboard', 'income', 'expenses', 'savings', 'stocks', 
                                  'weight', 'habits', 'cleaning', 'meals', 'timetable'
    """
    # Define constants
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
             'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4']
    
    # If no month specified, use current month
    if month is None:
        current_month = datetime.date.today().month - 1  # 0-based index
        month = months[current_month]
    
    # Define all possible tabs with their configurations at the beginning
    all_tabs = {
        "Income Tracker": {
            "headers": ["Date", "Source", "Amount", "Category", "Notes"],
            "color": "B0E0E6",  # Blue
            "sample_data": [
                [datetime.date.today(), "Salary", 3000, "Primary Income", "Monthly salary"],
                [datetime.date.today().replace(day=15), "Freelance", 1000, "Side Hustle", "Project X"]
            ]
        },
        "Expense Tracker": {
            "headers": ["Date", "Description", "Amount", "Category", "Payment Method", "Notes"],
            "color": "FFB6C1",  # Pink
            "sample_data": [
                [datetime.date.today(), "Groceries", 150, "Food", "Credit Card", "Weekly shopping"],
                [datetime.date.today(), "Electricity", 80, "Utilities", "Direct Debit", "Monthly bill"]
            ]
        },
        "Monthly Purchases": {
            "headers": ["Date", "Item", "Amount", "Type", "Category", "Notes"],
            "color": "D8BFD8",  # Light Purple
            "sample_data": [
                [datetime.date.today().replace(day=1), "Netflix", 15.99, "Subscription", "Entertainment", "Monthly plan"],
                [datetime.date.today().replace(day=5), "Gym Membership", 45.00, "Subscription", "Health", "Monthly membership"],
                [datetime.date.today().replace(day=10), "Office Chair", 199.99, "One-Time", "Furniture", "Ergonomic chair"]
            ]
        },
        "Savings Tracker": {
            "headers": ["Date", "Goal", "Target Amount", "Current Amount", "% Complete"],
            "color": "98FB98",  # Green
            "sample_data": [
                [datetime.date.today(), "Emergency Fund", 10000, 3500, "=D2/C2"]
            ]
        },
        "Stock Tracker": {
            "headers": ["Symbol", "Company", "Shares", "Avg Price", "Current Price", "Total Value", "Gain/Loss", "% Change"],
            "color": "D8BFD8",  # Purple
            "sample_data": [
                ["AAPL", "Apple Inc.", 10, 150, 175, "=C2*E2", "=F2-(C2*D2)", "=(E2-D2)/D2"]
            ]
        },
        "Weight Tracker": {
            "headers": ["Date", "Weight (kg)", "Body Fat %", "Notes"],
            "color": "FFDAB9",  # Orange
            "sample_data": [
                [datetime.date.today() - datetime.timedelta(days=7), 75.5, 22.0, "Started new diet"],
                [datetime.date.today(), 74.8, 21.5, "Feeling good!"]
            ]
        },
        "Habit Tracker": {
            "headers": ["Date", "Exercise", "Water (glasses)", "Sleep (hours)", "Meditation", "Reading", "Notes"],
            "color": "AFEEEE",  # Teal
            "sample_data": [
                [datetime.date.today() - datetime.timedelta(days=1), "✓", 8, 7.5, "✓", "30 min", "Felt great"],
                [datetime.date.today(), "✓", 6, 8.0, "✓", "15 min", "Tired"]
            ]
        },
        "Cleaning Checklist": {
            "headers": ["Task", "Frequency", "Last Done", "Next Due", "Notes"],
            "color": "E6E6FA",  # Lavender
            "sample_data": [
                ["Vacuum", "Weekly", datetime.date.today() - datetime.timedelta(days=2), datetime.date.today() + datetime.timedelta(days=5), "Living room and bedrooms"],
                ["Laundry", "Twice a week", datetime.date.today() - datetime.timedelta(days=1), datetime.date.today() + datetime.timedelta(days=2), "Whites and colors"]
            ]
        },
        "Meal Planner": {
            "headers": ["Day", "Breakfast", "Lunch", "Dinner", "Snacks", "Grocery Items"],
            "color": "FFDAB9",  # Peach (same as orange)
            "sample_data": [
                ["Monday", "Oatmeal", "Salad", "Grilled chicken", "Fruits", "Chicken, salad mix"],
                ["Tuesday", "Smoothie", "Sandwich", "Pasta", "Nuts", "Bread, pasta, nuts"]
            ]
        },
        "Debt Tracker": {
            "headers": ["Person/Company", "Type", "Amount Owed", "Amount Owe Me", "Due Date", "Status", "Priority", "Notes"],
            "color": "FF6B6B",  # Red
            "sample_data": []
        },
        "Time Table": {
            "headers": ["Time", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
            "color": "98FF98",  # Mint
            "sample_data": [
                ["9:00 AM", "Work", "Work", "Work", "Work", "Work", "Sleep in", "Brunch"],
                ["12:00 PM", "Lunch", "Lunch", "Lunch meeting", "Lunch", "Lunch", "Grocery shopping", "Relax"]
            ]
        }
    }

    # Section mapping for backward compatibility
    section_mapping = {
        'income': 'Income Tracker',
        'expenses': 'Expense Tracker',
        'monthly_purchases': 'Monthly Purchases',
        'savings': 'Savings Tracker',
        'stocks': 'Stock Tracker',
        'weight': 'Weight Tracker',
        'habits': 'Habit Tracker',
        'cleaning': 'Cleaning Checklist',
        'meals': 'Meal Planner',
        'timetable': 'Time Table'
    }
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    while len(wb.sheetnames) > 0:
        wb.remove(wb[wb.sheetnames[0]])

    # Color scheme
    colors = {
        'pink': 'FFB6C1',
        'purple': 'D8BFD8',
        'blue': 'B0E0E6',
        'green': '98FB98',
        'yellow': 'FFFACD',
        'orange': 'FFDAB9',
        'teal': 'AFEEEE',
        'lavender': 'E6E6FA',
        'peach': 'FFDAB9',
        'mint': '98FF98'
    }

    # Helper function to create headers
    def create_header(sheet, title, color):
        sheet.merge_cells('A1:J1')
        sheet['A1'] = title
        sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet['A1'].alignment = Alignment(horizontal="center")
        sheet.row_dimensions[1].height = 30
        
        # Add some spacing
        sheet.row_dimensions[2].height = 10

    # Determine which sections to include
    if sections is None:
        sections = list(section_mapping.keys())
    
    # Create sheets in the desired order
    sheet_order = []
    
    # Always create Welcome Guide first
    welcome = wb.create_sheet("Welcome Guide")
    create_header(welcome, "🌟 Welcome to Your Life & Budget Dashboard 2025 🌟", '6f42c1')
    sheet_order.append('Welcome Guide')
    
    # Create Dashboard
    dashboard = wb.create_sheet('Dashboard')
    create_dashboard(dashboard, month)
    sheet_order.append('Dashboard')
    
    create_excel_template_simple(month=month, sections=sections)
    
    welcome_text = [
        "Thank you for choosing the Life & Budget Dashboard! This comprehensive tool will help you manage:",
        "• Personal Finances (Income, Expenses, Savings, Investments)",
        "• Health & Wellness (Weight, Self-Care, Habits)",
        "• Daily Life (Meal Planning, Cleaning, Schedules)",
        "",
        "📊 HOW TO USE:",
        "1. Start by entering your monthly budget in the 'Dashboard' tab",
        "2. Track your income and expenses in their respective tabs",
        "3. Monitor your investments in the 'Stock Tracker'",
        "4. Use the health and lifestyle trackers to maintain balance",
        "5. Check the 'Dashboard' for insights and progress",
        "",
        "💡 TIP: Use the 'AI Insights' button to get personalized recommendations!",
        "",
        "🔒 Your data stays on your device. For cloud sync, save this file to your preferred cloud storage."
    ]
    
    for i, line in enumerate(welcome_text, start=3):
        cell = welcome.cell(row=i, column=1, value=line)
        if line.startswith("•"):
            cell.font = Font(bold=True)
        elif ":" in line:
            cell.font = Font(bold=True, color='6f42c1')
    
    # Create dashboard sheet - first remove any existing Dashboard sheet
    if 'Dashboard' in [sheet.title for sheet in wb.worksheets]:
        wb.remove(wb['Dashboard'])
    dashboard = wb.create_sheet("Dashboard")
    create_header(dashboard, f"📊 {month} - Financial Overview", '6f42c1')
    
    # Add sample data and charts
    dashboard['A3'] = "Financial Overview"
    dashboard['A3'].font = Font(bold=True, size=12, color='6f42c1')
    
    # Calculate monthly values based on the selected month
    month_index = months.index(month)
    monthly_income = 5000 + (month_index * 100)  # Sample data that increases slightly each month
    monthly_expenses = 3200 + (month_index * 50)  # Sample data that increases slightly each month
    monthly_savings = monthly_income - monthly_expenses
    
    # Calculate weekly breakdown
    weekly_income = [monthly_income * 0.25] * 4
    weekly_expenses = [
        monthly_expenses * 0.2,  # Week 1
        monthly_expenses * 0.3,  # Week 2
        monthly_expenses * 0.25, # Week 3
        monthly_expenses * 0.25  # Week 4
    ]
    weekly_savings = [i - e for i, e in zip(weekly_income, weekly_expenses)]
    
    # Add sample data with monthly focus
    dashboard_data = [
        ["Category", "Planned", "Actual", "Difference", "% of Budget"],
        ["Income", monthly_income, monthly_income * 1.04, "=C4-B4", "=C4/SUM(B4:B8)"],
        ["Expenses", monthly_expenses * 1.1, monthly_expenses, "=C5-B5", "=C5/SUM(B4:B8)"],
        ["Savings", monthly_savings * 1.1, monthly_savings, "=C6-B6", "=C6/SUM(B4:B8)"],
        ["Investments", 500, 550, "=C7-B7", "=C7/SUM(B4:B8)"],
        ["Total", "=SUM(B4:B7)", "=SUM(C4:C7)", "=C8-B8", "=C8/SUM(B4:B8)"]
    ]
    
    for r, row in enumerate(dashboard_data, start=3):
        for c, value in enumerate(row, start=1):
            cell = dashboard.cell(row=r, column=c, value=value)
            if r == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F2E6FF', end_color='F2E6FF', fill_type="solid")
    
    # Create or clear the Charts sheet
    if 'Charts' in wb.sheetnames:
        wb.remove(wb['Charts'])
    charts_sheet = wb.create_sheet('Charts')
    
    # Set up the Charts sheet
    charts_sheet.title = 'Charts'
    charts_sheet.sheet_view.showGridLines = False  # Hide gridlines for cleaner look
    
    # Title and styling
    charts_sheet.merge_cells('A1:J1')
    charts_sheet['A1'] = f"{month} - Financial Overview"
    charts_sheet['A1'].font = Font(size=18, bold=True, color='FFFFFF', name='Calibri')
    charts_sheet['A1'].fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type="solid")
    charts_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    charts_sheet.row_dimensions[1].height = 35
    
    # Add a subtitle
    charts_sheet.merge_cells('A2:J2')
    charts_sheet['A2'] = "Interactive Financial Visualizations"
    charts_sheet['A2'].font = Font(size=14, color='6f42c1', name='Calibri')
    charts_sheet['A2'].alignment = Alignment(horizontal="center")
    charts_sheet.row_dimensions[2].height = 25
    
    # 1. Weekly Financial Trend for the Month (Top-left)
    line_chart = LineChart()
    line_chart.title = f"{month} - Weekly Financial Trend"
    line_chart.style = 12
    line_chart.y_axis.title = 'Amount ($)'
    line_chart.x_axis.title = 'Week'
    line_chart.height = 15
    line_chart.width = 25
    
    # Add some padding for better visualization
    for row in range(5, 10):
        charts_sheet.row_dimensions[row].height = 15
    
    # Write weekly data to a dedicated section
    weekly_data_row = 50
    dashboard.cell(row=weekly_data_row, column=1, value="Week")
    dashboard.cell(row=weekly_data_row, column=2, value="Income")
    dashboard.cell(row=weekly_data_row, column=3, value="Expenses")
    dashboard.cell(row=weekly_data_row, column=4, value="Savings")
    
    for i, week in enumerate(weeks, 1):
        row = weekly_data_row + i
        dashboard.cell(row=row, column=1, value=week)
        dashboard.cell(row=row, column=2, value=weekly_income[i-1])
        dashboard.cell(row=row, column=3, value=weekly_expenses[i-1])
        dashboard.cell(row=row, column=4, value=weekly_savings[i-1])
    
    # Add data to the chart with proper series names
    data = Reference(dashboard, min_col=2, min_row=weekly_data_row-1, max_row=weekly_data_row+len(weeks)-1)
    line_chart.add_data(data, titles_from_data=True)
    
    # Add expenses data
    data = Reference(dashboard, min_col=3, min_row=weekly_data_row-1, max_row=weekly_data_row+len(weeks)-1)
    line_chart.add_data(data, titles_from_data=True)
    
    # Add savings data
    data = Reference(dashboard, min_col=4, min_row=weekly_data_row-1, max_row=weekly_data_row+len(weeks)-1)
    line_chart.add_data(data, titles_from_data=True)
    
    # Set series names explicitly using the correct openpyxl syntax
    from openpyxl.chart.series import SeriesLabel
    line_chart.series[0].title = SeriesLabel(v='Income')
    line_chart.series[1].title = SeriesLabel(v='Expenses')
    line_chart.series[2].title = SeriesLabel(v='Savings')
    
    # Set categories (x-axis)
    cats = Reference(dashboard, min_col=1, min_row=weekly_data_row+1, max_row=weekly_data_row+len(weeks))
    line_chart.set_categories(cats)
    
    # Customize the series with better colors
    line_chart.series[0].graphicalProperties.line.solidFill = "4CAF50"  # Green for income
    line_chart.series[1].graphicalProperties.line.solidFill = "F44336"  # Red for expenses
    line_chart.series[2].graphicalProperties.line.solidFill = "2196F3"  # Blue for savings
    
    # Add data labels to all points
    for series in line_chart.series:
        series.dLbls = openpyxl.chart.label.DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showSeriesName = True
    
    charts_sheet.add_chart(line_chart, "A3")
    
    # 2. Planned vs Actual Bar Chart (Top-right)
    chart1 = BarChart()
    data = Reference(dashboard, min_col=2, max_col=3, min_row=3, max_row=6)  # Exclude Total row
    cats = Reference(dashboard, min_col=1, min_row=4, max_row=6)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.title = "Planned vs Actual"
    chart1.style = 12
    chart1.y_axis.title = 'Amount ($)'
    chart1.x_axis.title = 'Category'
    chart1.height = 15
    chart1.width = 25
    
    # Customize colors
    chart1.series[0].graphicalProperties.solidFill = "4CAF50"  # Green for Planned
    chart1.series[1].graphicalProperties.solidFill = "2196F3"  # Blue for Actual
    
    # Add data labels
    chart1.dLbls = openpyxl.chart.label.DataLabelList()
    chart1.dLbls.showVal = True
    
    charts_sheet.add_chart(chart1, "L3")
    
    # Add sample expense breakdown for visualization with better formatting
    expense_data = [
        ["Category", "Amount", "Target", "Variance", "% of Total"],
        ["Housing", 1200, 1100, "=B2-C2", "=B2/SUM($B$2:$B$7)"],
        ["Food", 800, 700, "=B3-C3", "=B3/SUM($B$2:$B$7)"],
        ["Transportation", 400, 450, "=B4-C4", "=B4/SUM($B$2:$B$7)"],
        ["Utilities", 300, 350, "=B5-C5", "=B5/SUM($B$2:$B$7)"],
        ["Entertainment", 300, 250, "=B6-C6", "=B6/SUM($B$2:$B$7)"],
        ["Others", 200, 200, "=B7-C7", "=B7/SUM($B$2:$B$7)"],
        ["Total", "=SUM(B2:B7)", "=SUM(C2:C7)", "=SUM(D2:D7)", "=SUM(E2:E7)"]
    ]
    
    # Write expense breakdown data with formatting
    for r, row in enumerate(expense_data, start=25):
        for c, value in enumerate(row, start=1):
            cell = dashboard.cell(row=r, column=c, value=value)
            if r == 25:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
                cell.font = Font(color='FFFFFF')
            elif r == len(expense_data) + 24:  # Total row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F2E6FF', end_color='F2E6FF', fill_type='solid')
    
    # 3. Expense Breakdown Pie Chart (Bottom-left) - 3D with percentage and legend
    expense_pie = PieChart()
    data = Reference(dashboard, min_col=2, min_row=26, max_row=31)
    labels = Reference(dashboard, min_col=1, min_row=26, max_row=31)
    expense_pie.add_data(data, titles_from_data=False)
    expense_pie.set_categories(labels)
    expense_pie.title = f"{month} - Expenses Breakdown"
    expense_pie.height = 15
    expense_pie.width = 25
    
    # Add percentage data labels
    expense_pie.dataLabels = openpyxl.chart.label.DataLabelList()
    expense_pie.dataLabels.showPercent = True
    expense_pie.dataLabels.showVal = False
    expense_pie.dataLabels.showCatName = True
    expense_pie.legend.position = 'b'  # Bottom legend
    
    # Add some 3D rotation for better visibility
    expense_pie.rotX = 30
    expense_pie.rotY = 30
    
    charts_sheet.add_chart(expense_pie, "A25")
    
    # 4. Budget Distribution Donut Chart (Bottom-right)
    pie = PieChart()
    data = Reference(dashboard, min_col=2, min_row=4, max_row=6)
    labels = Reference(dashboard, min_col=1, min_row=4, max_row=6)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = f"{month} - Budget Distribution"
    pie.height = 15
    pie.width = 25
    
    # Make it a donut chart by setting a hole size
    pie.holeSize = 30  # Percentage of hole size
    
    # Add percentage data labels
    pie.dataLabels = openpyxl.chart.label.DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = True
    pie.legend.position = 'r'  # Right legend
    
    # Custom colors
    colors = ['4CAF50', '2196F3', 'FFC107']  # Green, Blue, Amber
    for i, series in enumerate(pie.series):
        series.graphicalProperties.solidFill = colors[i % len(colors)]
    
    # Add 3D rotation
    pie.rotX = 15
    pie.rotY = 30
    
    charts_sheet.add_chart(pie, "L25")
    
    # Weekly breakdown data is already calculated at the beginning of the function
    
    # Sample data for monthly trends
    income_data = [5000 + (i * 100) for i in range(12)]  # Increasing income
    expense_data = [3200 + (i * 50) for i in range(12)]  # Increasing expenses
    savings_data = [i - e for i, e in zip(income_data, expense_data)]  # Calculate savings
    
    # Write time series data with better formatting
    dashboard['A45'] = "Monthly Financial Trend"
    dashboard['A45'].font = Font(bold=True, size=12, color='6f42c1', name='Calibri')
    
    headers = ["Month", "Income", "Expenses", "Savings"]
    for col, header in enumerate(headers, 1):
        cell = dashboard.cell(row=46, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
    
    for i, (month, inc, exp, sav) in enumerate(zip(months, income_data, expense_data, savings_data), start=1):
        row = 46 + i
        dashboard.cell(row=row, column=1, value=month)
        
        # Format as currency
        for col, val in zip([2, 3, 4], [inc, exp, sav]):
            cell = dashboard.cell(row=row, column=col, value=val)
            cell.number_format = '$#,##0'
            
            # Add conditional formatting for negative savings
            if col == 4 and val < 0:
                cell.font = Font(color='FF0000')  # Red for negative savings
    
    # Auto-size columns for better visibility
    for column in dashboard.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        dashboard.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    # Define the exact order we want the sheets to appear in
    sheet_order = [
        'Welcome Guide',
        'Income Tracker',
        'Expense Tracker',
        'Savings Tracker',
        'Stock Tracker',
        'Weight Tracker',
        'Habit Tracker',
        'Cleaning Checklist',
        'Meal Planner',
        'Time Table',
        'AI Insights'
    ]
    
    # Map UI section names to tab names
    section_map = {
        'income': "Income Tracker",
        'expenses': "Expense Tracker",
        'savings': "Savings Tracker",
        'stocks': "Stock Tracker",
        'weight': "Weight Tracker",
        'habits': "Habit Tracker",
        'cleaning': "Cleaning Checklist",
        'meals': "Meal Planner",
        'timetable': "Time Table"
    }
    
    # Initialize tabs with core sheets
    tabs = {
        "Welcome Guide": all_tabs.get("Welcome Guide", {}),
        "AI Insights": all_tabs.get("AI Insights", {})
    }
    
    # If no sections specified, include all
    if not sections:
        sections = section_map.keys()
    
    # Add selected tabs to the tabs dictionary
    for section in sections:
        if section in section_map:
            tab_name = section_map[section]
            if tab_name in all_tabs:
                tabs[tab_name] = all_tabs[tab_name]
    
    # Process each tab in the desired order
    for tab_name in sheet_order:
        # Handle core sheets
        if tab_name in ['Welcome Guide', 'AI Insights']:
            tab_data = all_tabs.get(tab_name, {})
            if tab_name not in wb.sheetnames:
                sheet = wb.create_sheet(tab_name)
            else:
                sheet = wb[tab_name]
                sheet.delete_rows(1, sheet.max_row)
            
            if tab_name == 'Welcome Guide':
                create_welcome_guide(sheet)
            elif tab_name == 'AI Insights':
                create_header(sheet, "🤖 AI-Powered Financial Insights", '6f42c1')
                ai_text = [
                    "This section provides AI-generated insights based on your data:",
                    "",
                    "🔍 To get started:",
                    "1. Fill in your financial data in the respective sheets",
                    "2. Click the 'Generate AI Insights' button in the app",
                    "3. Review personalized recommendations here",
                    "",
                    "The AI will analyze your spending patterns, saving habits, and overall financial health to provide actionable advice.",
                    "",
                    "💡 Tip: The more data you provide, the more accurate the insights will be!"
                ]
                for i, line in enumerate(ai_text, start=3):
                    cell = sheet.cell(row=i, column=1, value=line)
                    if line.startswith("🔍") or line.startswith("💡"):
                        cell.font = Font(bold=True, color='6f42c1')
            continue
            
        # Skip Charts sheet for now (handled separately)
        if tab_name == 'Charts':
            continue
            
        # Skip if not in selected tabs
        if tab_name not in tabs:
            continue
            
        # Remove existing sheet if it exists (except for core sheets)
        if tab_name in wb.sheetnames and tab_name not in ['Welcome Guide', 'AI Insights']:
            try:
                wb.remove(wb[tab_name])
            except Exception as e:
                print(f"Warning: Could not remove sheet {tab_name}: {str(e)}")
            
        # Get the tab data (skip for AI Insights)
        tab_data = all_tabs.get(tab_name, {})
        
        # Create the sheet if it doesn't exist
        if tab_name not in wb.sheetnames:
            sheet = wb.create_sheet(tab_name)
        else:
            sheet = wb[tab_name]
            # Clear existing content but keep formatting
            sheet.delete_rows(1, sheet.max_row)
        
        # Handle Welcome Guide specially
        if tab_name == 'Welcome Guide':
            create_welcome_guide(sheet)
            continue
            
        # Handle AI Insights specially
        if tab_name == 'AI Insights':
            create_header(sheet, "🤖 AI-Powered Financial Insights", '6f42c1')
            ai_text = [
                "This section provides AI-generated insights based on your data:",
                "",
                "🔍 To get started:",
                "1. Fill in your financial data in the respective sheets",
                "2. Click the 'Generate AI Insights' button in the app",
                "3. Review personalized recommendations here",
                "",
                "The AI will analyze your spending patterns, saving habits, and overall financial health to provide actionable advice.",
                "",
                "💡 Tip: The more data you provide, the more accurate the insights will be!"
            ]
            
            for i, line in enumerate(ai_text, start=3):
                cell = sheet.cell(row=i, column=1, value=line)
                if line.startswith("🔍") or line.startswith("💡"):
                    cell.font = Font(bold=True, color='6f42c1')
            continue
            
        # For all other sheets, add the standard content
        title = f"{month} {tab_name}" if tab_name == 'Dashboard' else tab_name
        create_header(sheet, title, tab_data.get("color", "6f42c1"))
        
        # Add column headers
        headers = tab_data.get("headers", [])
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(
                start_color=tab_data.get("color", "6f42c1"), 
                end_color=tab_data.get("color", "6f42c1"), 
                fill_type='solid'
            )
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
        
        # Add sample data if available
        if "sample_data" in tab_data and tab_data["sample_data"]:
            data = tab_data["sample_data"]
            for row_num, row_data in enumerate(data, start=4):
                for col_num, cell_value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                    if isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
                        cell.number_format = '0.00'  # Removed currency symbol
            
            # Add total row if there are numeric columns
            if data and len(data[0]) > 1:  # Only if there are columns to sum
                total_row = len(data) + 4
                sheet.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
                
                # Calculate and add totals for numeric columns
                for col in range(2, len(headers) + 1):
                    if any(isinstance(sheet.cell(row=r, column=col).value, (int, float)) 
                          for r in range(4, total_row)):
                        # Calculate sum of the column
                        total = sum(sheet.cell(row=r, column=col).value or 0 
                                  for r in range(4, total_row) 
                                  if isinstance(sheet.cell(row=r, column=col).value, (int, float)))
                        cell = sheet.cell(row=total_row, column=col, value=total)
                        cell.font = Font(bold=True)
                        cell.number_format = '0.00'  # Format total without currency
                        
                        # Create a simple bar chart for the column
                        try:
                            if col > 1:  # Skip the first column (labels)
                                chart = BarChart()
                                chart.title = f"{tab_name} - {headers[col-1]}"
                                chart.style = 10
                                chart.y_axis.title = 'Amount'
                                chart.x_axis.title = 'Items'
                                
                                # Add data to chart
                                data_ref = Reference(sheet, min_col=col, min_row=3, max_row=total_row-1)
                                cats = Reference(sheet, min_col=1, min_row=4, max_row=total_row-1)
                                chart.add_data(data_ref, titles_from_data=True)
                                chart.set_categories(cats)
                                
                                # Position the chart
                                chart_cell = get_column_letter(col * 2) + str(total_row + 2)
                                sheet.add_chart(chart, chart_cell)
                        except Exception as e:
                            print(f"Error creating chart for {tab_name} column {col}: {str(e)}")
        
        # Auto-size columns for the current sheet
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    # Create Charts sheet if expenses or monthly_purchases are included
    if 'expenses' in sections or 'monthly_purchases' in sections:
        create_enhanced_charts(wb, month)
    
    # Create AI Insights sheet only if it doesn't exist
    if 'AI Insights' not in wb.sheetnames:
        ai_sheet = wb.create_sheet('AI Insights')
        create_ai_insights_placeholder(ai_sheet)
    
    # Reorder sheets to match TempJan2025.xlsx order
    desired_order = [
        'Welcome Guide',
        'Dashboard',
        'Charts',
        'Income Tracker',
        'Expense Tracker',
        'Monthly Purchases',
        'Savings Tracker',
        'Weight Tracker',
        'AI Insights'
    ]
    
    # Move sheets to desired order (in reverse to get correct positioning)
    for sheet_name in reversed(desired_order):
        if sheet_name in wb.sheetnames:
            wb.move_sheet(wb[sheet_name], 0)
    
    # Set the active sheet to Dashboard or Welcome Guide
    if 'Dashboard' in wb.sheetnames:
        wb.active = wb['Dashboard']
    elif 'Welcome Guide' in wb.sheetnames:
        wb.active = wb['Welcome Guide']
    
    # Ensure all sheets are visible
    for sheet in wb.worksheets:
        sheet.sheet_state = 'visible'
    
    return wb

def read_excel_data_optimized(file_path, selected_categories=None, for_ai_conversion=False):
    """Optimized function to read Excel data using openpyxl for better accuracy.
    
    Args:
        file_path (str): Path to the Excel file
        selected_categories (list, optional): Categories to include. Defaults to None (all).
        for_ai_conversion (bool): If True, prepares data for AI-friendly conversion.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        data = {}

        # Map categories to sheet names
        category_map = {
            'Income': 'Income Tracker',
            'Expenses': 'Expense Tracker',
            'Savings': 'Savings Tracker',
            'Investments': 'Stock Tracker',
            'Health': ['Weight Tracker', 'Habit Tracker'],
            'Lifestyle': ['Cleaning Checklist', 'Meal Planner', 'Time Table']
        }

        # Determine which sheets to process
        if selected_categories:
            sheets_to_process = []
            for category in selected_categories:
                if category in category_map:
                    sheets = category_map[category]
                    if isinstance(sheets, list):
                        sheets_to_process.extend(sheets)
                    else:
                        sheets_to_process.append(sheets)
        else:
            # Exclude non-data sheets
            excluded_sheets = ['Welcome Guide', 'Dashboard', 'Charts', 'AI Insights', 'Summary']
            sheets_to_process = [sheet.title for sheet in wb.worksheets 
                               if sheet.title not in excluded_sheets]

        # Process each sheet
        for sheet_name in sheets_to_process:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                df = worksheet_to_dataframe(sheet)
                
                if not df.empty:
                    # Clean the data
                    df = df.dropna(how='all')
                    df = df.dropna(axis=1, how='all')
                    
                    if not df.empty:
                        if for_ai_conversion:
                            # For AI conversion, we want to keep more structured data
                            data[sheet_name] = df
                        else:
                            # For display purposes, convert to string
                            data[sheet_name] = df.head(20).to_string()

        return data
    except Exception as e:
        return {"error": f"Error reading Excel file: {str(e)}"}

def create_ai_friendly_template(output_path="ai_finance_template.xlsx"):
    """Creates an AI-friendly Excel template with optimized structure for financial tracking."""
    from openpyxl.styles import Font, PatternFill
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("AI_Instructions")
    ws_instructions['A1'] = "AI-Friendly Data Formatting Instructions"
    ws_instructions['A1'].font = Font(bold=True, size=14)
    instructions = [
        "1. Keep one data type per column",
        "2. Use consistent date formats (YYYY-MM-DD recommended)",
        "3. Avoid merged cells",
        "4. Use headers in the first row",
        "5. No empty rows or columns within the data",
        "6. Use consistent category names",
        "7. Include all relevant metadata in separate columns",
        "8. Use numerical values without currency symbols"
    ]
    
    for i, instruction in enumerate(instructions, start=3):
        ws_instructions[f'A{i}'] = instruction
    
    # Add sample data sheets for different data types
    def add_sample_sheet(name, headers, data):
        ws = wb.create_sheet(name)
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 30)  # Cap at 30 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Transactions sample
    transactions_headers = ["Date", "Type", "Category", "Description", "Amount", "Account", "Tags"]
    transactions_data = [
        ["2023-11-01", "Expense", "Food", "Grocery Shopping", 150.00, "Checking", "groceries"],
        ["2023-11-02", "Income", "Salary", "Monthly Salary", 3000.00, "Savings", "salary"],
        ["2023-11-03", "Expense", "Transport", "Bus Fare", 50.00, "Credit Card", "commute"]
    ]
    add_sample_sheet("Transactions", transactions_headers, transactions_data)
    
    # Budget sample
    budget_headers = ["Category", "Budgeted", "Spent", "Remaining", "Period"]
    budget_data = [
        ["Groceries", 600.00, 150.00, 450.00, "Monthly"],
        ["Transportation", 200.00, 50.00, 150.00, "Monthly"],
        ["Entertainment", 100.00, 0.00, 100.00, "Monthly"]
    ]
    add_sample_sheet("Budget", budget_headers, budget_data)
    
    # Save the workbook
    wb.save(output_path)
    return output_path

def convert_to_ai_friendly(input_file, output_file):
    """Convert an existing Excel file to AI-friendly format.
    
    Args:
        input_file (str): Path to the input Excel file
        output_file (str): Path to save the converted file
        
    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        # Read the input file
        wb_in = openpyxl.load_workbook(input_file, data_only=True)
        wb_out = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb_out.sheetnames:
            wb_out.remove(wb_out['Sheet'])
        
        # Add instructions sheet
        ws_instructions = wb_out.create_sheet("AI_Instructions")
        ws_instructions['A1'] = "AI-Friendly Data Format"
        ws_instructions['A1'].font = Font(bold=True, size=14)
        
        instructions = [
            "This file has been converted to be more AI-friendly. Here's what was done:",
            "1. Removed formatting and merged cells",
            "2. Ensured consistent data types in columns",
            "3. Standardized date formats",
            "4. Added clear headers",
            "5. Removed empty rows and columns"
        ]
        
        for i, instruction in enumerate(instructions, start=3):
            ws_instructions[f'A{i}'] = instruction
        
        # Process each sheet
        for sheet_name in wb_in.sheetnames:
            # Skip non-data sheets
            if sheet_name in ['Welcome Guide', 'Charts', 'AI Insights', 'Dashboard', '_metadata']:
                continue
                
            ws_in = wb_in[sheet_name]
            df = worksheet_to_dataframe(ws_in)
            
            if not df.empty:
                # Clean the data
                df = df.dropna(how='all').dropna(axis=1, how='all')
                
                if not df.empty:
                    # Standardize column names
                    df.columns = [str(col).strip() for col in df.columns]
                    
                    # Standardize date formats if present
                    date_columns = [col for col in df.columns if 'date' in col.lower() or 'Date' in str(col)]
                    for col in date_columns:
                        try:
                            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
                        except:
                            pass
                    
                    # Create new sheet in output workbook
                    safe_sheet_name = re.sub(r'[\[\]\:*?/\\]', '_', sheet_name)[:31]  # Excel sheet name limit
                    ws_out = wb_out.create_sheet(safe_sheet_name)
                    
                    # Write data to new sheet
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws_out.append(r)
                    
                    # Format headers
                    for cell in ws_out[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    
                    # Auto-adjust column widths
                    for column in ws_out.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min((max_length + 2), 30)  # Cap at 30 characters
                        ws_out.column_dimensions[column_letter].width = adjusted_width
        
        # Add metadata sheet
        ws_meta = wb_out.create_sheet("_metadata")
        ws_meta['A1'] = "File Information"
        ws_meta['A1'].font = Font(bold=True, size=12)
        
        # Add version and schema information
        ws_meta['A2'] = "Data Schema Version"
        ws_meta['B2'] = "1.1"  # Bump version for Monthly Purchases addition
        ws_meta['A2'].font = Font(bold=True)
        
        meta_data = [
            ("Original file:", os.path.basename(input_file)),
            ("Converted on:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Conversion notes:", "This file has been optimized for AI analysis."),
            ("", ""),
            ("Monthly Purchases Schema:", ""),
            ("- Date:", "Date of the purchase"),
            ("- Item:", "Name of the item or service"),
            ("- Amount:", "Cost of the purchase"),
            ("- Type:", "Subscription or One-Time"),
            ("- Category:", "Category of the purchase"),
            ("- Notes:", "Additional details about the purchase")
        ]
        
        for i, (label, value) in enumerate(meta_data, start=3):
            ws_meta[f'A{i}'] = label
            ws_meta[f'B{i}'] = value
            ws_meta[f'A{i}'].font = Font(bold=True)
        
        # Save the output file
        wb_out.save(output_file)
        return True, ""
        
    except Exception as e:
        return False, f"Error during conversion: {str(e)}"

def generate_ai_insights(file_path, selected_categories=None):
    """
    Generate AI insights from the uploaded Excel file using Ollama.
    
    Args:
        file_path (str): Path to the Excel file
        selected_categories (list): List of categories to analyze
        
    Returns:
        tuple: (excel_data_str, insights) - The data string and AI-generated insights
    """
    try:
        # Read the Excel file
        excel_data = read_excel_data_optimized(file_path, selected_categories, for_ai_conversion=True)
        
        if not excel_data or "error" in excel_data:
            error_msg = excel_data.get("error", "No data found in the Excel file")
            return "", f"❌ Error processing Excel file: {error_msg}"
        
        # Convert data to a string format for the AI
        excel_data_str = ""
        for sheet_name, df in excel_data.items():
            excel_data_str += f"\n--- {sheet_name} ---\n"
            excel_data_str += df.to_string() + "\n\n"
        
        # Prepare an enhanced prompt for Ollama
        prompt = f"""# Financial and Lifestyle Analysis Report

## Data Overview
Analyze the following financial and personal data to provide comprehensive insights. The data includes:
- Income and expense tracking
- Savings and investment information
- Health and lifestyle metrics
- Debt tracking and obligations

## Analysis Instructions
For each relevant data section, provide:
1. **Key Observations**: 2-3 bullet points highlighting the most important findings
2. **Trends**: Any noticeable patterns or changes over time
3. **Strengths**: What's working well
4. **Areas for Improvement**: Specific, actionable recommendations
5. **Quick Wins**: Easy changes that could have immediate positive impact

## Data to Analyze:
{excel_data_str}

## Detailed Analysis Request

### 1. Financial Health Assessment
- Calculate and analyze the savings rate (savings/income)
- Identify top 3 spending categories by amount
- Compare fixed vs. variable expenses
- Evaluate emergency fund status (if data available)

### 2. Debt Analysis (if debt data available)
- Analyze total debt burden and net debt position
- Identify high-priority debts that need immediate attention
- Suggest debt repayment strategies (avalanche vs snowball method)
- Highlight any overdue or upcoming due dates
- Analyze debt-to-income ratio implications

### 3. Spending Analysis
- Identify any unusual or outlier transactions
- Highlight any recurring subscriptions or expenses that could be reduced
- Compare spending against common budgeting guidelines (e.g., 50/30/20 rule)

### 4. Income & Budget Optimization
- Analyze income stability and sources
- Suggest potential areas for expense reduction
- Recommend budget allocation improvements considering debt obligations

### 5. Savings & Investments
- Evaluate current savings rate
- Assess investment diversification (if data available)
- Suggest potential savings goals based on income/expense patterns
- Recommend how to balance debt repayment with savings goals

### 6. Lifestyle & Health (if data available)
- Analyze any health metrics for trends
- Correlate lifestyle choices with financial patterns
- Suggest holistic improvements that benefit both health and finances

## Required Output Format

# Financial & Lifestyle Insights Report

### 📊 Executive Summary
[2-3 sentence overview of the most important findings]

### 💰 Financial Health Score
[Score from 1-10 with brief explanation]

### 📈 Key Metrics
- **Savings Rate**: [X]% (Goal: 20%+)
- **Top Spending Category**: [Category] ([X]% of expenses)
- [Other relevant metrics]

### 🔍 Detailed Analysis
[Organized by category with clear headings and bullet points]

### 🎯 Actionable Recommendations
1. [Specific, actionable item with potential impact]
2. [Specific, actionable item with potential impact]
3. [Specific, actionable item with potential impact]

### 🚀 Quick Wins
- [Quick action with minimal effort]
- [Quick action with minimal effort]

### 📅 30-Day Action Plan
1. Week 1: [Specific task]
2. Week 2: [Specific task]
3. Week 3-4: [Specific tasks]

Note: Be specific with numbers and percentages where possible. Use emojis for better readability. Keep the tone positive and encouraging while being direct about areas needing improvement."""
        
        try:
            # First, check if Ollama is running
            try:
                print("Attempting to list Ollama models...")  # Debug log
                
                # List models using Ollama
                models_response = ollama.list()
                print(f"Ollama response type: {type(models_response)}")  # Debug log
                
                # Initialize available_models list
                available_models = []
                
                # Handle the case where models is a direct attribute
                if hasattr(models_response, 'models'):
                    print("Found models attribute in response")
                    for model in models_response.models:
                        if hasattr(model, 'model'):  # Check for model attribute
                            available_models.append(model.model)
                            print(f"Found model: {model.model}")
                
                # If no models found yet, try to access the models list directly
                if not available_models and isinstance(models_response, list):
                    print("Response is a list of models")
                    for model in models_response:
                        if hasattr(model, 'model'):
                            available_models.append(model.model)
                            print(f"Found model in list: {model.model}")
                
                # If still no models, try to access the models attribute through __dict__
                if not available_models and hasattr(models_response, '__dict__'):
                    models_dict = models_response.__dict__
                    if 'models' in models_dict and models_dict['models']:
                        for model in models_dict['models']:
                            if hasattr(model, 'model'):
                                available_models.append(model.model)
                                print(f"Found model in __dict__: {model.model}")
                
                print(f"Available models: {available_models}")
                
                if not available_models:
                    # As a last resort, try to use the first available model directly
                    try:
                        response = ollama.generate(model='llama2', prompt='Test')
                        if response and hasattr(response, 'response'):
                            available_models = ['llama2']
                            print("Successfully tested llama2 model")
                    except Exception as e:
                        print(f"Could not use llama2 model: {str(e)}")
                
            except Exception as e:
                raise Exception(f"Could not connect to Ollama. Is it running? Error: {str(e)}")
            
            # If we have available models, try to use one
            if available_models:
                # Prefer gemma3:4b if available, then llama2, otherwise use the first available model
                model_to_use = next(
                    (m for m in available_models if 'gemma3:4b' in m.lower()) or 
                    (m for m in available_models if 'llama2' in m.lower()) or 
                    available_models,
                    available_models[0]
                )
                print(f"Using model: {model_to_use}")
                
                try:
                    # Generate insights using the selected model
                    response = ollama.generate(
                        model=model_to_use,
                        prompt=prompt,
                        stream=False
                    )
                    
                    if hasattr(response, 'response'):
                        insights = response.response
                    elif isinstance(response, dict) and 'response' in response:
                        insights = response['response']
                    else:
                        insights = str(response)
                    
                    return excel_data_str, insights
                    
                except Exception as e:
                    raise Exception(f"Error generating insights with model '{model_to_use}': {str(e)}")
            else:
                raise Exception("No suitable AI models found. Please install a model with 'ollama pull <model>'")
            
            try:
                # Try to use the model to generate insights
                response = ollama.generate(
                    model=model_to_use,
                    prompt=prompt,
                    stream=False
                )
                insights = response.get('response', 'No insights generated')
                return excel_data_str, insights
                
            except Exception as e:
                raise Exception(f"Error generating insights with model '{model_to_use}': {str(e)}")
                
        except Exception as e:
            # Format the error message with installation instructions
            error_msg = f"""
            ⚠️ Error generating AI insights: {str(e)}
            
            To enable AI analysis, you need to install an AI model first.
            
            Try running one of these commands in your terminal:
            - `ollama pull llama2`  # Recommended for most users
            - `ollama pull mistral` # Good alternative
            - `ollama pull gemma`   # Lightweight option
            
            After installation, restart this app and try again.
            
            Here's your data for manual review:
            {excel_data_str}
            """
            return excel_data_str, error_msg
            
    except Exception as e:
        return "", f"❌ Error: {str(e)}"

def parse_monzo_statement(file_content):
    """Parse Monzo bank statement from CSV or PDF"""
    try:
        # First try to detect if it's a PDF
        if file_content.startswith(b'%PDF'):
            return parse_monzo_pdf_statement(file_content)
        
        # Try to parse as CSV
        df = pd.read_csv(StringIO(file_content.decode('utf-8')))
        
        # Debug: Show the columns we found
        print(f"Monzo statement columns found: {list(df.columns)}")
        print(f"First few rows:\n{df.head()}")
        
        # Standardize Monzo columns - be more flexible with column names
        # Common Monzo column variations
        column_mapping = {
            'date': 'Date',
            'time': 'Time',
            'type': 'Type',
            'description': 'Description',
            'category': 'Category',
            'amount': 'Amount',
            'balance': 'Balance',
            'transaction_type': 'Transaction Type'
        }
        
        # Try to find the correct columns
        date_col = None
        desc_col = None
        amount_col = None
        category_col = None
        
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if col_lower in ['date', 'transaction date', 'posted date']:
                date_col = col
            elif col_lower in ['description', 'details', 'transaction description', 'memo', 'notes']:
                desc_col = col
            elif col_lower in ['amount', 'value', 'debit', 'credit', 'transaction amount']:
                amount_col = col
            elif col_lower in ['category', 'type', 'transaction type']:
                category_col = col
        
        if date_col and desc_col and amount_col:
            # Rename columns to standard names
            df = df.rename(columns={
                date_col: 'Date',
                desc_col: 'Description',
                amount_col: 'Amount'
            })
            
            # Ensure Amount is numeric
            df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
            
            # Add bank and transaction type
            df['Bank'] = 'Monzo'
            df['Transaction Type'] = df['Amount'].apply(lambda x: 'Income' if x > 0 else 'Expense')
            
            # Handle Monzo categories - if they exist, use them; otherwise categorize by description
            if category_col and category_col in df.columns:
                # Check if Monzo already has pot transfer categories
                df['Category'] = df[category_col].apply(lambda x: 
                    'Pot Transfer' if 'pot' in str(x).lower() else categorize_transaction(x)
                )
                # Drop the original category column to avoid confusion
                df = df.drop(columns=[category_col])
            else:
                df['Category'] = df['Description'].apply(categorize_transaction)
            
            return df
        else:
            print(f"Could not find required columns. Date: {date_col}, Description: {desc_col}, Amount: {amount_col}")
            return None
    except Exception as e:
        print(f"Error parsing Monzo statement: {str(e)}")
        return None

def extract_monzo_balance_summary(text):
    """Extract balance summary information from Monzo PDF header"""
    balance_info = {}
    
    try:
        # Extract Total balance (Including all Pots and Cashback)
        total_balance_match = re.search(r'Total balance\(Including all Pots and Cashback\)£(\d{1,3}(?:,\d{3})*\.\d{2})', text)
        if total_balance_match:
            balance_info['total_balance_including_pots'] = float(total_balance_match.group(1).replace(',', ''))
        
        # Extract Personal Account balance (Excluding all Pots)
        personal_balance_match = re.search(r'Personal Account balance\(Excluding all Pots\)£(\d{1,3}(?:,\d{3})*\.\d{2})', text)
        if personal_balance_match:
            balance_info['personal_account_balance'] = float(personal_balance_match.group(1).replace(',', ''))
        
        # Extract Balance in Pots
        pots_balance_match = re.search(r'Balance in Pots\(This includes both Regular Pots with Monzo and SavingsPots with external providers\)£(\d{1,3}(?:,\d{3})*\.\d{2})', text)
        if pots_balance_match:
            balance_info['balance_in_pots'] = float(pots_balance_match.group(1).replace(',', ''))
        
        # Extract Cashback Balance
        cashback_balance_match = re.search(r'Cashback Balance-£(\d{1,3}(?:,\d{3})*\.\d{2})', text)
        if cashback_balance_match:
            balance_info['cashback_balance'] = float(cashback_balance_match.group(1).replace(',', ''))
        
        # Extract Total outgoings
        outgoings_match = re.search(r'Total outgoings\+£(\d{1,3}(?:,\d{3})*\.\d{2})', text)
        if outgoings_match:
            balance_info['total_outgoings'] = float(outgoings_match.group(1).replace(',', ''))
        
        # Extract Total deposits
        deposits_match = re.search(r'Total deposits£(\d{1,3}(?:,\d{3})*\.\d{2})', text)
        if deposits_match:
            balance_info['total_deposits'] = float(deposits_match.group(1).replace(',', ''))
            
    except Exception as e:
        print(f"Error extracting balance summary: {str(e)}")
    
    return balance_info

def parse_monzo_pdf_statement(file_content):
    """Parse Monzo PDF statement by extracting text with improved parsing for concatenated format"""
    try:
        import PyPDF2
        import io
        import re
        
        # Read PDF
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
        text = ""
        
        # Extract text from all pages
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        
        print(f"Extracted PDF text (first 1000 chars):\n{text[:1000]}")
        
        # Extract balance summary from header
        balance_info = extract_monzo_balance_summary(text)
        
        # Parse Monzo PDF format - handle concatenated transaction format
        all_transactions = []
        
        # Use regex to find all transaction patterns in the concatenated text
        # Pattern: DD/MM/YYYYDescriptionAmountBalanceAmount
        transaction_pattern = r'(\d{2}/\d{2}/\d{4})([A-Za-z0-9\s\.\-\(\)\/]+?)(-?\d+\.\d{2})(-?\d+\.\d{2})'
        
        matches = re.findall(transaction_pattern, text)
        print(f"Found {len(matches)} transaction matches using regex")
        
        for match in matches:
            date_str, description, amount_str, balance_str = match
            
            try:
                amount = float(amount_str)
                balance = float(balance_str)
                
                # Clean up description
                description = description.strip()
                
                # Skip obvious invalid transactions but keep pot transfers
                if (len(description) < 2 or 
                    description.isdigit() or
                    any(keyword in description.lower() for keyword in [
                        'balance', 'total', 'account', 'statement', 'anthony', 'kinyua', 'gathukia',
                        'flat', 'room', 'house', 'court', 'road', 'nottingham', 'ng7', 'united', 'kingdom'
                    ])):
                    continue
                
                # Only consider reasonable transaction amounts
                if abs(amount) < 0.01 or abs(amount) > 10000:
                    continue
                
                all_transactions.append({
                    'Date': datetime.datetime.strptime(date_str, '%d/%m/%Y').strftime('%Y-%m-%d'),
                    'Description': description,
                    'Amount': amount,
                    'Balance': balance,
                    'Bank': 'Monzo',
                    'Transaction Type': 'Income' if amount > 0 else 'Expense',
                    'Category': categorize_transaction(description)
                })
                
            except (ValueError, IndexError) as e:
                print(f"Error parsing transaction: {match}, Error: {e}")
                continue
        
        # If regex didn't work well, try alternative approach
        if len(all_transactions) < 10:
            print("Regex approach found few transactions, trying alternative parsing...")
            
            # Alternative: split by date pattern
            date_pattern = r'(\d{2}/\d{2}/\d{4})'
            parts = re.split(date_pattern, text)
            
            current_date = None
            for i in range(1, len(parts), 2):  # Skip the first empty part
                if i + 1 < len(parts):
                    date_str = parts[i]
                    content = parts[i + 1]
                    
                    try:
                        current_date = datetime.datetime.strptime(date_str, '%d/%m/%Y').strftime('%Y-%m-%d')
                        
                        # Find all amounts in this section
                        amount_matches = re.findall(r'(-?\d+\.\d{2})', content)
                        
                        if len(amount_matches) >= 2:  # Need at least amount and balance
                            amount = float(amount_matches[0])
                            
                            # Extract description (text between date and first amount)
                            description = content[:content.find(amount_matches[0])].strip()
                            
                            # Skip invalid transactions but keep pot transfers
                            if (len(description) >= 2 and 
                                not description.isdigit() and
                                abs(amount) >= 0.01 and abs(amount) <= 10000 and
                                not any(keyword in description.lower() for keyword in [
                                    'balance', 'total', 'account', 'statement'
                                ])):
                                
                                all_transactions.append({
                                    'Date': current_date,
                                    'Description': description,
                                    'Amount': amount,
                                    'Bank': 'Monzo',
                                    'Transaction Type': 'Income' if amount > 0 else 'Expense',
                                    'Category': categorize_transaction(description)
                                })
                    
                    except (ValueError, IndexError):
                        continue
        
        # Filter and keep ALL valid transactions (including duplicates)
        valid_transactions = []
        
        for trans in all_transactions:
            desc = trans['Description'].strip()
            desc_lower = desc.lower()
            
            # Skip obvious invalid transactions but keep pot transfers
            if (len(desc) < 2 or 
                desc.isdigit() or
                any(keyword in desc_lower for keyword in [
                    'balance', 'total', 'account', 'statement', 'anthony', 'kinyua', 'gathukia',
                    'flat', 'room', 'house', 'court', 'road', 'nottingham', 'ng7', 'united', 'kingdom'
                ])):
                continue
            
            # Keep ALL valid transactions, no deduplication
            valid_transactions.append(trans)
        
        if valid_transactions:
            df = pd.DataFrame(valid_transactions)
            print(f"Successfully parsed {len(valid_transactions)} transactions from PDF (from {len(all_transactions)} raw matches)")
            print(f"Date range: {df['Date'].min()} to {df['Date'].max()}")
            
            # Add balance summary information to the dataframe
            if balance_info:
                for key, value in balance_info.items():
                    df.attrs[key] = value
            
            return df
        else:
            print("No transactions found in PDF")
            return None
            
    except ImportError:
        print("PyPDF2 not installed. Install with: pip install PyPDF2")
        return None
    except Exception as e:
        print(f"Error parsing Monzo PDF: {str(e)}")
        return None

def parse_lloyds_statement(file_content):
    """Parse Lloyds bank statement"""
    try:
        # Try CSV format
        df = pd.read_csv(StringIO(file_content.decode('utf-8')))
        
        # Standardize Lloyds columns
        if 'Transaction Date' in df.columns or 'Date' in df.columns:
            date_col = 'Transaction Date' if 'Transaction Date' in df.columns else 'Date'
            df = df.rename(columns={date_col: 'Date'})
            
            if 'Description' not in df.columns and 'Transaction Details' in df.columns:
                df = df.rename(columns={'Transaction Details': 'Description'})
                
            if 'Amount' not in df.columns and 'Debit Amount' in df.columns:
                df['Amount'] = df['Debit Amount'].fillna(0) - df['Credit Amount'].fillna(0)
            
            df['Bank'] = 'Lloyds'
            df['Transaction Type'] = df['Amount'].apply(lambda x: 'Income' if x > 0 else 'Expense')
            df['Category'] = df['Description'].apply(categorize_transaction)
            return df
        else:
            return None
    except:
        return None

def parse_barclays_statement(file_content):
    """Parse Barclays bank statement"""
    try:
        # Try CSV format
        df = pd.read_csv(StringIO(file_content.decode('utf-8')))
        
        # Standardize Barclays columns
        if 'Transaction Date' in df.columns or 'Date' in df.columns:
            date_col = 'Transaction Date' if 'Transaction Date' in df.columns else 'Date'
            df = df.rename(columns={date_col: 'Date'})
            
            if 'Description' not in df.columns and 'Transaction Description' in df.columns:
                df = df.rename(columns={'Transaction Description': 'Description'})
                
            df['Bank'] = 'Barclays'
            df['Transaction Type'] = df['Amount'].apply(lambda x: 'Income' if x > 0 else 'Expense')
            df['Category'] = df['Description'].apply(categorize_transaction)
            return df
        else:
            return None
    except:
        return None

def categorize_transaction(description):
    """Categorize transactions based on description"""
    description = str(description).lower().strip()
    
    # Pot transfers - handle these first as they're special
    if any(keyword in description for keyword in [
        'transfer from pot', 'transfer to pot', 'pot transfer', 'pot to pot', 'between pots', 
        'move to pot', 'pot withdrawal', 'monzo pot', 'pot deposit', 'pot withdrawal', 
        'savings pot', 'bills pot', 'expenses pot', 'shared pot', 'monzo plus pot'
    ]):
        return 'Pot Transfer'
    
    # P2P payments - handle before general categories
    if any(keyword in description for keyword in ['p2p payment', 'payment to', 'paid to']):
        return 'Transfers & Payments'
    
    # Income categories
    if any(keyword in description for keyword in ['salary', 'wages', 'pay', 'income']):
        return 'Income'
    elif any(keyword in description for keyword in ['refund', 'return', 'cashback']):
        return 'Refund'
    
    # Housing & Utilities
    elif any(keyword in description for keyword in ['rent', 'mortgage', 'property', 'council tax']):
        return 'Housing'
    elif any(keyword in description for keyword in ['electric', 'gas', 'water', 'bill', 'utility', 'broadband', 'internet']):
        return 'Utilities'
    
    # Food & Dining
    elif any(keyword in description for keyword in ['tesco', 'sainsbury', 'asda', 'morrisons', 'grocery', 'food', 'supermarket']):
        return 'Groceries'
    elif any(keyword in description for keyword in ['restaurant', 'cafe', 'coffee', 'dining', 'eat', 'takeaway', 'deliveroo', 'just eat']):
        return 'Dining Out'
    elif any(keyword in description for keyword in ['pub', 'bar', 'wine', 'beer']):
        return 'Alcohol & Social'
    
    # Transportation
    elif any(keyword in description for keyword in ['uber', 'taxi', 'bus', 'train', 'tube', 'transport', 'tfl', 'national rail']):
        return 'Transportation'
    elif any(keyword in description for keyword in ['petrol', 'gas', 'fuel', 'parking']):
        return 'Car Expenses'
    
    # Shopping
    elif any(keyword in description for keyword in ['amazon', 'ebay', 'shop', 'store', 'purchase', 'retail']):
        return 'Shopping'
    elif any(keyword in description for keyword in ['clothing', 'fashion', 'h&m', 'zara', 'primark']):
        return 'Clothing'
    elif any(keyword in description for keyword in ['pharmacy', 'boots', 'superdrug', 'medicine']):
        return 'Health & Pharmacy'
    
    # Entertainment & Subscriptions
    elif any(keyword in description for keyword in ['netflix', 'spotify', 'subscription', 'prime', 'disney+']):
        return 'Subscriptions'
    elif any(keyword in description for keyword in ['cinema', 'movie', 'entertainment', 'theatre', 'concert']):
        return 'Entertainment'
    elif any(keyword in description for keyword in ['gym', 'fitness', 'health', 'exercise']):
        return 'Health & Fitness'
    
    # Financial Services (exclude pot transfers which are handled above)
    elif any(keyword in description for keyword in ['bank', 'interest', 'fee', 'charge', 'payment']):
        return 'Banking & Fees'
    
    # Personal Care
    elif any(keyword in description for keyword in ['hair', 'beauty', 'salon', 'barber']):
        return 'Personal Care'
    
    # Technology
    elif any(keyword in description for keyword in ['apple', 'google', 'microsoft', 'app', 'software']):
        return 'Technology'
    
    # Travel
    elif any(keyword in description for keyword in ['hotel', 'flight', 'holiday', 'travel', 'airbnb', 'booking']):
        return 'Travel'
    
    # Education
    elif any(keyword in description for keyword in ['course', 'education', 'book', 'university']):
        return 'Education'
    
    # Charity
    elif any(keyword in description for keyword in ['charity', 'donation', 'fund']):
        return 'Charity & Donations'
    
    else:
        return 'Other'

def analyze_financial_performance(df):
    """Analyze financial performance using Ollama"""
    try:
        # Separate pot transfers from regular transactions
        pot_transfers = df[df['Category'] == 'Pot Transfer']
        regular_transactions = df[df['Category'] != 'Pot Transfer']
        
        # Prepare summary statistics (excluding pot transfers for main calculations)
        total_income = regular_transactions[regular_transactions['Amount'] > 0]['Amount'].sum()
        total_expenses = abs(regular_transactions[regular_transactions['Amount'] < 0]['Amount'].sum())
        net_savings = total_income - total_expenses
        pot_transfer_total = abs(pot_transfers['Amount'].sum()) if not pot_transfers.empty else 0
        
        # Category breakdown (exclude pot transfers from expenses)
        expense_by_category = regular_transactions[regular_transactions['Amount'] < 0].groupby('Category')['Amount'].sum().abs().sort_values(ascending=False)
        
        # Recent transactions (show more transactions)
        recent_transactions = df.head(20).to_dict('records')
        
        # Create an optimized prompt for faster processing
        # Limit data sent to Ollama to prevent timeouts
        category_summary = expense_by_category.head(10).to_string() if len(expense_by_category) > 10 else expense_by_category.to_string()
        
        prompt = f"""Analyze this financial data:
        
        SUMMARY:
        - Total Income: ${total_income:,.2f}
        - Total Expenses: ${total_expenses:,.2f}
        - Net Savings: ${net_savings:,.2f}
        - Pot Transfers: ${pot_transfer_total:,.2f}
        - Balance Left: ${total_income - total_expenses:,.2f}
        - Total Transactions: {len(df)}
        - Regular Transactions: {len(regular_transactions)}
        - Pot Transfer Transactions: {len(pot_transfers)}
        
        TOP EXPENSE CATEGORIES:
        {category_summary}
        
        RECENT TRANSACTIONS (sample):
        {str(recent_transactions[:5])}
        
        Provide analysis covering:
        1. Financial health assessment
        2. Main spending patterns
        3. Top savings opportunities
        4. 3-4 specific recommendations
        5. Note about pot transfers if present (these are internal money movements between accounts/pots)
        
        Keep response concise but thorough."""
        
        # Call Ollama without timeout for unlimited processing time
        try:
            result = subprocess.run(
                ['ollama', 'run', 'llama2', prompt],
                capture_output=True,
                text=True
            )
            
            if result.returncode == 0:
                return result.stdout
            else:
                return f"⚠️ Ollama analysis failed: {result.stderr}"
                
        except Exception as e:
            return f"⚠️ Error during analysis: {str(e)}"
            
    except Exception as e:
        return f"❌ Error preparing analysis: {str(e)}"

def main():
    st.title("📊 Life & Budget Dashboard")
    st.markdown("### Your All-in-One Financial and Personal Management Tool")
    
    # Add AI-friendly template download button
    if st.sidebar.button("📥 Download AI-Friendly Template"):
        with st.spinner("Creating your template..."):
            template_path = create_ai_friendly_template()
            with open(template_path, "rb") as f:
                template_data = f.read()
            
            st.sidebar.success("Template created!")
            st.sidebar.download_button(
                label="Download AI-Friendly Template",
                data=template_data,
                file_name="ai_finance_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Clean up the temporary file
            try:
                os.remove(template_path)
            except:
                pass
    
    # Sidebar for navigation
    st.sidebar.title("Navigation")
    page = st.sidebar.radio("Go to", ["Home", "Generate Template", "AI Insights", "AI Template Converter", "Bank Statement Analysis"])
    
    if page == "Home":
        st.markdown("""
        Welcome to your personal Life & Budget Dashboard! This tool helps you track your finances, 
        health, and daily activities all in one place.
        
        ### Features:
        - **Financial Tracking**: Income, expenses, savings, and investments
        - **Health & Wellness**: Weight tracking and habit formation
        - **Life Organization**: Cleaning schedules and meal planning
        
        Get started by generating a new template or upload your existing data for AI-powered insights.
        """)
        
    elif page == "Generate Template":
        st.header("📝 Generate New Template")
        st.write("Create a new Excel template with the sections you need.")
        
        # Month selection
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        selected_month = st.selectbox("Select Month", months, index=datetime.date.today().month - 1)
        
        # Section selection
        st.subheader("Select Sections to Include")
        
        # Create columns for better organization
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("**Financial**")
            income = st.checkbox("Income", value=True, key="income_cb")
            expenses = st.checkbox("Expenses", value=True, key="expenses_cb")
            monthly_purchases = st.checkbox("Monthly Purchases", value=True, key="monthly_purchases_cb")
            savings = st.checkbox("Savings", value=True, key="savings_cb")
            stocks = st.checkbox("📊 Stock Tracker", value=True, key="stocks_cb")
            debt = st.checkbox("💳 Debt Tracker", value=True, key="debt_cb")
        
        with col2:
            st.markdown("**Health**")
            weight = st.checkbox("Weight", value=True, key="weight_cb")
            habits = st.checkbox("Habits", value=True, key="habits_cb")
        
        with col3:
            st.markdown("**Life Organization**")
            cleaning = st.checkbox("Cleaning", value=True, key="cleaning_cb")
            meals = st.checkbox("Meal Planning", value=True, key="meals_cb")
            timetable = st.checkbox("Time Table", value=True, key="timetable_cb")
        
        st.markdown("---")

        # Create a single, correct list of selected sections to be used everywhere
        selected_sections = []
        if income: selected_sections.append("income")
        if expenses: selected_sections.append("expenses")
        if savings: selected_sections.append("savings")
        if stocks: selected_sections.append("stocks")
        if debt: selected_sections.append("debt")
        if weight: selected_sections.append("weight")
        if habits: selected_sections.append("habits")
        if cleaning: selected_sections.append("cleaning")
        if meals: selected_sections.append("meals")
        if timetable: selected_sections.append("timetable")
        # Always include monthly purchases with expenses
        if expenses: selected_sections.append("monthly_purchases")

        # If no sections are selected, include all
        if not selected_sections:
            selected_sections = [
                'income', 'expenses', 'monthly_purchases', 'savings', 'stocks', 
                'debt', 'weight', 'habits', 'cleaning', 'meals', 'timetable'
            ]

        if st.button("🔍 Generate Preview"):
            with st.spinner(f"Creating your {selected_month} template..."):
                try:
                    # Generate the workbook in memory
                    wb = create_excel_template(month=selected_month, sections=selected_sections)
                    
                    st.subheader("Excel Sheet Preview")
                    for sheet in wb.worksheets:
                        with st.expander(f"Sheet: {sheet.title}"):
                            # Convert sheet to DataFrame for preview
                            df = worksheet_to_dataframe(sheet)
                            st.dataframe(df)
                except Exception as e:
                    st.error(f"Error generating preview: {str(e)}")
        
        if st.button("✨ Generate Template"):
            with st.spinner(f"Creating your {selected_month} template..."):
                # Create a temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    file_path = tmp.name
                
                try:
                    # Create the template with the selected month and sections
                    wb = create_excel_template(month=selected_month, sections=selected_sections)
                    
                    # Debug: Print the sheets that were created
                    st.sidebar.write("Sheets created:", wb.sheetnames)
                    
                    # Add Monthly Purchases sheet if expenses are included
                    if 'Expense Tracker' in wb.sheetnames and 'Monthly Purchases' not in wb.sheetnames and monthly_purchases:
                        monthly_purchases_sheet = wb.create_sheet("Monthly Purchases")
                        monthly_purchases_sheet.append(["Date", "Item", "Amount", "Type", "Category", "Notes"])
                        
                        # Add sample data
                        sample_data = [
                            [datetime.date.today().replace(day=1), "Netflix", 15.99, "Subscription", "Entertainment", "Monthly plan"],
                            [datetime.date.today().replace(day=5), "Gym Membership", 45.00, "Subscription", "Health", "Monthly membership"],
                            [datetime.date.today().replace(day=10), "Office Chair", 199.99, "One-Time", "Furniture", "Ergonomic chair"]
                        ]
                        for row in sample_data:
                            monthly_purchases_sheet.append(row)
                            
                        # Style the header
                        for cell in monthly_purchases_sheet[1]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
                    
                except Exception as e:
                    st.error(f"Error creating Excel template: {str(e)}")
                    st.stop()
                wb.save(file_path)
                
                # Create download link
                with open(file_path, 'rb') as f:
                    bytes_data = f.read()
                
                st.success(f"{selected_month} template created successfully!")
                
                # Create columns for download options
                col1, col2 = st.columns(2)
                
                with col1:
                    st.download_button(
                        label=f"📥 Download Excel Template",
                        data=bytes_data,
                        file_name=f"life_budget_tracker_{selected_month}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with col2:
                    # Generate PDF report
                    with st.spinner("Creating PDF report..."):
                        try:
                            pdf_data = create_pdf_report(month=selected_month, sections=selected_sections)
                            if pdf_data:
                                st.download_button(
                                    label=f"📄 Download PDF Report",
                                    data=pdf_data,
                                    file_name=f"financial_report_{selected_month}.pdf",
                                    mime="application/pdf"
                                )
                            else:
                                st.error("Failed to generate PDF report")
                        except Exception as e:
                            st.error(f"Error creating PDF: {str(e)}")
                
                # Show features comparison
                st.markdown("---")
                st.markdown("### 📋 Export Options Comparison")
                
                comparison_data = [
                    ["Feature", "Excel Template", "PDF Report"],
                    ["📊 Charts & Graphs", "✅ Interactive", "✅ High-Quality"],
                    ["📝 Data Entry", "✅ Editable", "❌ Read-only"],
                    ["📱 Mobile Friendly", "⚠️ Limited", "✅ Optimized"],
                    ["🎨 Professional Design", "✅ Enhanced", "✅ Premium"],
                    ["📈 Financial Analysis", "✅ Formulas", "✅ Insights"],
                    ["🔄 Real-time Updates", "✅ Dynamic", "❌ Static"]
                ]
                
                comparison_table = st.table(comparison_data)
                
                # Clean up
                os.unlink(file_path)
                
    elif page == "AI Template Converter":
        st.header("🤖 Convert to AI-Friendly Format")
        st.write("Upload your Excel file to convert it to a more AI-friendly format that works better with Ollama.")
        
        uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])
        
        if uploaded_file is not None:
            with st.spinner("Converting your file..."):
                # Create a temporary file for the uploaded file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_in:
                    tmp_in.write(uploaded_file.getvalue())
                    input_path = tmp_in.name
                
                # Create a temporary output file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_out:
                    output_path = tmp_out.name
                
                try:
                    # Convert the file
                    success, message = convert_to_ai_friendly(input_path, output_path)
                    
                    if success:
                        st.success("File converted successfully!")
                        
                        # Create download link
                        with open(output_path, 'rb') as f:
                            bytes_data = f.read()
                        
                        st.download_button(
                            label="💾 Download AI-Friendly File",
                            data=bytes_data,
                            file_name=f"ai_friendly_{uploaded_file.name}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        st.markdown("### What was improved:")
                        st.markdown("""
                        - 🧹 Removed formatting and merged cells
                        - 📊 Standardized data types and formats
                        - 📅 Ensured consistent date formatting
                        - 🏷️ Added clear headers
                        - 🗑️ Removed empty rows and columns
                        - 📝 Added metadata and instructions
                        """)
                    else:
                        st.error(f"Error during conversion: {message}")
                        
                except Exception as e:
                    st.error(f"An error occurred: {str(e)}")
                    
                finally:
                    # Clean up temporary files
                    for path in [input_path, output_path]:
                        try:
                            if os.path.exists(path):
                                os.unlink(path)
                        except:
                            pass
    
    elif page == "AI Insights":
        st.header("📊 Upload Your Data for AI Insights")
        
        # Add category selection
        categories = [
            'Income',
            'Expenses',
            'Savings',
            'Investments',
            'Debt',
            'Health',
            'Lifestyle'
        ]
        
        selected_categories = st.multiselect(
            "Select categories to analyze:",
            categories,
            default=categories[:2]  # Default to Income and Expenses
        )
        
        uploaded_file = st.file_uploader("Upload your filled Excel file", type=["xlsx"])
        
        if uploaded_file is not None:
            if st.button("🤖 Generate AI Insights", key="ai_insights_btn"):
                with st.spinner("Analyzing your data with AI..."):
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        tmp.write(uploaded_file.getvalue())
                        tmp_path = tmp.name
                    
                    try:
                        excel_data_str, insights = generate_ai_insights(tmp_path, selected_categories)
                        
                        # --- DEBUG: Show the data sent to the AI --- #
                        with st.expander("View Data Sent to AI (for debugging)"):
                            st.text(excel_data_str)
                        # --- END DEBUG --- #

                        if not insights.startswith(('❌', '⚠️')):
                            st.success("AI Analysis Complete!")
                            st.markdown("### 🎯 Your Personalized Insights")
                            
                            with st.expander("View Insights", expanded=True):
                                st.markdown(insights)
                            
                            # Download buttons
                            col1, col2 = st.columns(2)
                            with col1:
                                st.download_button(
                                    label="📝 Download as Text",
                                    data=insights,
                                    file_name="financial_insights.txt",
                                    mime="text/plain"
                                )
                            with col2:
                                pdf_path = generate_pdf(insights)
                                with open(pdf_path, "rb") as f:
                                    st.download_button(
                                        label="📄 Download as PDF",
                                        data=f,
                                        file_name="financial_insights.pdf",
                                        mime="application/pdf"
                                    )
                                os.unlink(pdf_path)
                        else:
                            st.error(insights)
                    finally:
                        if os.path.exists(tmp_path):
                            os.unlink(tmp_path)
    
    elif page == "Bank Statement Analysis":
        st.header("🏦 Bank Statement Analysis")
        st.write("Upload your bank statement for AI-powered financial performance analysis using Ollama.")
        
        # Bank selection
        st.subheader("Select Your Bank")
        bank_options = ["Monzo", "Lloyds", "Barclays"]
        selected_bank = st.selectbox("Choose your bank:", bank_options)
        
        # File upload
        st.subheader(f"Upload {selected_bank} Statement")
        uploaded_file = st.file_uploader(
            f"Choose your {selected_bank} statement file",
            type=["csv", "pdf", "xlsx", "xls"],
            key="bank_statement"
        )
        
        if uploaded_file is not None:
            st.success(f"✅ {selected_bank} statement uploaded successfully!")
            
            # Parse the statement based on bank
            with st.spinner(f"Parsing {selected_bank} statement..."):
                file_content = uploaded_file.getvalue()
                
                if selected_bank == "Monzo":
                    df = parse_monzo_statement(file_content)
                elif selected_bank == "Lloyds":
                    df = parse_lloyds_statement(file_content)
                elif selected_bank == "Barclays":
                    df = parse_barclays_statement(file_content)
                
                if df is not None:
                    st.success("✅ Statement parsed successfully!")
                    
                    # Display Monzo balance summary if available
                    if hasattr(df, 'attrs') and df.attrs:
                        st.subheader("🏦 Monzo Balance Summary")
                        balance_cols = st.columns(3)
                        
                        with balance_cols[0]:
                            if 'total_balance_including_pots' in df.attrs:
                                st.metric("Total Balance (Inc. Pots)", f"£{df.attrs['total_balance_including_pots']:,.2f}")
                            if 'personal_account_balance' in df.attrs:
                                st.metric("Personal Account Balance", f"£{df.attrs['personal_account_balance']:,.2f}")
                        
                        with balance_cols[1]:
                            if 'balance_in_pots' in df.attrs:
                                st.metric("Balance in Pots", f"£{df.attrs['balance_in_pots']:,.2f}")
                            if 'cashback_balance' in df.attrs:
                                st.metric("Cashback Balance", f"£{df.attrs['cashback_balance']:,.2f}")
                        
                        with balance_cols[2]:
                            if 'total_outgoings' in df.attrs:
                                st.metric("Total Outgoings", f"£{df.attrs['total_outgoings']:,.2f}")
                            if 'total_deposits' in df.attrs:
                                st.metric("Total Deposits", f"£{df.attrs['total_deposits']:,.2f}")
                    
                    # Show summary statistics with pot transfer handling
                    st.subheader("📊 Transaction Summary")
                    col1, col2, col3, col4, col5, col6 = st.columns(6)
                    
                    # Separate pot transfers from regular transactions
                    pot_transfers = df[df['Category'] == 'Pot Transfer']
                    regular_transactions = df[df['Category'] != 'Pot Transfer']
                    
                    with col1:
                        total_income = regular_transactions[regular_transactions['Amount'] > 0]['Amount'].sum()
                        st.metric("Total Income", f"£{total_income:,.2f}")
                    
                    with col2:
                        total_expenses = abs(regular_transactions[regular_transactions['Amount'] < 0]['Amount'].sum())
                        st.metric("Total Expenses", f"£{total_expenses:,.2f}")
                    
                    with col3:
                        net_savings = total_income - total_expenses
                        st.metric("Net Savings", f"£{net_savings:,.2f}")
                    
                    with col4:
                        pot_in = pot_transfers[pot_transfers['Amount'] > 0]['Amount'].sum() if not pot_transfers.empty else 0
                        st.metric("Pot Money In", f"£{pot_in:,.2f}")
                    
                    with col5:
                        pot_out = abs(pot_transfers[pot_transfers['Amount'] < 0]['Amount'].sum() if not pot_transfers.empty else 0)
                        st.metric("Pot Money Out", f"£{pot_out:,.2f}")
                    
                    with col6:
                        # Calculate balance left (income - expenses, excluding pot transfers)
                        balance_left = total_income - total_expenses
                        st.metric("Balance Left", f"£{balance_left:,.2f}")
                    
                    # Show pot transfer details if any exist
                    if not pot_transfers.empty:
                        st.subheader("🔄 Pot Transfer Transactions")
                        st.dataframe(pot_transfers, width='stretch')
                        
                        # Pot transfer summary
                        pot_in = pot_transfers[pot_transfers['Amount'] > 0]['Amount'].sum()
                        pot_out = abs(pot_transfers[pot_transfers['Amount'] < 0]['Amount'].sum())
                        
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Pot Money In", f"${pot_in:,.2f}")
                        with col2:
                            st.metric("Pot Money Out", f"${pot_out:,.2f}")
                        with col3:
                            st.metric("Net Pot Movement", f"${pot_in - pot_out:,.2f}")
                    
                    # Show transaction preview
                    st.subheader("📋 Recent Transactions")
                    st.dataframe(df.head(10), width='stretch')
                    
                    # Category breakdown (exclude pot transfers from expenses)
                    st.subheader("📈 Spending by Category")
                    expense_df = regular_transactions[regular_transactions['Amount'] < 0]
                    if not expense_df.empty:
                        category_summary = expense_df.groupby('Category')['Amount'].sum().abs().sort_values(ascending=False)
                        
                        # Create pie chart
                        fig = px.pie(
                            values=category_summary.values,
                            names=category_summary.index,
                            title="Expense Breakdown (Excluding Pot Transfers)"
                        )
                        st.plotly_chart(fig, width='stretch')
                    
                    # Transaction count summary
                    st.subheader("📊 Transaction Summary")
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        regular_count = len(regular_transactions)
                        st.metric("Regular Transactions", regular_count)
                    
                    with col2:
                        pot_count = len(pot_transfers)
                        st.metric("Pot Transfer Transactions", pot_count)
                    
                    with col3:
                        total_count = len(df)
                        st.metric("Total Transactions", total_count)
                    
                    # AI Analysis button
                    if st.button(f"🤖 Analyze with Ollama", type="primary"):
                        with st.spinner("Analyzing your financial performance with AI..."):
                            analysis = analyze_financial_performance(df)
                            
                            st.subheader("🧠 AI Financial Analysis")
                            if analysis.startswith("⚠️") or analysis.startswith("❌"):
                                st.error(analysis)
                                
                                # Show basic analysis even if AI fails
                                st.subheader("📊 Basic Analysis")
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    st.write("**Financial Summary:**")
                                    st.write(f"- Total Income: £{total_income:,.2f}")
                                    st.write(f"- Total Expenses: £{total_expenses:,.2f}")
                                    st.write(f"- Net Savings: £{net_savings:,.2f}")
                                    st.write(f"- Balance Left: £{balance_left:,.2f}")
                                    savings_rate = (net_savings/total_income*100) if total_income > 0 else 0
                                    st.write(f"- Savings Rate: {savings_rate:.1f}%")
                                    
                                    st.write("**Pot Transfer Summary:**")
                                    st.write(f"- Pot Money In: £{pot_in:,.2f}")
                                    st.write(f"- Pot Money Out: £{pot_out:,.2f}")
                                    st.write(f"- Net Pot Movement: £{pot_in - pot_out:,.2f}")
                                
                                with col2:
                                    st.write("**Top Spending Categories:**")
                                    if not expense_df.empty:
                                        category_summary = expense_df.groupby('Category')['Amount'].sum().abs().sort_values(ascending=False)
                                        for cat, amount in category_summary.head(3).items():
                                            st.write(f"- {cat}: £{abs(amount):,.2f}")
                                
                                # Basic recommendations
                                st.subheader("💡 Quick Recommendations")
                                if net_savings < 0:
                                    st.warning("⚠️ You're spending more than you earn. Consider reviewing expenses.")
                                elif net_savings < total_income * 0.1:  # Less than 10% savings
                                    st.info("💰 Try to increase savings to at least 10% of income.")
                                else:
                                    st.success("✅ Good savings rate! Keep tracking expenses.")
                                    
                                # Pot transfer insights
                                if pot_in > 0 or pot_out > 0:
                                    st.info(f"🔄 You moved £{pot_out:,.2f} out of pots and £{pot_in:,.2f} into pots. Net pot movement: £{pot_in - pot_out:,.2f}")
                            else:
                                st.markdown(analysis)
                                
                                # Show analysis confidence
                                st.info("🤖 Analysis powered by Ollama Llama2")
                            
                            # Download analysis
                            analysis_bytes = analysis.encode('utf-8')
                            st.download_button(
                                label="📥 Download Analysis",
                                data=analysis_bytes,
                                file_name=f"financial_analysis_{selected_bank}_{datetime.datetime.now().strftime('%Y%m%d')}.txt",
                                mime="text/plain"
                            )
                    
                    # Export processed data
                    st.subheader("💾 Export Processed Data")
                    csv = df.to_csv(index=False)
                    st.download_button(
                        label="📊 Download CSV",
                        data=csv,
                        file_name=f"processed_{selected_bank}_statement.csv",
                        mime="text/csv"
                    )
                    
                else:
                    st.error(f"❌ Failed to parse {selected_bank} statement.")
                    st.info("🔍 **Debugging Info:** Check the terminal/console output for detailed parsing information.")
                    
                    # Show file format help
                    with st.expander("📋 Expected File Format"):
                        st.markdown(f"""
                        **{selected_bank} Statement Requirements:**
                        
                        **Required Columns (any variation of these names):**
                        - Date/Transaction Date/Posted Date
                        - Description/Details/Transaction Description
                        - Amount/Value/Debit/Credit
                        
                        **Supported File Types:**
                        - CSV (.csv)
                        - Excel (.xlsx, .xls)
                        - PDF (limited support)
                        
                        **Common Column Name Variations:**
                        - Date: 'date', 'transaction date', 'posted date'
                        - Description: 'description', 'details', 'memo', 'notes'
                        - Amount: 'amount', 'value', 'debit', 'credit'
                        
                        **Tips:**
                        - Ensure your file has headers in the first row
                        - Check for extra spaces in column names
                        - Make sure the file is not password protected
                        """)
                    
                    # Show raw file preview
                    try:
                        st.subheader("🔍 File Preview")
                        preview_df = pd.read_csv(StringIO(file_content.decode('utf-8')))
                        st.write("**Columns found:**", list(preview_df.columns))
                        st.write("**First 5 rows:**")
                        st.dataframe(preview_df.head(), use_container_width=True)
                    except Exception as e:
                        st.error(f"Could not preview file: {str(e)}")
    
    st.markdown("---")
    st.markdown("### 📱 Features at a Glance")
    
    features = st.columns(3)
    
    with features[0]:
        st.markdown("""
        #### 💰 Financial Tracking
        - Income & Expense tracking
        - Budget planning
        - Savings goals
        - Investment portfolio
        """)
    
    with features[1]:
        st.markdown("""
        #### 🏋️ Health & Wellness
        - Weight tracking
        - Habit formation
        - Self-care routines
        - Meal planning
        """)
    
    with features[2]:
        st.markdown("""
        #### 🏠 Life Organization
        - Cleaning checklist
        - Weekly schedule
        - Time management
        - Task tracking
        """)
    
    st.markdown("---")
    st.markdown("""
    ### Getting Started
    1. Download the Excel template above
    2. Fill in your financial and personal data
    3. Upload it back to get AI-powered insights
    4. Use the dashboard to track your progress
    """)

def generate_pdf(insights_text, excel_data_str=None):
    """
    Generate a professional PDF report from insights and optional Excel data
    
    Args:
        insights_text (str): The insights text to include in the PDF
        excel_data_str (str, optional): Formatted Excel data as string. Defaults to None.
    
    Returns:
        str: Path to the generated PDF file
    """
    try:
        # Create PDF object with proper margins
        pdf = FPDF()
        
        # Set document properties
        pdf.set_title("Financial Insights Report")
        pdf.set_author("Finance Budget System")
        pdf.set_auto_page_break(auto=True, margin=20)
        
        # Add a page
        pdf.add_page()
        
        # Define styles
        styles = {
            'title': {'font': 'Arial', 'style': 'B', 'size': 24, 'color': (0, 51, 102)},
            'subtitle': {'font': 'Arial', 'style': '', 'size': 12, 'color': (100, 100, 100)},
            'header1': {'font': 'Arial', 'style': 'B', 'size': 16, 'color': (0, 51, 102)},
            'header2': {'font': 'Arial', 'style': 'B', 'size': 14, 'color': (0, 71, 133)},
            'header3': {'font': 'Arial', 'style': 'B', 'size': 12, 'color': (0, 91, 150)},
            'normal': {'font': 'Arial', 'style': '', 'size': 11, 'color': (0, 0, 0)},
            'footer': {'font': 'Arial', 'style': 'I', 'size': 8, 'color': (100, 100, 100)},
            'table_header': {'fill': True, 'fill_color': (240, 240, 240), 'text_color': (0, 0, 0), 'border': 1},
            'table_row': {'fill': False, 'fill_color': (255, 255, 255), 'text_color': (0, 0, 0), 'border': 1}
        }
        
        # Set default font
        pdf.set_font(styles['normal']['font'], styles['normal']['style'], styles['normal']['size'])
        
        # Add header function
        def add_header():
            # Add company logo (commented out - add path to your logo)
            # try:
            #     pdf.image('path/to/logo.png', 10, 8, 33)
            # except:
            #     pass  # Skip if logo not found
            
            # Add title
            pdf.set_font(styles['title']['font'], styles['title']['style'], styles['title']['size'])
            pdf.set_text_color(*styles['title']['color'])
            pdf.cell(0, 10, 'FINANCIAL INSIGHTS REPORT', 0, 1, 'C')
            
            # Add subtitle with date
            pdf.set_font(styles['subtitle']['font'], styles['subtitle']['style'], styles['subtitle']['size'])
            pdf.set_text_color(*styles['subtitle']['color'])
            pdf.cell(0, 5, f'Generated on: {datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p")}', 0, 1, 'C')
            
            # Add a line
            pdf.set_draw_color(200, 200, 200)
            pdf.line(10, pdf.get_y() + 5, 200, pdf.get_y() + 5)
            pdf.ln(10)
        
        # Add footer function
        def add_footer():
            # Position at 1.5 cm from bottom
            pdf.set_y(-15)
            pdf.set_font(styles['footer']['font'], styles['footer']['style'], styles['footer']['size'])
            pdf.set_text_color(*styles['footer']['color'])
            
            # Page number
            page_num = pdf.page_no()
            pdf.cell(0, 10, f'Page {page_num}', 0, 0, 'C')
            
            # Add a line
            pdf.set_draw_color(200, 200, 200)
            pdf.line(10, 280, 200, 280)
        
        # Clean the text to ensure professional formatting
        def clean_text(text):
            if not isinstance(text, str):
                text = str(text)
            
            # Remove emojis and other non-ASCII characters
            text = text.encode('ascii', 'ignore').decode('ascii')
            
            # Replace common problematic characters
            replacements = {
                '–': '-', '—': '-', '•': '*',
                '“': '"', '”': '"', '‘': "'", '’': "'",
                '…': '...', '—': '--', '–': '-', '•': '*',
                '·': '*', '`': "'", '´': "'"
            }
            
            for old, new in replacements.items():
                text = text.replace(old, new)
                
            # Clean up any remaining non-printable characters
            text = ''.join(char for char in text if char.isprintable() or char.isspace())
            return text.strip()
        
        # Add a section with proper styling
        def add_section(title, level=1, add_toc=True):
            if pdf.get_y() > 250:  # Prevent section headers at the bottom of the page
                pdf.add_page()
            
            # Add to table of contents
            if add_toc and hasattr(pdf, 'toc'):
                pdf.toc.append((level, title, pdf.page_no()))
            
            # Add section header
            if level == 1:
                style = styles['header1']
                spacing = 10
            elif level == 2:
                style = styles['header2']
                spacing = 8
            else:
                style = styles['header3']
                spacing = 6
            
            pdf.set_font(style['font'], style['style'], style['size'])
            pdf.set_text_color(*style['color'])
            pdf.ln(spacing)
            
            # Add a small colored line under the header
            pdf.cell(0, 5, title, 0, 1, 'L')
            pdf.set_draw_color(style['color'][0], style['color'][1], style['color'][2])
            pdf.line(pdf.l_margin, pdf.get_y(), 50, pdf.get_y())
            pdf.ln(5)
            
            # Reset to normal text
            pdf.set_font(styles['normal']['font'], styles['normal']['style'], styles['normal']['size'])
            pdf.set_text_color(*styles['normal']['color'])
        
        # Add header to the first page
        add_header()
        
        # Initialize table of contents
        pdf.toc = []
        
        # Add table of contents page
        add_section('Table of Contents', 1, False)
        
        # Add sections
        add_section('Executive Summary', 1)
        
        # Add insights text with proper formatting
        current_paragraph = []
        
        for line in str(insights_text).split('\n'):
            line = clean_text(line.strip())
            if not line:
                if current_paragraph:
                    pdf.multi_cell(0, 6, ' '.join(current_paragraph))
                    current_paragraph = []
                pdf.ln(4)
                continue
                
            # Handle bullet points
            if line.startswith(('*', '-')):
                if current_paragraph:
                    pdf.multi_cell(0, 6, ' '.join(current_paragraph))
                    current_paragraph = []
                pdf.cell(10, 6, '•', 0, 0)
                pdf.multi_cell(0, 6, line[1:].strip())
            # Handle section headers in the text
            elif line.endswith(':'):
                if current_paragraph:
                    pdf.multi_cell(0, 6, ' '.join(current_paragraph))
                    current_paragraph = []
                add_section(line[:-1], 2)  # Remove the colon
            else:
                current_paragraph.append(line)
        
        # Add any remaining content
        if current_paragraph:
            pdf.multi_cell(0, 6, ' '.join(current_paragraph))
        
        # Add Excel data if provided
        if excel_data_str:
            add_section('Detailed Financial Analysis', 1)
            
            # Add a note about the data
            pdf.set_font(styles['normal']['font'], 'I', styles['normal']['size'] - 1)
            pdf.multi_cell(0, 5, "The following section contains detailed financial data from your records. "
                                "For a more comprehensive analysis, please review the full dataset in the Excel file.")
            pdf.ln(5)
            
            # Add the data in a clean, monospaced font
            pdf.set_font('Courier', '', 9)
            
            # Split the data into lines and add with proper formatting
            data_lines = clean_text(str(excel_data_str)).split('\n')
            for line in data_lines:
                if pdf.get_y() > 270:  # Add new page if needed
                    add_footer()
                    pdf.add_page()
                    add_header()
                pdf.cell(0, 4, line, 0, 1)
        
        # Add table of contents at the beginning
        def add_table_of_contents():
            # Save current page number
            current_page = pdf.page_no()
            
            # Go to first page
            pdf.page = 0
            
            # Find the TOC position (after the header)
            pdf.set_y(40)
            
            # Add TOC title
            pdf.set_font(styles['header1']['font'], 'B', styles['header1']['size'])
            pdf.set_text_color(*styles['header1']['color'])
            pdf.cell(0, 10, 'Table of Contents', 0, 1, 'L')
            pdf.ln(5)
            
            # Add TOC entries
            pdf.set_font(styles['normal']['font'], '', styles['normal']['size'] - 1)
            
            for level, title, page in pdf.toc:
                # Indent based on level
                indent = (level - 1) * 10
                
                # Calculate y position and add new page if needed
                if pdf.get_y() > 260:
                    pdf.add_page()
                    add_header()
                    pdf.set_y(40)
                
                # Add entry
                pdf.cell(indent, 6, '')
                pdf.cell(140 - indent, 6, title, 0, 0, 'L')
                
                # Add dotted line
                dot_leader = '.' * (60 - len(title) - len(str(page)))
                pdf.cell(0, 6, f"{dot_leader}{page}", 0, 1, 'R')
            
            # Restore to the last page
            pdf.page = current_page
        
        # Add TOC if we have entries
        if hasattr(pdf, 'toc') and pdf.toc:
            add_table_of_contents()
        
        # Add footer to all pages
        for i in range(1, pdf.page_no() + 1):
            pdf.page = i - 1
            add_footer()
        
        # Save the PDF to a temporary file
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        pdf_path = temp_pdf.name
        pdf.output(pdf_path)
        
        return pdf_path
        
    except Exception as e:
        # Create a simple error PDF if something goes wrong
        error_pdf = FPDF()
        error_pdf.add_page()
        error_pdf.set_font('Arial', 'B', 16)
        error_pdf.cell(0, 10, 'Error Generating Report', 0, 1, 'C')
        error_pdf.set_font('Arial', '', 12)
        error_pdf.multi_cell(0, 10, f'An error occurred while generating the PDF report:\n\n{str(e)}')
        
        # Save the error PDF
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        error_pdf_path = temp_pdf.name
        error_pdf.output(error_pdf_path)
        return error_pdf_path
    
    return pdf_path

if __name__ == "__main__":
    main()