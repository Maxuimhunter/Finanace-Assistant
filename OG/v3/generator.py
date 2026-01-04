# app.py
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, PieChart3D
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import os
import datetime
import json
import ollama
import time
from io import BytesIO
from fpdf import FPDF
import base64
import io
from pathlib import Path

# Set page config
st.set_page_config(
    page_title="Life & Budget Dashboard",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for styling
st.markdown("""
<style>
    .main {
        background-color: #f8f9fa;
    }
    .stButton>button {
        background-color: #6f42c1;
        color: white;
        border-radius: 8px;
        padding: 0.5rem 1rem;
    }
    .stButton>button:hover {
        background-color: #5a32a3;
        color: white;
    }
    .stTextInput>div>div>input {
        border: 1px solid #6f42c1;
        border-radius: 4px;
    }
    .css-1d391kg {
        padding-top: 3rem;
    }
</style>
""", unsafe_allow_html=True)

def create_welcome_guide(sheet):
    """Create the Welcome Guide sheet with instructions and overview."""
    # Set column widths
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 50
    
    # Title
    title_cell = sheet['B2']
    title_cell.value = "Welcome to Your Personal Finance & Life Tracker"
    title_cell.font = Font(size=18, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    sheet.merge_cells('B2:D2')
    
    # Subtitle
    subtitle = sheet['B4']
    subtitle.value = "Your All-in-One Solution for Managing Finances and Life"
    subtitle.font = Font(size=12, italic=True)
    sheet.merge_cells('B4:D4')
    
    # Sections
    sections = [
        ("📊 Dashboard", "Overview of your financial health and key metrics"),
        ("💰 Income Tracker", "Log and categorize all your income sources"),
        ("💸 Expense Tracker", "Track and analyze your spending"),
        ("🏦 Savings Tracker", "Monitor your savings progress and goals"),
        ("📈 Stock Tracker", "Keep an eye on your investments"),
        ("⚖️ Weight Tracker", "Track your weight and health metrics"),
        ("✅ Habit Tracker", "Build and maintain positive habits"),
        ("🧹 Cleaning Schedule", "Stay on top of household chores"),
        ("🍽️ Meal Planner", "Plan your meals and grocery shopping"),
        ("📅 Weekly Schedule", "Organize your weekly activities")
    ]
    
    # Add sections
    for i, (title, desc) in enumerate(sections, start=6):
        # Section title
        title_cell = sheet.cell(row=i, column=2, value=title)
        title_cell.font = Font(bold=True, size=12, color='2F5496')
        
        # Section description
        desc_cell = sheet.cell(row=i, column=3, value=desc)
        
        # Add a subtle bottom border
        for col in range(2, 5):
            cell = sheet.cell(row=i, column=col)
            cell.border = Border(bottom=Side(style='thin', color='D9D9D9'))
    
    # Instructions
    instructions = [
        ("Getting Started", "1. Begin by reviewing each tab to understand what's available"),
        ("", "2. Update the Dashboard with your monthly budget and goals"),
        ("", "3. Enter your income and expenses as they occur"),
        ("", "4. Use the trackers to monitor your progress"),
        ("", "5. Check back regularly to stay on top of your finances and life!")
    ]
    
    # Add instructions
    start_row = len(sections) + 8
    for i, (title, instr) in enumerate(instructions, start=start_row):
        if title:  # Only add the title if it's not empty
            title_cell = sheet.cell(row=i, column=2, value=title)
            title_cell.font = Font(bold=True, size=12, color='2F5496')
            sheet.merge_cells(f'B{i}:D{i}')
            i += 1
        
        instr_cell = sheet.cell(row=i, column=3, value=instr)
        sheet.merge_cells(f'C{i}:D{i}')
    
    # Add a footer
    footer_row = i + 3
    footer = sheet.cell(row=footer_row, column=2, 
                       value="Thank you for using the Personal Finance & Life Tracker! 🚀")
    footer.font = Font(italic=True, color='7F7F7F')
    sheet.merge_cells(f'B{footer_row}:D{footer_row}')
    
    # Set row heights for better spacing
    for i in range(1, footer_row + 2):
        sheet.row_dimensions[i].height = 20
    
    # Add some padding at the bottom
    sheet.row_dimensions[footer_row + 2].height = 10


def is_merged_cell(sheet, row, col):
    """Check if a cell is part of a merged range."""
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and 
            merged_range.min_col <= col <= merged_range.max_col):
            return True
    return False

def create_excel_template(month=None, sections=None):
    """
    Create an Excel template for a specific month with the specified sections.
    
    Args:
        month (str, optional): The month to generate the template for (e.g., 'Jan', 'Feb'). 
                              If None, uses current month.
        sections (list, optional): List of section names to include. If None, includes all sections.
                                  Possible values: 'dashboard', 'income', 'expenses', 'savings', 'stocks', 
                                  'weight', 'habits', 'cleaning', 'meals', 'timetable'
    """
    # Define constants
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
             'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4']
    
    # If no month specified, use current month
    if month is None:
        current_month = datetime.date.today().month - 1  # 0-based index
        month = months[current_month]
    
    # Define all possible tabs with their configurations at the beginning
    all_tabs = {
        "Dashboard": {
            "headers": [],
            "color": "6f42c1",
            "sample_data": []
        },
        "Income Tracker": {
            "headers": ["Date", "Source", "Amount", "Category", "Notes"],
            "color": "B0E0E6",  # Blue
            "sample_data": [
                [datetime.date.today(), "Salary", 3000, "Primary Income", "Monthly salary"],
                [datetime.date.today().replace(day=15), "Freelance", 1000, "Side Hustle", "Project X"]
            ]
        },
        "Expense Tracker": {
            "headers": ["Date", "Description", "Amount", "Category", "Payment Method", "Notes"],
            "color": "FFB6C1",  # Pink
            "sample_data": [
                [datetime.date.today(), "Groceries", 150, "Food", "Credit Card", "Weekly shopping"],
                [datetime.date.today(), "Electricity", 80, "Utilities", "Direct Debit", "Monthly bill"]
            ]
        },
        "Savings Tracker": {
            "headers": ["Date", "Goal", "Target Amount", "Current Amount", "% Complete"],
            "color": "98FB98",  # Green
            "sample_data": [
                [datetime.date.today(), "Emergency Fund", 10000, 3500, "=D2/C2"]
            ]
        },
        "Stock Tracker": {
            "headers": ["Symbol", "Company", "Shares", "Avg Price", "Current Price", "Total Value", "Gain/Loss", "% Change"],
            "color": "D8BFD8",  # Purple
            "sample_data": [
                ["AAPL", "Apple Inc.", 10, 150, 175, "=C2*E2", "=F2-(C2*D2)", "=(E2-D2)/D2"]
            ]
        },
        "Weight Tracker": {
            "headers": ["Date", "Weight (kg)", "Body Fat %", "Notes"],
            "color": "FFDAB9",  # Orange
            "sample_data": [
                [datetime.date.today() - datetime.timedelta(days=7), 75.5, 22.0, "Started new diet"],
                [datetime.date.today(), 74.8, 21.5, "Feeling good!"]
            ]
        },
        "Habit Tracker": {
            "headers": ["Date", "Exercise", "Water (glasses)", "Sleep (hours)", "Meditation", "Reading", "Notes"],
            "color": "AFEEEE",  # Teal
            "sample_data": [
                [datetime.date.today() - datetime.timedelta(days=1), "✓", 8, 7.5, "✓", "30 min", "Felt great"],
                [datetime.date.today(), "✓", 6, 8.0, "✓", "15 min", "Tired"]
            ]
        },
        "Cleaning Checklist": {
            "headers": ["Task", "Frequency", "Last Done", "Next Due", "Notes"],
            "color": "E6E6FA",  # Lavender
            "sample_data": [
                ["Vacuum", "Weekly", datetime.date.today() - datetime.timedelta(days=2), datetime.date.today() + datetime.timedelta(days=5), "Living room and bedrooms"],
                ["Laundry", "Twice a week", datetime.date.today() - datetime.timedelta(days=1), datetime.date.today() + datetime.timedelta(days=2), "Whites and colors"]
            ]
        },
        "Meal Planner": {
            "headers": ["Day", "Breakfast", "Lunch", "Dinner", "Snacks", "Grocery Items"],
            "color": "FFDAB9",  # Peach (same as orange)
            "sample_data": [
                ["Monday", "Oatmeal", "Salad", "Grilled chicken", "Fruits", "Chicken, salad mix"],
                ["Tuesday", "Smoothie", "Sandwich", "Pasta", "Nuts", "Bread, pasta, nuts"]
            ]
        },
        "Time Table": {
            "headers": ["Time", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
            "color": "98FF98",  # Mint
            "sample_data": [
                ["9:00 AM", "Work", "Work", "Work", "Work", "Work", "Sleep in", "Brunch"],
                ["12:00 PM", "Lunch", "Lunch", "Lunch meeting", "Lunch", "Lunch", "Grocery shopping", "Relax"]
            ]
        }
    }

    wb = openpyxl.Workbook()
    # Remove default sheet
    while len(wb.sheetnames) > 0:
        wb.remove(wb[wb.sheetnames[0]])

    # Color scheme
    colors = {
        'pink': 'FFB6C1',
        'purple': 'D8BFD8',
        'blue': 'B0E0E6',
        'green': '98FB98',
        'yellow': 'FFFACD',
        'orange': 'FFDAB9',
        'teal': 'AFEEEE',
        'lavender': 'E6E6FA',
        'peach': 'FFDAB9',
        'mint': '98FF98'
    }

    # Helper function to create headers
    def create_header(sheet, title, color):
        sheet.merge_cells('A1:J1')
        sheet['A1'] = title
        sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet['A1'].alignment = Alignment(horizontal="center")
        sheet.row_dimensions[1].height = 30
        
        # Add some spacing
        sheet.row_dimensions[2].height = 10

    # Welcome Guide
    welcome = wb.create_sheet("Welcome Guide")
    create_header(welcome, "🌟 Welcome to Your Life & Budget Dashboard 2025 🌟", '6f42c1')
    
    welcome_text = [
        "Thank you for choosing the Life & Budget Dashboard! This comprehensive tool will help you manage:",
        "• Personal Finances (Income, Expenses, Savings, Investments)",
        "• Health & Wellness (Weight, Self-Care, Habits)",
        "• Daily Life (Meal Planning, Cleaning, Schedules)",
        "",
        "📊 HOW TO USE:",
        "1. Start by entering your monthly budget in the 'Dashboard' tab",
        "2. Track your income and expenses in their respective tabs",
        "3. Monitor your investments in the 'Stock Tracker'",
        "4. Use the health and lifestyle trackers to maintain balance",
        "5. Check the 'Dashboard' for insights and progress",
        "",
        "💡 TIP: Use the 'AI Insights' button to get personalized recommendations!",
        "",
        "🔒 Your data stays on your device. For cloud sync, save this file to your preferred cloud storage."
    ]
    
    for i, line in enumerate(welcome_text, start=3):
        cell = welcome.cell(row=i, column=1, value=line)
        if line.startswith("•"):
            cell.font = Font(bold=True)
        elif ":" in line:
            cell.font = Font(bold=True, color='6f42c1')
    
    # Create dashboard sheet - first remove any existing Dashboard sheet
    if 'Dashboard' in [sheet.title for sheet in wb.worksheets]:
        wb.remove(wb['Dashboard'])
    dashboard = wb.create_sheet("Dashboard")
    create_header(dashboard, f"📊 {month} - Financial Overview", '6f42c1')
    
    # Add sample data and charts
    dashboard['A3'] = "Financial Overview"
    dashboard['A3'].font = Font(bold=True, size=12, color='6f42c1')
    
    # Calculate monthly values based on the selected month
    month_index = months.index(month)
    monthly_income = 5000 + (month_index * 100)  # Sample data that increases slightly each month
    monthly_expenses = 3200 + (month_index * 50)  # Sample data that increases slightly each month
    monthly_savings = monthly_income - monthly_expenses
    
    # Calculate weekly breakdown
    weekly_income = [monthly_income * 0.25] * 4
    weekly_expenses = [
        monthly_expenses * 0.2,  # Week 1
        monthly_expenses * 0.3,  # Week 2
        monthly_expenses * 0.25, # Week 3
        monthly_expenses * 0.25  # Week 4
    ]
    weekly_savings = [i - e for i, e in zip(weekly_income, weekly_expenses)]
    
    # Add sample data with monthly focus
    dashboard_data = [
        ["Category", "Planned", "Actual", "Difference", "% of Budget"],
        ["Income", monthly_income, monthly_income * 1.04, "=C4-B4", "=C4/SUM(B4:B8)"],
        ["Expenses", monthly_expenses * 1.1, monthly_expenses, "=C5-B5", "=C5/SUM(B4:B8)"],
        ["Savings", monthly_savings * 1.1, monthly_savings, "=C6-B6", "=C6/SUM(B4:B8)"],
        ["Investments", 500, 550, "=C7-B7", "=C7/SUM(B4:B8)"],
        ["Total", "=SUM(B4:B7)", "=SUM(C4:C7)", "=C8-B8", "=C8/SUM(B4:B8)"]
    ]
    
    for r, row in enumerate(dashboard_data, start=3):
        for c, value in enumerate(row, start=1):
            cell = dashboard.cell(row=r, column=c, value=value)
            if r == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F2E6FF', end_color='F2E6FF', fill_type="solid")
    
    # Create or clear the Charts sheet
    if 'Charts' in wb.sheetnames:
        wb.remove(wb['Charts'])
    charts_sheet = wb.create_sheet('Charts')
    
    # Set up the Charts sheet
    charts_sheet.title = 'Charts'
    charts_sheet.sheet_view.showGridLines = False  # Hide gridlines for cleaner look
    
    # Title and styling
    charts_sheet.merge_cells('A1:J1')
    charts_sheet['A1'] = f"{month} - Financial Overview"
    charts_sheet['A1'].font = Font(size=18, bold=True, color='FFFFFF', name='Calibri')
    charts_sheet['A1'].fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type="solid")
    charts_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    charts_sheet.row_dimensions[1].height = 35
    
    # Add a subtitle
    charts_sheet.merge_cells('A2:J2')
    charts_sheet['A2'] = "Interactive Financial Visualizations"
    charts_sheet['A2'].font = Font(size=14, color='6f42c1', name='Calibri')
    charts_sheet['A2'].alignment = Alignment(horizontal="center")
    charts_sheet.row_dimensions[2].height = 25
    
    # 1. Weekly Financial Trend for the Month (Top-left)
    line_chart = LineChart()
    line_chart.title = f"{month} - Weekly Financial Trend"
    line_chart.style = 12
    line_chart.y_axis.title = 'Amount ($)'
    line_chart.x_axis.title = 'Week'
    line_chart.height = 15
    line_chart.width = 25
    
    # Add some padding for better visualization
    for row in range(5, 10):
        charts_sheet.row_dimensions[row].height = 15
    
    # Write weekly data to a dedicated section
    weekly_data_row = 50
    dashboard.cell(row=weekly_data_row, column=1, value="Week")
    dashboard.cell(row=weekly_data_row, column=2, value="Income")
    dashboard.cell(row=weekly_data_row, column=3, value="Expenses")
    dashboard.cell(row=weekly_data_row, column=4, value="Savings")
    
    for i, week in enumerate(weeks, 1):
        row = weekly_data_row + i
        dashboard.cell(row=row, column=1, value=week)
        dashboard.cell(row=row, column=2, value=weekly_income[i-1])
        dashboard.cell(row=row, column=3, value=weekly_expenses[i-1])
        dashboard.cell(row=row, column=4, value=weekly_savings[i-1])
    
    # Add data to the chart with proper series names
    data = Reference(dashboard, min_col=2, min_row=weekly_data_row-1, max_row=weekly_data_row+len(weeks)-1)
    line_chart.add_data(data, titles_from_data=True)
    
    # Add expenses data
    data = Reference(dashboard, min_col=3, min_row=weekly_data_row-1, max_row=weekly_data_row+len(weeks)-1)
    line_chart.add_data(data, titles_from_data=True)
    
    # Add savings data
    data = Reference(dashboard, min_col=4, min_row=weekly_data_row-1, max_row=weekly_data_row+len(weeks)-1)
    line_chart.add_data(data, titles_from_data=True)
    
    # Set series names explicitly using the correct openpyxl syntax
    from openpyxl.chart.series import SeriesLabel
    line_chart.series[0].title = SeriesLabel(v='Income')
    line_chart.series[1].title = SeriesLabel(v='Expenses')
    line_chart.series[2].title = SeriesLabel(v='Savings')
    
    # Set categories (x-axis)
    cats = Reference(dashboard, min_col=1, min_row=weekly_data_row+1, max_row=weekly_data_row+len(weeks))
    line_chart.set_categories(cats)
    
    # Customize the series with better colors
    line_chart.series[0].graphicalProperties.line.solidFill = "4CAF50"  # Green for income
    line_chart.series[1].graphicalProperties.line.solidFill = "F44336"  # Red for expenses
    line_chart.series[2].graphicalProperties.line.solidFill = "2196F3"  # Blue for savings
    
    # Add data labels to all points
    for series in line_chart.series:
        series.dLbls = openpyxl.chart.label.DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showSeriesName = True
    
    charts_sheet.add_chart(line_chart, "A3")
    
    # 2. Planned vs Actual Bar Chart (Top-right)
    chart1 = BarChart()
    data = Reference(dashboard, min_col=2, max_col=3, min_row=3, max_row=6)  # Exclude Total row
    cats = Reference(dashboard, min_col=1, min_row=4, max_row=6)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.title = "Planned vs Actual"
    chart1.style = 12
    chart1.y_axis.title = 'Amount ($)'
    chart1.x_axis.title = 'Category'
    chart1.height = 15
    chart1.width = 25
    
    # Customize colors
    chart1.series[0].graphicalProperties.solidFill = "4CAF50"  # Green for Planned
    chart1.series[1].graphicalProperties.solidFill = "2196F3"  # Blue for Actual
    
    # Add data labels
    chart1.dLbls = openpyxl.chart.label.DataLabelList()
    chart1.dLbls.showVal = True
    
    charts_sheet.add_chart(chart1, "L3")
    
    # Add sample expense breakdown for visualization with better formatting
    expense_data = [
        ["Category", "Amount", "Target", "Variance", "% of Total"],
        ["Housing", 1200, 1100, "=B2-C2", "=B2/SUM($B$2:$B$7)"],
        ["Food", 800, 700, "=B3-C3", "=B3/SUM($B$2:$B$7)"],
        ["Transportation", 400, 450, "=B4-C4", "=B4/SUM($B$2:$B$7)"],
        ["Utilities", 300, 350, "=B5-C5", "=B5/SUM($B$2:$B$7)"],
        ["Entertainment", 300, 250, "=B6-C6", "=B6/SUM($B$2:$B$7)"],
        ["Others", 200, 200, "=B7-C7", "=B7/SUM($B$2:$B$7)"],
        ["Total", "=SUM(B2:B7)", "=SUM(C2:C7)", "=SUM(D2:D7)", "=SUM(E2:E7)"]
    ]
    
    # Write expense breakdown data with formatting
    for r, row in enumerate(expense_data, start=25):
        for c, value in enumerate(row, start=1):
            cell = dashboard.cell(row=r, column=c, value=value)
            if r == 25:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
                cell.font = Font(color='FFFFFF')
            elif r == len(expense_data) + 24:  # Total row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F2E6FF', end_color='F2E6FF', fill_type='solid')
    
    # 3. Expense Breakdown Pie Chart (Bottom-left) - 3D with percentage and legend
    expense_pie = PieChart3D()
    data = Reference(dashboard, min_col=2, min_row=26, max_row=31)
    labels = Reference(dashboard, min_col=1, min_row=26, max_row=31)
    expense_pie.add_data(data, titles_from_data=False)
    expense_pie.set_categories(labels)
    expense_pie.title = f"{month} - Expenses Breakdown"
    expense_pie.height = 15
    expense_pie.width = 25
    
    # Add percentage data labels
    expense_pie.dataLabels = openpyxl.chart.label.DataLabelList()
    expense_pie.dataLabels.showPercent = True
    expense_pie.dataLabels.showVal = False
    expense_pie.dataLabels.showCatName = True
    expense_pie.legend.position = 'b'  # Bottom legend
    
    # Add some 3D rotation for better visibility
    expense_pie.rotX = 30
    expense_pie.rotY = 30
    
    charts_sheet.add_chart(expense_pie, "A25")
    
    # 4. Budget Distribution Donut Chart (Bottom-right)
    pie = PieChart3D()
    data = Reference(dashboard, min_col=2, min_row=4, max_row=6)
    labels = Reference(dashboard, min_col=1, min_row=4, max_row=6)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = f"{month} - Budget Distribution"
    pie.height = 15
    pie.width = 25
    
    # Make it a donut chart by setting a hole size
    pie.holeSize = 30  # Percentage of hole size
    
    # Add percentage data labels
    pie.dataLabels = openpyxl.chart.label.DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = True
    pie.legend.position = 'r'  # Right legend
    
    # Custom colors
    colors = ['4CAF50', '2196F3', 'FFC107']  # Green, Blue, Amber
    for i, series in enumerate(pie.series):
        series.graphicalProperties.solidFill = colors[i % len(colors)]
    
    # Add 3D rotation
    pie.rotX = 15
    pie.rotY = 30
    
    charts_sheet.add_chart(pie, "L25")
    
    # Weekly breakdown data is already calculated at the beginning of the function
    
    # Sample data for monthly trends
    income_data = [5000 + (i * 100) for i in range(12)]  # Increasing income
    expense_data = [3200 + (i * 50) for i in range(12)]  # Increasing expenses
    savings_data = [i - e for i, e in zip(income_data, expense_data)]  # Calculate savings
    
    # Write time series data with better formatting
    dashboard['A45'] = "Monthly Financial Trend"
    dashboard['A45'].font = Font(bold=True, size=12, color='6f42c1', name='Calibri')
    
    headers = ["Month", "Income", "Expenses", "Savings"]
    for col, header in enumerate(headers, 1):
        cell = dashboard.cell(row=46, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
    
    for i, (month, inc, exp, sav) in enumerate(zip(months, income_data, expense_data, savings_data), start=1):
        row = 46 + i
        dashboard.cell(row=row, column=1, value=month)
        
        # Format as currency
        for col, val in zip([2, 3, 4], [inc, exp, sav]):
            cell = dashboard.cell(row=row, column=col, value=val)
            cell.number_format = '$#,##0'
            
            # Add conditional formatting for negative savings
            if col == 4 and val < 0:
                cell.font = Font(color='FF0000')  # Red for negative savings
    
    # Auto-size columns for better visibility
    for column in dashboard.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        dashboard.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    # Define the exact order we want the sheets to appear in
    sheet_order = [
        'Welcome Guide',
        'Dashboard',
        'Charts',
        'Income Tracker',
        'Expense Tracker',
        'Savings Tracker',
        'Stock Tracker',
        'Weight Tracker',
        'Habit Tracker',
        'Cleaning Checklist',
        'Meal Planner',
        'Time Table',
        'AI Insights'
    ]
    
    # Map UI section names to tab names
    section_map = {
        'income': "Income Tracker",
        'expenses': "Expense Tracker",
        'savings': "Savings Tracker",
        'stocks': "Stock Tracker",
        'weight': "Weight Tracker",
        'habits': "Habit Tracker",
        'cleaning': "Cleaning Checklist",
        'meals': "Meal Planner",
        'timetable': "Time Table"
    }
    
    # Initialize tabs with core sheets
    tabs = {
        "Dashboard": all_tabs.get("Dashboard", {}),
        "Welcome Guide": all_tabs.get("Welcome Guide", {}),
        "AI Insights": all_tabs.get("AI Insights", {})
    }
    
    # If no sections specified, include all
    if not sections:
        sections = section_map.keys()
    
    # Add selected tabs to the tabs dictionary
    for section in sections:
        if section in section_map:
            tab_name = section_map[section]
            if tab_name in all_tabs:
                tabs[tab_name] = all_tabs[tab_name]
    
    # Process each tab in the desired order
    for tab_name in sheet_order:
        # Handle core sheets
        if tab_name in ['Welcome Guide', 'Dashboard', 'AI Insights']:
            tab_data = all_tabs.get(tab_name, {})
            if tab_name not in wb.sheetnames:
                sheet = wb.create_sheet(tab_name)
            else:
                sheet = wb[tab_name]
                sheet.delete_rows(1, sheet.max_row)
            
            if tab_name == 'Welcome Guide':
                create_welcome_guide(sheet)
            elif tab_name == 'AI Insights':
                create_header(sheet, "🤖 AI-Powered Financial Insights", '6f42c1')
                ai_text = [
                    "This section provides AI-generated insights based on your data:",
                    "",
                    "🔍 To get started:",
                    "1. Fill in your financial data in the respective sheets",
                    "2. Click the 'Generate AI Insights' button in the app",
                    "3. Review personalized recommendations here",
                    "",
                    "The AI will analyze your spending patterns, saving habits, and overall financial health to provide actionable advice.",
                    "",
                    "💡 Tip: The more data you provide, the more accurate the insights will be!"
                ]
                for i, line in enumerate(ai_text, start=3):
                    cell = sheet.cell(row=i, column=1, value=line)
                    if line.startswith("🔍") or line.startswith("💡"):
                        cell.font = Font(bold=True, color='6f42c1')
            continue
            
        # Skip Charts sheet for now (handled separately)
        if tab_name == 'Charts':
            continue
            
        # Skip if not in selected tabs
        if tab_name not in tabs:
            continue
            
        # Remove existing sheet if it exists (except for core sheets)
        if tab_name in wb.sheetnames and tab_name not in ['Welcome Guide', 'Dashboard', 'Charts', 'AI Insights']:
            try:
                wb.remove(wb[tab_name])
            except Exception as e:
                print(f"Warning: Could not remove sheet {tab_name}: {str(e)}")
            
        # Get the tab data (skip for AI Insights)
        tab_data = all_tabs.get(tab_name, {})
        
        # Create the sheet if it doesn't exist
        if tab_name not in wb.sheetnames:
            sheet = wb.create_sheet(tab_name)
        else:
            sheet = wb[tab_name]
            # Clear existing content but keep formatting
            sheet.delete_rows(1, sheet.max_row)
        
        # Handle Welcome Guide specially
        if tab_name == 'Welcome Guide':
            create_welcome_guide(sheet)
            continue
            
        # Handle AI Insights specially
        if tab_name == 'AI Insights':
            create_header(sheet, "🤖 AI-Powered Financial Insights", '6f42c1')
            ai_text = [
                "This section provides AI-generated insights based on your data:",
                "",
                "🔍 To get started:",
                "1. Fill in your financial data in the respective sheets",
                "2. Click the 'Generate AI Insights' button in the app",
                "3. Review personalized recommendations here",
                "",
                "The AI will analyze your spending patterns, saving habits, and overall financial health to provide actionable advice.",
                "",
                "💡 Tip: The more data you provide, the more accurate the insights will be!"
            ]
            
            for i, line in enumerate(ai_text, start=3):
                cell = sheet.cell(row=i, column=1, value=line)
                if line.startswith("🔍") or line.startswith("💡"):
                    cell.font = Font(bold=True, color='6f42c1')
            continue
            
        # For all other sheets, add the standard content
        title = f"{month} {tab_name}" if tab_name == 'Dashboard' else tab_name
        create_header(sheet, title, tab_data.get("color", "6f42c1"))
        
        # Add column headers
        headers = tab_data.get("headers", [])
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(
                start_color=tab_data.get("color", "6f42c1"), 
                end_color=tab_data.get("color", "6f42c1"), 
                fill_type='solid'
            )
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
        
        # Add sample data if available
        if "sample_data" in tab_data and tab_data["sample_data"]:
            try:
                for row_num, row_data in enumerate(tab_data["sample_data"], start=4):
                    for col_num, cell_value in enumerate(row_data, start=1):
                        try:
                            cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                            if isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
                                cell.number_format = '$#,##0.00'
                        except Exception as e:
                            print(f"Error setting cell value at row {row_num}, col {col_num}: {str(e)}")
            except Exception as e:
                print(f"Error adding sample data to {tab_name}: {str(e)}")
        
        # Auto-size columns for the current sheet
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    # Reorder sheets to match our desired order (in reverse to get correct positioning)
    for sheet_name in reversed(sheet_order):
        if sheet_name in wb.sheetnames:
            wb.move_sheet(wb[sheet_name], 0)
    
    # Set the active sheet to Dashboard or Welcome Guide
    if 'Dashboard' in wb.sheetnames:
        wb.active = wb['Dashboard']
    elif 'Welcome Guide' in wb.sheetnames:
        wb.active = wb['Welcome Guide']
    
    # Ensure all sheets are visible
    for sheet in wb.worksheets:
        sheet.sheet_state = 'visible'
    
    # Section mapping is now at the beginning of the function
    
    return wb

def read_excel_data_optimized(file_path, selected_categories=None):
    """Optimized function to read only the necessary data from Excel."""
    try:
        xls = pd.ExcelFile(file_path)
        data = {}
        
        # Define which sheets to process based on selected categories
        sheets_to_process = []
        if selected_categories:
            # Map categories to their sheet names
            category_map = {
                'Income': 'Income Tracker',
                'Expenses': 'Expense Tracker',
                'Savings': 'Savings Tracker',
                'Investments': 'Stock Tracker',
                'Health': ['Weight Tracker', 'Self-Care Tracker', 'Habit Tracker'],
                'Lifestyle': ['Cleaning Checklist', 'Meal Planner', 'Time Table']
            }
            
            for category in selected_categories:
                if category in category_map:
                    sheets = category_map[category]
                    if isinstance(sheets, list):
                        sheets_to_process.extend(sheets)
                    else:
                        sheets_to_process.append(sheets)
        else:
            # If no categories selected, process all sheets except AI Insights
            sheets_to_process = [sheet for sheet in xls.sheet_names if sheet != 'AI Insights']
        
        # Process only the selected sheets
        for sheet_name in sheets_to_process:
            if sheet_name in xls.sheet_names:
                try:
                    # Read only the first 100 rows to avoid processing too much data
                    df = pd.read_excel(xls, sheet_name=sheet_name, nrows=100)
                    if not df.empty:
                        # Convert to string and clean up
                        data[sheet_name] = df.head(20).to_string()  # Limit to first 20 rows
                except Exception as e:
                    print(f"Error processing sheet {sheet_name}: {str(e)}")
        
        return data
    except Exception as e:
        return {"error": f"Error reading Excel file: {str(e)}"}

def create_download_link(val, filename):
    """Generates a link to download a file."""
    b64 = base64.b64encode(val)  # val looks like b'...'
    return f'<a href="data:application/octet-stream;base64,{b64.decode()}" download="{filename}">Download {filename}</a>'

def generate_pdf(insights_text, filename="financial_insights.pdf"):
    """Generate a PDF from the insights text."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    # Add title
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, txt="Financial Insights Report", ln=True, align='C')
    pdf.ln(10)
    
    # Add date
    pdf.set_font("Arial", 'I', 10)
    pdf.cell(200, 10, txt=f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
    pdf.ln(10)
    
    # Add insights
    pdf.set_font("Arial", size=12)
    for line in insights_text.split('\n'):
        if line.strip().startswith(('📈', '💡', '💰', '🌱', '⚠️', '🎯', '🔧')):
            pdf.set_font('', 'B')
            pdf.cell(200, 10, txt=line, ln=True)
            pdf.set_font('', '')
        else:
            pdf.multi_cell(0, 10, txt=line)
    
    # Save to temporary file
    temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdf_path = temp_pdf.name
    pdf.output(pdf_path)
    return pdf_path

def generate_ai_insights(file_path, selected_categories=None):
    """Generate AI insights using Ollama.
    
    Args:
        file_path (str): Path to the Excel file
        selected_categories (list, optional): List of categories to analyze. Defaults to None (all categories).
    """
    try:
        # First, check if Ollama is running
        try:
            models = ollama.list()
            if not models.get('models'):
                raise Exception("No models available. Please install a model with 'ollama pull gemma3:4b'")
        except Exception as e:
            return f"❌ Error connecting to Ollama: {str(e)}\n\nPlease ensure Ollama is running and you have at least one model installed.\n\nTo install the recommended model, run: ollama pull gemma3:4b"
        
        # Read data from Excel
        excel_data = read_excel_data_optimized(file_path, selected_categories)
        if isinstance(excel_data, dict) and 'error' in excel_data:
            return excel_data['error']
        
        # Prepare the prompt for the AI
        categories_note = f"\n\nNote: The user is specifically interested in these categories: {', '.join(selected_categories)}. Focus your analysis on these areas." if selected_categories else ""
        
        prompt = f"""You are a financial advisor analyzing personal finance data. 
        Please provide a detailed analysis including:
        1. Financial health overview (savings rate, debt-to-income, etc.)
        2. Spending patterns and potential savings opportunities
        3. Progress towards financial goals
        4. Any concerning trends or anomalies
        5. 3-5 specific, actionable recommendations
        
        Format your response with clear sections and use emojis for better readability.
        Be specific and reference the actual numbers from the data when possible.
        {categories_note}
        
        Here's the data to analyze:
        """
        
        # Limit the amount of data we send to the model
        total_chars = 0
        max_chars = 4000  # Limit total prompt size to ~4k characters
        
        # Add data from each sheet to the prompt, but limit total size
        for sheet_name, content in excel_data.items():
            # Take first 3 lines of each sheet's content
            sheet_summary = '\n'.join(str(content).split('\n')[:3])
            sheet_data = f"\n\n--- {sheet_name} Data ---\n{sheet_summary}..."
            
            # Check if we can add this data without exceeding max_chars
            if total_chars + len(sheet_data) > max_chars:
                break
                
            prompt += sheet_data
            total_chars += len(sheet_data)
        
        # Show progress to user
        progress_text = st.empty()
        progress_text.text("🤖 Analyzing your financial data...")
        
        # Generate response using Ollama with optimized parameters
        response = ollama.generate(
            model='gemma3:4b',
            prompt=prompt,
            options={
                'temperature': 0.5,  # Lower temperature for more focused responses
                'max_tokens': 1000,   # Limit response length
                'num_ctx': 2048,      # Smaller context window
                'num_predict': 500,   # Limit prediction length
                'top_k': 40,          # Limit the sampling pool
                'top_p': 0.9,         # Slightly more focused than default
                'repeat_penalty': 1.1  # Discourage repetition
            }
        )
        
        # Clear progress text
        progress_text.empty()
        
        return response['response']
        
    except Exception as e:
        return f"❌ Error generating insights: {str(e)}\n\nPlease ensure Ollama is running and you have a model installed.\n\nTo install the recommended model, run: ollama pull gemma3:4b"

def main():
    st.title("📊 Life & Budget Dashboard")
    st.markdown("### Your All-in-One Financial and Personal Management Tool")
    
    # Sidebar for navigation
    st.sidebar.title("Navigation")
    page = st.sidebar.radio("Go to", ["Home", "Generate Template", "AI Insights"])
    
    if page == "Home":
        st.markdown("""
        Welcome to your personal Life & Budget Dashboard! This tool helps you track your finances, 
        health, and daily activities all in one place.
        
        ### Features:
        - **Financial Tracking**: Income, expenses, savings, and investments
        - **Health & Wellness**: Weight tracking and habit formation
        - **Life Organization**: Cleaning schedules and meal planning
        
        Get started by generating a new template or upload your existing data for AI-powered insights.
        """)
        
    elif page == "Generate Template":
        st.header("📝 Generate New Template")
        st.write("Create a new Excel template with the sections you need.")
        
        # Month selection
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        selected_month = st.selectbox("Select Month", months, index=datetime.date.today().month - 1)
        
        # Section selection
        st.subheader("Select Sections to Include")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            financial = st.checkbox("Financial Tracking", True)
            if financial:
                income = st.checkbox("Income Tracker", True)
                expenses = st.checkbox("Expense Tracker", True)
                savings = st.checkbox("Savings Tracker", True)
                stocks = st.checkbox("Stock Tracker", True)
        
        with col2:
            health = st.checkbox("Health & Wellness", True)
            if health:
                weight = st.checkbox("Weight Tracker", True)
                habits = st.checkbox("Habit Tracker", True)
        
        with col3:
            life = st.checkbox("Life Organization", True)
            if life:
                cleaning = st.checkbox("Cleaning Checklist", True)
                meals = st.checkbox("Meal Planner", True)
                timetable = st.checkbox("Weekly Schedule", True)
        
        # Create a list of selected sections
        selected_sections = []
        
        # Financial section
        if financial:
            if income: selected_sections.append("income")
            if expenses: selected_sections.append("expenses")
            if savings: selected_sections.append("savings")
            if stocks: selected_sections.append("stocks")
        
        # Health section
        if health:
            if weight: selected_sections.append("weight")
            if habits: selected_sections.append("habits")
        
        # Life organization section
        if life:
            if cleaning: selected_sections.append("cleaning")
            if meals: selected_sections.append("meals")
            if timetable: selected_sections.append("timetable")
        
        # If no sections are selected, include all
        if not selected_sections and not financial and not health and not life:
            selected_sections = [
                'income', 'expenses', 'savings', 'stocks',
                'weight', 'habits', 'cleaning', 'meals', 'timetable'
            ]
        
        if st.button("✨ Generate Template"):
            with st.spinner(f"Creating your {selected_month} template..."):
                # Create a temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    file_path = tmp.name
                
                try:
                    # Create the template with the selected month and sections
                    wb = create_excel_template(month=selected_month, sections=selected_sections)
                    
                    # Debug: Print the sheets that were created
                    st.sidebar.write("Sheets created:", wb.sheetnames)
                    
                except Exception as e:
                    st.error(f"Error creating Excel template: {str(e)}")
                    st.stop()
                wb.save(file_path)
                
                # Create download link
                with open(file_path, 'rb') as f:
                    bytes_data = f.read()
                
                st.success(f"{selected_month} template created successfully!")
                st.download_button(
                    label=f"📥 Download {selected_month} Template",
                    data=bytes_data,
                    file_name=f"life_budget_tracker_{selected_month}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Clean up
                os.unlink(file_path)
                
    elif page == "AI Insights":
        st.header("📊 Upload Your Data for AI Insights")
        
        # Add category selection
        categories = [
            'Income',
            'Expenses',
            'Savings',
            'Investments',
            'Health',
            'Lifestyle'
        ]
        
        selected_categories = st.multiselect(
            "Select categories to analyze:",
            categories,
            default=categories[:2]  # Default to Income and Expenses
        )
        
        uploaded_file = st.file_uploader("Upload your filled Excel file", type=["xlsx"])
        
        if uploaded_file is not None:
            if st.button("🤖 Generate AI Insights", key="ai_insights_btn"):
                with st.spinner("Analyzing your data with AI..."):
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        tmp.write(uploaded_file.getvalue())
                        tmp_path = tmp.name
                    
                    try:
                        insights = generate_ai_insights(tmp_path, selected_categories)
                        
                        if not insights.startswith(('❌', '⚠️')):
                            st.success("AI Analysis Complete!")
                            st.markdown("### 🎯 Your Personalized Insights")
                            
                            with st.expander("View Insights", expanded=True):
                                st.markdown(insights)
                            
                            # Download buttons
                            col1, col2 = st.columns(2)
                            with col1:
                                st.download_button(
                                    label="📝 Download as Text",
                                    data=insights,
                                    file_name="financial_insights.txt",
                                    mime="text/plain"
                                )
                            with col2:
                                pdf_path = generate_pdf(insights)
                                with open(pdf_path, "rb") as f:
                                    st.download_button(
                                        label="📄 Download as PDF",
                                        data=f,
                                        file_name="financial_insights.pdf",
                                        mime="application/pdf"
                                    )
                                os.unlink(pdf_path)
                        else:
                            st.error(insights)
                    finally:
                        if os.path.exists(tmp_path):
                            os.unlink(tmp_path)
    
    st.markdown("---")
    st.markdown("### 📱 Features at a Glance")
    
    features = st.columns(3)
    
    with features[0]:
        st.markdown("""
        #### 💰 Financial Tracking
        - Income & Expense tracking
        - Budget planning
        - Savings goals
        - Investment portfolio
        """)
    
    with features[1]:
        st.markdown("""
        #### 🏋️ Health & Wellness
        - Weight tracking
        - Habit formation
        - Self-care routines
        - Meal planning
        """)
    
    with features[2]:
        st.markdown("""
        #### 🏠 Life Organization
        - Cleaning checklist
        - Weekly schedule
        - Time management
        - Task tracking
        """)
    
    st.markdown("---")
    st.markdown("""
    ### Getting Started
    1. Download the Excel template above
    2. Fill in your financial and personal data
    3. Upload it back to get AI-powered insights
    4. Use the dashboard to track your progress
    """)

if __name__ == "__main__":
    main()