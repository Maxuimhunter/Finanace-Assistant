# app.py
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, PieChart3D
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import os
import datetime
import json
import ollama
import time
from io import BytesIO
from fpdf import FPDF
import base64
import io
from pathlib import Path
import re

# Set page config
st.set_page_config(
    page_title="Life & Budget Dashboard",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for styling
st.markdown("""
<style>
    .main {
        background-color: #f8f9fa;
    }
    .stButton>button {
        background-color: #6f42c1;
        color: white;
        border-radius: 8px;
        padding: 0.5rem 1rem;
    }
    .stButton>button:hover {
        background-color: #5a32a3;
        color: white;
    }
    .stTextInput>div>div>input {
        border: 1px solid #6f42c1;
        border-radius: 4px;
    }
    .css-1d391kg {
        padding-top: 3rem;
    }
</style>
""", unsafe_allow_html=True)

def create_welcome_guide(sheet):
    """Create a visually appealing Welcome Guide sheet."""
    # Set background color and hide gridlines
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = '6f42c1' # Purple

    # Set column widths for a better layout
    sheet.column_dimensions['A'].width = 4
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 60
    sheet.column_dimensions['D'].width = 4

    # Main Title
    sheet.merge_cells('B2:C2')
    title_cell = sheet['B2']
    title_cell.value = "🚀 Welcome to Your All-in-One Life Dashboard"
    title_cell.font = Font(size=24, bold=True, color='FFFFFF', name='Calibri')
    title_cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[2].height = 40

    # Subtitle
    sheet.merge_cells('B4:C4')
    subtitle_cell = sheet['B4']
    subtitle_cell.value = "Track finances, build habits, and organize your life with ease."
    subtitle_cell.font = Font(size=14, italic=True, color='595959', name='Calibri')
    subtitle_cell.alignment = Alignment(horizontal='center')
    sheet.row_dimensions[4].height = 25

    # Section: Core Features
    sheet.merge_cells('B6:C6')
    features_title = sheet['B6']
    features_title.value = "Core Features"
    features_title.font = Font(size=16, bold=True, color='6f42c1', name='Calibri')
    features_title.alignment = Alignment(horizontal='left')
    sheet.row_dimensions[6].height = 30

    features = [
        ("📊 Dashboard", "A high-level overview of your financial health and key life metrics."),
        ("💰 Financial Trackers", "Log income, expenses, savings, and investments in dedicated tabs."),
        ("💚 Health & Wellness", "Monitor weight, build habits, and focus on self-care."),
        ("📅 Life Organization", "Plan meals, manage cleaning schedules, and organize your time."),
        ("🤖 AI Insights", "Upload your sheet to get personalized financial recommendations.")
    ]

    thin_border = Border(left=Side(style='thin', color='DDDDDD'), 
                       right=Side(style='thin', color='DDDDDD'), 
                       top=Side(style='thin', color='DDDDDD'), 
                       bottom=Side(style='thin', color='DDDDDD'))

    for i, (title, desc) in enumerate(features, start=7):
        sheet.row_dimensions[i].height = 25
        # Feature Title
        title_cell = sheet.cell(row=i, column=2, value=title)
        title_cell.font = Font(size=12, bold=True, name='Calibri')
        title_cell.alignment = Alignment(vertical='center')
        # Feature Description
        desc_cell = sheet.cell(row=i, column=3, value=desc)
        desc_cell.font = Font(size=12, name='Calibri')
        desc_cell.alignment = Alignment(vertical='center', wrap_text=True)

    # Section: Getting Started
    start_row = len(features) + 9
    sheet.merge_cells(f'B{start_row}:C{start_row}')
    getting_started_title = sheet[f'B{start_row}']
    getting_started_title.value = "How to Get Started"
    getting_started_title.font = Font(size=16, bold=True, color='6f42c1', name='Calibri')
    sheet.row_dimensions[start_row].height = 30

    instructions = [
        ("1.", "Explore each tab to see what's available."),
        ("2.", "Fill in your data in the respective trackers (income, expenses, etc.)."),
        ("3.", "The 'Dashboard' and 'Charts' tabs will update automatically."),
        ("4.", "For AI analysis, save the file and upload it in the web app.")
    ]

    for i, (num, text) in enumerate(instructions, start=start_row + 1):
        sheet.row_dimensions[i].height = 25
        sheet.cell(row=i, column=2, value=num).font = Font(size=12, bold=True, name='Calibri')
        sheet.cell(row=i, column=3, value=text).font = Font(size=12, name='Calibri')

    # Footer
    footer_row = start_row + len(instructions) + 3
    sheet.merge_cells(f'B{footer_row}:C{footer_row}')
    footer_cell = sheet[f'B{footer_row}']
    footer_cell.value = "Happy tracking! ✨"
    footer_cell.font = Font(size=12, italic=True, color='595959', name='Calibri')
    footer_cell.alignment = Alignment(horizontal='center')


def worksheet_to_dataframe(sheet):
    """Converts an openpyxl worksheet to a pandas DataFrame.
    
    This function is designed to be robust and work with various Excel formats.
    It will:
    1. Find the first row that looks like a header
    2. Read all subsequent rows as data
    3. Clean up empty rows and columns
    4. Preserve all data (no aggressive filtering)
    """
    try:
        # Get all rows with values
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            return pd.DataFrame()

        # 1. Find the header row by looking for common header keywords
        header_row_index = -1
        for i, row in enumerate(all_rows):
            if not any(cell is None for cell in row[:5]):  # First 5 columns shouldn't be empty in a header
                header_candidates = [str(cell).lower() for cell in row if cell is not None]
                header_keywords = ['date', 'amount', 'source', 'description', 'category', 'type', 'name', 'item']
                if any(any(kw in str(cell).lower() for kw in header_keywords) for cell in row if cell):
                    header_row_index = i
                    break
        
        # If no header found, try to use the first non-empty row as header
        if header_row_index == -1:
            for i, row in enumerate(all_rows):
                if any(cell is not None for cell in row):
                    header_row_index = i
                    break
            
            if header_row_index == -1:
                return pd.DataFrame()  # No data found

        # 2. Get header and data
        header = [str(cell) if cell is not None else f'Column_{i+1}' 
                 for i, cell in enumerate(all_rows[header_row_index])]
        data = all_rows[header_row_index + 1:]

        # 3. Create DataFrame with proper handling of empty cells
        df = pd.DataFrame(data, columns=header)
        
        # 4. Clean the DataFrame
        # Drop completely empty rows and columns
        df = df.dropna(how='all')
        df = df.dropna(how='all', axis=1)
        
        # Convert all columns to string and strip whitespace
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.strip()
        
        # Remove completely empty rows again after type conversion
        df = df[df.astype(str).ne('None').any(axis=1)]
        df = df[df.astype(str).ne('').any(axis=1)]
        
        # Reset index after all filtering
        df.reset_index(drop=True, inplace=True)
        
        return df
        
    except Exception as e:
        print(f"Error processing sheet {sheet.title}: {str(e)}")
        return pd.DataFrame()

def is_merged_cell(sheet, row, col):
    """Check if a cell is part of a merged range."""
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and 
            merged_range.min_col <= col <= merged_range.max_col):
            return True
    return False

def create_excel_template(month=None, sections=None):
    """
    Create an Excel template for a specific month with the specified sections.
    
    Args:
        month (str, optional): The month to generate the template for (e.g., 'Jan', 'Feb'). 
                              If None, uses current month.
        sections (list, optional): List of section names to include. If None, includes all sections.
                                  Possible values: 'dashboard', 'income', 'expenses', 'savings', 'stocks', 
                                  'weight', 'habits', 'cleaning', 'meals', 'timetable'
    """
    # Define constants
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
             'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    weeks = ['Week 1', 'Week 2', 'Week 3', 'Week 4']
    
    # If no month specified, use current month
    if month is None:
        current_month = datetime.date.today().month - 1  # 0-based index
        month = months[current_month]
    
    # Define all possible tabs with their configurations at the beginning
    all_tabs = {
        "Income Tracker": {
            "headers": ["Date", "Source", "Amount", "Category", "Notes"],
            "color": "B0E0E6",  # Blue
            "sample_data": [
                [datetime.date.today(), "Salary", 3000, "Primary Income", "Monthly salary"],
                [datetime.date.today().replace(day=15), "Freelance", 1000, "Side Hustle", "Project X"]
            ]
        },
        "Expense Tracker": {
            "headers": ["Date", "Description", "Amount", "Category", "Payment Method", "Notes"],
            "color": "FFB6C1",  # Pink
            "sample_data": [
                [datetime.date.today(), "Groceries", 150, "Food", "Credit Card", "Weekly shopping"],
                [datetime.date.today(), "Electricity", 80, "Utilities", "Direct Debit", "Monthly bill"]
            ]
        },
        "Savings Tracker": {
            "headers": ["Date", "Goal", "Target Amount", "Current Amount", "% Complete"],
            "color": "98FB98",  # Green
            "sample_data": [
                [datetime.date.today(), "Emergency Fund", 10000, 3500, "=D2/C2"]
            ]
        },
        "Stock Tracker": {
            "headers": ["Symbol", "Company", "Shares", "Avg Price", "Current Price", "Total Value", "Gain/Loss", "% Change"],
            "color": "D8BFD8",  # Purple
            "sample_data": [
                ["AAPL", "Apple Inc.", 10, 150, 175, "=C2*E2", "=F2-(C2*D2)", "=(E2-D2)/D2"]
            ]
        },
        "Weight Tracker": {
            "headers": ["Date", "Weight (kg)", "Body Fat %", "Notes"],
            "color": "FFDAB9",  # Orange
            "sample_data": [
                [datetime.date.today() - datetime.timedelta(days=7), 75.5, 22.0, "Started new diet"],
                [datetime.date.today(), 74.8, 21.5, "Feeling good!"]
            ]
        },
        "Habit Tracker": {
            "headers": ["Date", "Exercise", "Water (glasses)", "Sleep (hours)", "Meditation", "Reading", "Notes"],
            "color": "AFEEEE",  # Teal
            "sample_data": [
                [datetime.date.today() - datetime.timedelta(days=1), "✓", 8, 7.5, "✓", "30 min", "Felt great"],
                [datetime.date.today(), "✓", 6, 8.0, "✓", "15 min", "Tired"]
            ]
        },
        "Cleaning Checklist": {
            "headers": ["Task", "Frequency", "Last Done", "Next Due", "Notes"],
            "color": "E6E6FA",  # Lavender
            "sample_data": [
                ["Vacuum", "Weekly", datetime.date.today() - datetime.timedelta(days=2), datetime.date.today() + datetime.timedelta(days=5), "Living room and bedrooms"],
                ["Laundry", "Twice a week", datetime.date.today() - datetime.timedelta(days=1), datetime.date.today() + datetime.timedelta(days=2), "Whites and colors"]
            ]
        },
        "Meal Planner": {
            "headers": ["Day", "Breakfast", "Lunch", "Dinner", "Snacks", "Grocery Items"],
            "color": "FFDAB9",  # Peach (same as orange)
            "sample_data": [
                ["Monday", "Oatmeal", "Salad", "Grilled chicken", "Fruits", "Chicken, salad mix"],
                ["Tuesday", "Smoothie", "Sandwich", "Pasta", "Nuts", "Bread, pasta, nuts"]
            ]
        },
        "Time Table": {
            "headers": ["Time", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
            "color": "98FF98",  # Mint
            "sample_data": [
                ["9:00 AM", "Work", "Work", "Work", "Work", "Work", "Sleep in", "Brunch"],
                ["12:00 PM", "Lunch", "Lunch", "Lunch meeting", "Lunch", "Lunch", "Grocery shopping", "Relax"]
            ]
        }
    }

    wb = openpyxl.Workbook()
    # Remove default sheet
    while len(wb.sheetnames) > 0:
        wb.remove(wb[wb.sheetnames[0]])

    # Color scheme
    colors = {
        'pink': 'FFB6C1',
        'purple': 'D8BFD8',
        'blue': 'B0E0E6',
        'green': '98FB98',
        'yellow': 'FFFACD',
        'orange': 'FFDAB9',
        'teal': 'AFEEEE',
        'lavender': 'E6E6FA',
        'peach': 'FFDAB9',
        'mint': '98FF98'
    }

    # Helper function to create headers
    def create_header(sheet, title, color):
        sheet.merge_cells('A1:J1')
        sheet['A1'] = title
        sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet['A1'].alignment = Alignment(horizontal="center")
        sheet.row_dimensions[1].height = 30
        
        # Add some spacing
        sheet.row_dimensions[2].height = 10

    # Welcome Guide
    welcome = wb.create_sheet("Welcome Guide")
    create_header(welcome, "🌟 Welcome to Your Life & Budget Dashboard 2025 🌟", '6f42c1')
    
    welcome_text = [
        "Thank you for choosing the Life & Budget Dashboard! This comprehensive tool will help you manage:",
        "• Personal Finances (Income, Expenses, Savings, Investments)",
        "• Health & Wellness (Weight, Self-Care, Habits)",
        "• Daily Life (Meal Planning, Cleaning, Schedules)",
        "",
        "📊 HOW TO USE:",
        "1. Start by entering your monthly budget in the 'Dashboard' tab",
        "2. Track your income and expenses in their respective tabs",
        "3. Monitor your investments in the 'Stock Tracker'",
        "4. Use the health and lifestyle trackers to maintain balance",
        "5. Check the 'Dashboard' for insights and progress",
        "",
        "💡 TIP: Use the 'AI Insights' button to get personalized recommendations!",
        "",
        "🔒 Your data stays on your device. For cloud sync, save this file to your preferred cloud storage."
    ]
    
    for i, line in enumerate(welcome_text, start=3):
        cell = welcome.cell(row=i, column=1, value=line)
        if line.startswith("•"):
            cell.font = Font(bold=True)
        elif ":" in line:
            cell.font = Font(bold=True, color='6f42c1')
    
    # Create dashboard sheet - first remove any existing Dashboard sheet
    if 'Dashboard' in [sheet.title for sheet in wb.worksheets]:
        wb.remove(wb['Dashboard'])
    dashboard = wb.create_sheet("Dashboard")
    create_header(dashboard, f"📊 {month} - Financial Overview", '6f42c1')
    
    # Add sample data and charts
    dashboard['A3'] = "Financial Overview"
    dashboard['A3'].font = Font(bold=True, size=12, color='6f42c1')
    
    # Calculate monthly values based on the selected month
    month_index = months.index(month)
    monthly_income = 5000 + (month_index * 100)  # Sample data that increases slightly each month
    monthly_expenses = 3200 + (month_index * 50)  # Sample data that increases slightly each month
    monthly_savings = monthly_income - monthly_expenses
    
    # Calculate weekly breakdown
    weekly_income = [monthly_income * 0.25] * 4
    weekly_expenses = [
        monthly_expenses * 0.2,  # Week 1
        monthly_expenses * 0.3,  # Week 2
        monthly_expenses * 0.25, # Week 3
        monthly_expenses * 0.25  # Week 4
    ]
    weekly_savings = [i - e for i, e in zip(weekly_income, weekly_expenses)]
    
    # Add sample data with monthly focus
    dashboard_data = [
        ["Category", "Planned", "Actual", "Difference", "% of Budget"],
        ["Income", monthly_income, monthly_income * 1.04, "=C4-B4", "=C4/SUM(B4:B8)"],
        ["Expenses", monthly_expenses * 1.1, monthly_expenses, "=C5-B5", "=C5/SUM(B4:B8)"],
        ["Savings", monthly_savings * 1.1, monthly_savings, "=C6-B6", "=C6/SUM(B4:B8)"],
        ["Investments", 500, 550, "=C7-B7", "=C7/SUM(B4:B8)"],
        ["Total", "=SUM(B4:B7)", "=SUM(C4:C7)", "=C8-B8", "=C8/SUM(B4:B8)"]
    ]
    
    for r, row in enumerate(dashboard_data, start=3):
        for c, value in enumerate(row, start=1):
            cell = dashboard.cell(row=r, column=c, value=value)
            if r == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F2E6FF', end_color='F2E6FF', fill_type="solid")
    
    # Create or clear the Charts sheet
    if 'Charts' in wb.sheetnames:
        wb.remove(wb['Charts'])
    charts_sheet = wb.create_sheet('Charts')
    
    # Set up the Charts sheet
    charts_sheet.title = 'Charts'
    charts_sheet.sheet_view.showGridLines = False  # Hide gridlines for cleaner look
    
    # Title and styling
    charts_sheet.merge_cells('A1:J1')
    charts_sheet['A1'] = f"{month} - Financial Overview"
    charts_sheet['A1'].font = Font(size=18, bold=True, color='FFFFFF', name='Calibri')
    charts_sheet['A1'].fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type="solid")
    charts_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    charts_sheet.row_dimensions[1].height = 35
    
    # Add a subtitle
    charts_sheet.merge_cells('A2:J2')
    charts_sheet['A2'] = "Interactive Financial Visualizations"
    charts_sheet['A2'].font = Font(size=14, color='6f42c1', name='Calibri')
    charts_sheet['A2'].alignment = Alignment(horizontal="center")
    charts_sheet.row_dimensions[2].height = 25
    
    # 1. Weekly Financial Trend for the Month (Top-left)
    line_chart = LineChart()
    line_chart.title = f"{month} - Weekly Financial Trend"
    line_chart.style = 12
    line_chart.y_axis.title = 'Amount ($)'
    line_chart.x_axis.title = 'Week'
    line_chart.height = 15
    line_chart.width = 25
    
    # Add some padding for better visualization
    for row in range(5, 10):
        charts_sheet.row_dimensions[row].height = 15
    
    # Write weekly data to a dedicated section
    weekly_data_row = 50
    dashboard.cell(row=weekly_data_row, column=1, value="Week")
    dashboard.cell(row=weekly_data_row, column=2, value="Income")
    dashboard.cell(row=weekly_data_row, column=3, value="Expenses")
    dashboard.cell(row=weekly_data_row, column=4, value="Savings")
    
    for i, week in enumerate(weeks, 1):
        row = weekly_data_row + i
        dashboard.cell(row=row, column=1, value=week)
        dashboard.cell(row=row, column=2, value=weekly_income[i-1])
        dashboard.cell(row=row, column=3, value=weekly_expenses[i-1])
        dashboard.cell(row=row, column=4, value=weekly_savings[i-1])
    
    # Add data to the chart with proper series names
    data = Reference(dashboard, min_col=2, min_row=weekly_data_row-1, max_row=weekly_data_row+len(weeks)-1)
    line_chart.add_data(data, titles_from_data=True)
    
    # Add expenses data
    data = Reference(dashboard, min_col=3, min_row=weekly_data_row-1, max_row=weekly_data_row+len(weeks)-1)
    line_chart.add_data(data, titles_from_data=True)
    
    # Add savings data
    data = Reference(dashboard, min_col=4, min_row=weekly_data_row-1, max_row=weekly_data_row+len(weeks)-1)
    line_chart.add_data(data, titles_from_data=True)
    
    # Set series names explicitly using the correct openpyxl syntax
    from openpyxl.chart.series import SeriesLabel
    line_chart.series[0].title = SeriesLabel(v='Income')
    line_chart.series[1].title = SeriesLabel(v='Expenses')
    line_chart.series[2].title = SeriesLabel(v='Savings')
    
    # Set categories (x-axis)
    cats = Reference(dashboard, min_col=1, min_row=weekly_data_row+1, max_row=weekly_data_row+len(weeks))
    line_chart.set_categories(cats)
    
    # Customize the series with better colors
    line_chart.series[0].graphicalProperties.line.solidFill = "4CAF50"  # Green for income
    line_chart.series[1].graphicalProperties.line.solidFill = "F44336"  # Red for expenses
    line_chart.series[2].graphicalProperties.line.solidFill = "2196F3"  # Blue for savings
    
    # Add data labels to all points
    for series in line_chart.series:
        series.dLbls = openpyxl.chart.label.DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showSeriesName = True
    
    charts_sheet.add_chart(line_chart, "A3")
    
    # 2. Planned vs Actual Bar Chart (Top-right)
    chart1 = BarChart()
    data = Reference(dashboard, min_col=2, max_col=3, min_row=3, max_row=6)  # Exclude Total row
    cats = Reference(dashboard, min_col=1, min_row=4, max_row=6)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.title = "Planned vs Actual"
    chart1.style = 12
    chart1.y_axis.title = 'Amount ($)'
    chart1.x_axis.title = 'Category'
    chart1.height = 15
    chart1.width = 25
    
    # Customize colors
    chart1.series[0].graphicalProperties.solidFill = "4CAF50"  # Green for Planned
    chart1.series[1].graphicalProperties.solidFill = "2196F3"  # Blue for Actual
    
    # Add data labels
    chart1.dLbls = openpyxl.chart.label.DataLabelList()
    chart1.dLbls.showVal = True
    
    charts_sheet.add_chart(chart1, "L3")
    
    # Add sample expense breakdown for visualization with better formatting
    expense_data = [
        ["Category", "Amount", "Target", "Variance", "% of Total"],
        ["Housing", 1200, 1100, "=B2-C2", "=B2/SUM($B$2:$B$7)"],
        ["Food", 800, 700, "=B3-C3", "=B3/SUM($B$2:$B$7)"],
        ["Transportation", 400, 450, "=B4-C4", "=B4/SUM($B$2:$B$7)"],
        ["Utilities", 300, 350, "=B5-C5", "=B5/SUM($B$2:$B$7)"],
        ["Entertainment", 300, 250, "=B6-C6", "=B6/SUM($B$2:$B$7)"],
        ["Others", 200, 200, "=B7-C7", "=B7/SUM($B$2:$B$7)"],
        ["Total", "=SUM(B2:B7)", "=SUM(C2:C7)", "=SUM(D2:D7)", "=SUM(E2:E7)"]
    ]
    
    # Write expense breakdown data with formatting
    for r, row in enumerate(expense_data, start=25):
        for c, value in enumerate(row, start=1):
            cell = dashboard.cell(row=r, column=c, value=value)
            if r == 25:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
                cell.font = Font(color='FFFFFF')
            elif r == len(expense_data) + 24:  # Total row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F2E6FF', end_color='F2E6FF', fill_type='solid')
    
    # 3. Expense Breakdown Pie Chart (Bottom-left) - 3D with percentage and legend
    expense_pie = PieChart3D()
    data = Reference(dashboard, min_col=2, min_row=26, max_row=31)
    labels = Reference(dashboard, min_col=1, min_row=26, max_row=31)
    expense_pie.add_data(data, titles_from_data=False)
    expense_pie.set_categories(labels)
    expense_pie.title = f"{month} - Expenses Breakdown"
    expense_pie.height = 15
    expense_pie.width = 25
    
    # Add percentage data labels
    expense_pie.dataLabels = openpyxl.chart.label.DataLabelList()
    expense_pie.dataLabels.showPercent = True
    expense_pie.dataLabels.showVal = False
    expense_pie.dataLabels.showCatName = True
    expense_pie.legend.position = 'b'  # Bottom legend
    
    # Add some 3D rotation for better visibility
    expense_pie.rotX = 30
    expense_pie.rotY = 30
    
    charts_sheet.add_chart(expense_pie, "A25")
    
    # 4. Budget Distribution Donut Chart (Bottom-right)
    pie = PieChart3D()
    data = Reference(dashboard, min_col=2, min_row=4, max_row=6)
    labels = Reference(dashboard, min_col=1, min_row=4, max_row=6)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = f"{month} - Budget Distribution"
    pie.height = 15
    pie.width = 25
    
    # Make it a donut chart by setting a hole size
    pie.holeSize = 30  # Percentage of hole size
    
    # Add percentage data labels
    pie.dataLabels = openpyxl.chart.label.DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = True
    pie.legend.position = 'r'  # Right legend
    
    # Custom colors
    colors = ['4CAF50', '2196F3', 'FFC107']  # Green, Blue, Amber
    for i, series in enumerate(pie.series):
        series.graphicalProperties.solidFill = colors[i % len(colors)]
    
    # Add 3D rotation
    pie.rotX = 15
    pie.rotY = 30
    
    charts_sheet.add_chart(pie, "L25")
    
    # Weekly breakdown data is already calculated at the beginning of the function
    
    # Sample data for monthly trends
    income_data = [5000 + (i * 100) for i in range(12)]  # Increasing income
    expense_data = [3200 + (i * 50) for i in range(12)]  # Increasing expenses
    savings_data = [i - e for i, e in zip(income_data, expense_data)]  # Calculate savings
    
    # Write time series data with better formatting
    dashboard['A45'] = "Monthly Financial Trend"
    dashboard['A45'].font = Font(bold=True, size=12, color='6f42c1', name='Calibri')
    
    headers = ["Month", "Income", "Expenses", "Savings"]
    for col, header in enumerate(headers, 1):
        cell = dashboard.cell(row=46, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
    
    for i, (month, inc, exp, sav) in enumerate(zip(months, income_data, expense_data, savings_data), start=1):
        row = 46 + i
        dashboard.cell(row=row, column=1, value=month)
        
        # Format as currency
        for col, val in zip([2, 3, 4], [inc, exp, sav]):
            cell = dashboard.cell(row=row, column=col, value=val)
            cell.number_format = '$#,##0'
            
            # Add conditional formatting for negative savings
            if col == 4 and val < 0:
                cell.font = Font(color='FF0000')  # Red for negative savings
    
    # Auto-size columns for better visibility
    for column in dashboard.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        dashboard.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    # Define the exact order we want the sheets to appear in
    sheet_order = [
        'Welcome Guide',
        'Income Tracker',
        'Expense Tracker',
        'Savings Tracker',
        'Stock Tracker',
        'Weight Tracker',
        'Habit Tracker',
        'Cleaning Checklist',
        'Meal Planner',
        'Time Table',
        'AI Insights'
    ]
    
    # Map UI section names to tab names
    section_map = {
        'income': "Income Tracker",
        'expenses': "Expense Tracker",
        'savings': "Savings Tracker",
        'stocks': "Stock Tracker",
        'weight': "Weight Tracker",
        'habits': "Habit Tracker",
        'cleaning': "Cleaning Checklist",
        'meals': "Meal Planner",
        'timetable': "Time Table"
    }
    
    # Initialize tabs with core sheets
    tabs = {
        "Welcome Guide": all_tabs.get("Welcome Guide", {}),
        "AI Insights": all_tabs.get("AI Insights", {})
    }
    
    # If no sections specified, include all
    if not sections:
        sections = section_map.keys()
    
    # Add selected tabs to the tabs dictionary
    for section in sections:
        if section in section_map:
            tab_name = section_map[section]
            if tab_name in all_tabs:
                tabs[tab_name] = all_tabs[tab_name]
    
    # Process each tab in the desired order
    for tab_name in sheet_order:
        # Handle core sheets
        if tab_name in ['Welcome Guide', 'AI Insights']:
            tab_data = all_tabs.get(tab_name, {})
            if tab_name not in wb.sheetnames:
                sheet = wb.create_sheet(tab_name)
            else:
                sheet = wb[tab_name]
                sheet.delete_rows(1, sheet.max_row)
            
            if tab_name == 'Welcome Guide':
                create_welcome_guide(sheet)
            elif tab_name == 'AI Insights':
                create_header(sheet, "🤖 AI-Powered Financial Insights", '6f42c1')
                ai_text = [
                    "This section provides AI-generated insights based on your data:",
                    "",
                    "🔍 To get started:",
                    "1. Fill in your financial data in the respective sheets",
                    "2. Click the 'Generate AI Insights' button in the app",
                    "3. Review personalized recommendations here",
                    "",
                    "The AI will analyze your spending patterns, saving habits, and overall financial health to provide actionable advice.",
                    "",
                    "💡 Tip: The more data you provide, the more accurate the insights will be!"
                ]
                for i, line in enumerate(ai_text, start=3):
                    cell = sheet.cell(row=i, column=1, value=line)
                    if line.startswith("🔍") or line.startswith("💡"):
                        cell.font = Font(bold=True, color='6f42c1')
            continue
            
        # Skip Charts sheet for now (handled separately)
        if tab_name == 'Charts':
            continue
            
        # Skip if not in selected tabs
        if tab_name not in tabs:
            continue
            
        # Remove existing sheet if it exists (except for core sheets)
        if tab_name in wb.sheetnames and tab_name not in ['Welcome Guide', 'AI Insights']:
            try:
                wb.remove(wb[tab_name])
            except Exception as e:
                print(f"Warning: Could not remove sheet {tab_name}: {str(e)}")
            
        # Get the tab data (skip for AI Insights)
        tab_data = all_tabs.get(tab_name, {})
        
        # Create the sheet if it doesn't exist
        if tab_name not in wb.sheetnames:
            sheet = wb.create_sheet(tab_name)
        else:
            sheet = wb[tab_name]
            # Clear existing content but keep formatting
            sheet.delete_rows(1, sheet.max_row)
        
        # Handle Welcome Guide specially
        if tab_name == 'Welcome Guide':
            create_welcome_guide(sheet)
            continue
            
        # Handle AI Insights specially
        if tab_name == 'AI Insights':
            create_header(sheet, "🤖 AI-Powered Financial Insights", '6f42c1')
            ai_text = [
                "This section provides AI-generated insights based on your data:",
                "",
                "🔍 To get started:",
                "1. Fill in your financial data in the respective sheets",
                "2. Click the 'Generate AI Insights' button in the app",
                "3. Review personalized recommendations here",
                "",
                "The AI will analyze your spending patterns, saving habits, and overall financial health to provide actionable advice.",
                "",
                "💡 Tip: The more data you provide, the more accurate the insights will be!"
            ]
            
            for i, line in enumerate(ai_text, start=3):
                cell = sheet.cell(row=i, column=1, value=line)
                if line.startswith("🔍") or line.startswith("💡"):
                    cell.font = Font(bold=True, color='6f42c1')
            continue
            
        # For all other sheets, add the standard content
        title = f"{month} {tab_name}" if tab_name == 'Dashboard' else tab_name
        create_header(sheet, title, tab_data.get("color", "6f42c1"))
        
        # Add column headers
        headers = tab_data.get("headers", [])
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(
                start_color=tab_data.get("color", "6f42c1"), 
                end_color=tab_data.get("color", "6f42c1"), 
                fill_type='solid'
            )
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
        
        # Add sample data if available
        if "sample_data" in tab_data and tab_data["sample_data"]:
            data = tab_data["sample_data"]
            for row_num, row_data in enumerate(data, start=4):
                for col_num, cell_value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                    if isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
                        cell.number_format = '0.00'  # Removed currency symbol
            
            # Add total row if there are numeric columns
            if data and len(data[0]) > 1:  # Only if there are columns to sum
                total_row = len(data) + 4
                sheet.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
                
                # Calculate and add totals for numeric columns
                for col in range(2, len(headers) + 1):
                    if any(isinstance(sheet.cell(row=r, column=col).value, (int, float)) 
                          for r in range(4, total_row)):
                        # Calculate sum of the column
                        total = sum(sheet.cell(row=r, column=col).value or 0 
                                  for r in range(4, total_row) 
                                  if isinstance(sheet.cell(row=r, column=col).value, (int, float)))
                        cell = sheet.cell(row=total_row, column=col, value=total)
                        cell.font = Font(bold=True)
                        cell.number_format = '0.00'  # Format total without currency
                        
                        # Create a simple bar chart for the column
                        try:
                            if col > 1:  # Skip the first column (labels)
                                chart = BarChart()
                                chart.title = f"{tab_name} - {headers[col-1]}"
                                chart.style = 10
                                chart.y_axis.title = 'Amount'
                                chart.x_axis.title = 'Items'
                                
                                # Add data to chart
                                data_ref = Reference(sheet, min_col=col, min_row=3, max_row=total_row-1)
                                cats = Reference(sheet, min_col=1, min_row=4, max_row=total_row-1)
                                chart.add_data(data_ref, titles_from_data=True)
                                chart.set_categories(cats)
                                
                                # Position the chart
                                chart_cell = get_column_letter(col * 2) + str(total_row + 2)
                                sheet.add_chart(chart, chart_cell)
                        except Exception as e:
                            print(f"Error creating chart for {tab_name} column {col}: {str(e)}")
        
        # Auto-size columns for the current sheet
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    # Reorder sheets to match our desired order (in reverse to get correct positioning)
    for sheet_name in reversed(sheet_order):
        if sheet_name in wb.sheetnames:
            wb.move_sheet(wb[sheet_name], 0)
    
    # Set the active sheet to Dashboard or Welcome Guide
    if 'Dashboard' in wb.sheetnames:
        wb.active = wb['Dashboard']
    elif 'Welcome Guide' in wb.sheetnames:
        wb.active = wb['Welcome Guide']
    
    # Ensure all sheets are visible
    for sheet in wb.worksheets:
        sheet.sheet_state = 'visible'
    
    # Section mapping is now at the beginning of the function
    
    return wb

def read_excel_data_optimized(file_path, selected_categories=None, for_ai_conversion=False):
    """Optimized function to read Excel data using openpyxl for better accuracy.
    
    Args:
        file_path (str): Path to the Excel file
        selected_categories (list, optional): Categories to include. Defaults to None (all).
        for_ai_conversion (bool): If True, prepares data for AI-friendly conversion.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        data = {}

        # Map categories to sheet names
        category_map = {
            'Income': 'Income Tracker',
            'Expenses': 'Expense Tracker',
            'Savings': 'Savings Tracker',
            'Investments': 'Stock Tracker',
            'Health': ['Weight Tracker', 'Habit Tracker'],
            'Lifestyle': ['Cleaning Checklist', 'Meal Planner', 'Time Table']
        }

        # Determine which sheets to process
        if selected_categories:
            sheets_to_process = []
            for category in selected_categories:
                if category in category_map:
                    sheets = category_map[category]
                    if isinstance(sheets, list):
                        sheets_to_process.extend(sheets)
                    else:
                        sheets_to_process.append(sheets)
        else:
            # Exclude non-data sheets
            excluded_sheets = ['Welcome Guide', 'Dashboard', 'Charts', 'AI Insights', 'Summary']
            sheets_to_process = [sheet.title for sheet in wb.worksheets 
                               if sheet.title not in excluded_sheets]

        # Process each sheet
        for sheet_name in sheets_to_process:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                df = worksheet_to_dataframe(sheet)
                
                if not df.empty:
                    # Clean the data
                    df = df.dropna(how='all')
                    df = df.dropna(axis=1, how='all')
                    
                    if not df.empty:
                        if for_ai_conversion:
                            # For AI conversion, we want to keep more structured data
                            data[sheet_name] = df
                        else:
                            # For display purposes, convert to string
                            data[sheet_name] = df.head(20).to_string()

        return data
    except Exception as e:
        return {"error": f"Error reading Excel file: {str(e)}"}

def create_ai_friendly_template(output_path="ai_finance_template.xlsx"):
    """Creates an AI-friendly Excel template with optimized structure for financial tracking."""
    from openpyxl.styles import Font, PatternFill
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("AI_Instructions")
    ws_instructions['A1'] = "AI-Friendly Data Formatting Instructions"
    ws_instructions['A1'].font = Font(bold=True, size=14)
    instructions = [
        "1. Keep one data type per column",
        "2. Use consistent date formats (YYYY-MM-DD recommended)",
        "3. Avoid merged cells",
        "4. Use headers in the first row",
        "5. No empty rows or columns within the data",
        "6. Use consistent category names",
        "7. Include all relevant metadata in separate columns",
        "8. Use numerical values without currency symbols"
    ]
    
    for i, instruction in enumerate(instructions, start=3):
        ws_instructions[f'A{i}'] = instruction
    
    # Add sample data sheets for different data types
    def add_sample_sheet(name, headers, data):
        ws = wb.create_sheet(name)
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 30)  # Cap at 30 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Transactions sample
    transactions_headers = ["Date", "Type", "Category", "Description", "Amount", "Account", "Tags"]
    transactions_data = [
        ["2023-11-01", "Expense", "Food", "Grocery Shopping", 150.00, "Checking", "groceries"],
        ["2023-11-02", "Income", "Salary", "Monthly Salary", 3000.00, "Savings", "salary"],
        ["2023-11-03", "Expense", "Transport", "Bus Fare", 50.00, "Credit Card", "commute"]
    ]
    add_sample_sheet("Transactions", transactions_headers, transactions_data)
    
    # Budget sample
    budget_headers = ["Category", "Budgeted", "Spent", "Remaining", "Period"]
    budget_data = [
        ["Groceries", 600.00, 150.00, 450.00, "Monthly"],
        ["Transportation", 200.00, 50.00, 150.00, "Monthly"],
        ["Entertainment", 100.00, 0.00, 100.00, "Monthly"]
    ]
    add_sample_sheet("Budget", budget_headers, budget_data)
    
    # Save the workbook
    wb.save(output_path)
    return output_path

def convert_to_ai_friendly(input_file, output_file):
    """Convert an existing Excel file to AI-friendly format.
    
    Args:
        input_file (str): Path to the input Excel file
        output_file (str): Path to save the converted file
        
    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        # Read the input file
        wb_in = openpyxl.load_workbook(input_file, data_only=True)
        wb_out = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb_out.sheetnames:
            wb_out.remove(wb_out['Sheet'])
        
        # Add instructions sheet
        ws_instructions = wb_out.create_sheet("AI_Instructions")
        ws_instructions['A1'] = "AI-Friendly Data Format"
        ws_instructions['A1'].font = Font(bold=True, size=14)
        
        instructions = [
            "This file has been converted to be more AI-friendly. Here's what was done:",
            "1. Removed formatting and merged cells",
            "2. Ensured consistent data types in columns",
            "3. Standardized date formats",
            "4. Added clear headers",
            "5. Removed empty rows and columns"
        ]
        
        for i, instruction in enumerate(instructions, start=3):
            ws_instructions[f'A{i}'] = instruction
        
        # Process each sheet
        for sheet_name in wb_in.sheetnames:
            # Skip non-data sheets
            if sheet_name in ['Welcome Guide', 'Charts', 'AI Insights', 'Dashboard']:
                continue
                
            ws_in = wb_in[sheet_name]
            df = worksheet_to_dataframe(ws_in)
            
            if not df.empty:
                # Clean the data
                df = df.dropna(how='all').dropna(axis=1, how='all')
                
                if not df.empty:
                    # Standardize column names
                    df.columns = [str(col).strip() for col in df.columns]
                    
                    # Standardize date formats if present
                    date_columns = [col for col in df.columns if 'date' in col.lower() or 'Date' in str(col)]
                    for col in date_columns:
                        try:
                            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
                        except:
                            pass
                    
                    # Create new sheet in output workbook
                    safe_sheet_name = re.sub(r'[\[\]\:*?/\\]', '_', sheet_name)[:31]  # Excel sheet name limit
                    ws_out = wb_out.create_sheet(safe_sheet_name)
                    
                    # Write data to new sheet
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws_out.append(r)
                    
                    # Format headers
                    for cell in ws_out[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    
                    # Auto-adjust column widths
                    for column in ws_out.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min((max_length + 2), 30)  # Cap at 30 characters
                        ws_out.column_dimensions[column_letter].width = adjusted_width
        
        # Add metadata sheet
        ws_meta = wb_out.create_sheet("_metadata")
        ws_meta['A1'] = "File Information"
        ws_meta['A1'].font = Font(bold=True, size=12)
        
        meta_data = [
            ("Original file:", os.path.basename(input_file)),
            ("Converted on:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Conversion notes:", "This file has been optimized for AI analysis.")
        ]
        
        for i, (label, value) in enumerate(meta_data, start=3):
            ws_meta[f'A{i}'] = label
            ws_meta[f'B{i}'] = value
            ws_meta[f'A{i}'].font = Font(bold=True)
        
        # Save the output file
        wb_out.save(output_file)
        return True, ""
        
    except Exception as e:
        return False, f"Error during conversion: {str(e)}"

def generate_ai_insights(file_path, selected_categories=None):
    """
    Generate AI insights from the uploaded Excel file using Ollama.
    
    Args:
        file_path (str): Path to the Excel file
        selected_categories (list): List of categories to analyze
        
    Returns:
        tuple: (excel_data_str, insights) - The data string and AI-generated insights
    """
    try:
        # Read the Excel file
        excel_data = read_excel_data_optimized(file_path, selected_categories, for_ai_conversion=True)
        
        if not excel_data or "error" in excel_data:
            error_msg = excel_data.get("error", "No data found in the Excel file")
            return "", f"❌ Error processing Excel file: {error_msg}"
        
        # Convert data to a string format for the AI
        excel_data_str = ""
        for sheet_name, df in excel_data.items():
            excel_data_str += f"\n--- {sheet_name} ---\n"
            excel_data_str += df.to_string() + "\n\n"
        
        # Prepare the prompt for Ollama
        prompt = f"""Analyze the following financial and personal data and provide insights.
        Focus on spending patterns, savings rate, and potential areas for improvement.
        Be specific and provide actionable recommendations.
        
        Data:
        {excel_data_str}
        
        Key points to analyze:
        1. What are the main spending categories?
        2. Are there any unusual transactions or patterns?
        3. How does income compare to expenses?
        4. What are the savings trends?
        5. Any recommendations for budget optimization?
        """
        
        try:
            # First, check if Ollama is running
            try:
                print("Attempting to list Ollama models...")  # Debug log
                
                # List models using Ollama
                models_response = ollama.list()
                print(f"Ollama response type: {type(models_response)}")  # Debug log
                
                # Initialize available_models list
                available_models = []
                
                # Handle the case where models is a direct attribute
                if hasattr(models_response, 'models'):
                    print("Found models attribute in response")
                    for model in models_response.models:
                        if hasattr(model, 'model'):  # Check for model attribute
                            available_models.append(model.model)
                            print(f"Found model: {model.model}")
                
                # If no models found yet, try to access the models list directly
                if not available_models and isinstance(models_response, list):
                    print("Response is a list of models")
                    for model in models_response:
                        if hasattr(model, 'model'):
                            available_models.append(model.model)
                            print(f"Found model in list: {model.model}")
                
                # If still no models, try to access the models attribute through __dict__
                if not available_models and hasattr(models_response, '__dict__'):
                    models_dict = models_response.__dict__
                    if 'models' in models_dict and models_dict['models']:
                        for model in models_dict['models']:
                            if hasattr(model, 'model'):
                                available_models.append(model.model)
                                print(f"Found model in __dict__: {model.model}")
                
                print(f"Available models: {available_models}")
                
                if not available_models:
                    # As a last resort, try to use the first available model directly
                    try:
                        response = ollama.generate(model='llama2', prompt='Test')
                        if response and hasattr(response, 'response'):
                            available_models = ['llama2']
                            print("Successfully tested llama2 model")
                    except Exception as e:
                        print(f"Could not use llama2 model: {str(e)}")
                
            except Exception as e:
                raise Exception(f"Could not connect to Ollama. Is it running? Error: {str(e)}")
            
            # If we have available models, try to use one
            if available_models:
                # Prefer gemma3:4b if available, then llama2, otherwise use the first available model
                model_to_use = next(
                    (m for m in available_models if 'gemma3:4b' in m.lower()) or 
                    (m for m in available_models if 'llama2' in m.lower()) or 
                    available_models,
                    available_models[0]
                )
                print(f"Using model: {model_to_use}")
                
                try:
                    # Generate insights using the selected model
                    response = ollama.generate(
                        model=model_to_use,
                        prompt=prompt,
                        stream=False
                    )
                    
                    if hasattr(response, 'response'):
                        insights = response.response
                    elif isinstance(response, dict) and 'response' in response:
                        insights = response['response']
                    else:
                        insights = str(response)
                    
                    return excel_data_str, insights
                    
                except Exception as e:
                    raise Exception(f"Error generating insights with model '{model_to_use}': {str(e)}")
            else:
                raise Exception("No suitable AI models found. Please install a model with 'ollama pull <model>'")
            
            try:
                # Try to use the model to generate insights
                response = ollama.generate(
                    model=model_to_use,
                    prompt=prompt,
                    stream=False
                )
                insights = response.get('response', 'No insights generated')
                return excel_data_str, insights
                
            except Exception as e:
                raise Exception(f"Error generating insights with model '{model_to_use}': {str(e)}")
                
        except Exception as e:
            # Format the error message with installation instructions
            error_msg = f"""
            ⚠️ Error generating AI insights: {str(e)}
            
            To enable AI analysis, you need to install an AI model first.
            
            Try running one of these commands in your terminal:
            - `ollama pull llama2`  # Recommended for most users
            - `ollama pull mistral` # Good alternative
            - `ollama pull gemma`   # Lightweight option
            
            After installation, restart this app and try again.
            
            Here's your data for manual review:
            {excel_data_str}
            """
            return excel_data_str, error_msg
            
    except Exception as e:
        return "", f"❌ Error: {str(e)}"

def main():
    st.title("📊 Life & Budget Dashboard")
    st.markdown("### Your All-in-One Financial and Personal Management Tool")
    
    # Add AI-friendly template download button
    if st.sidebar.button("📥 Download AI-Friendly Template"):
        with st.spinner("Creating your template..."):
            template_path = create_ai_friendly_template()
            with open(template_path, "rb") as f:
                template_data = f.read()
            
            st.sidebar.success("Template created!")
            st.sidebar.download_button(
                label="Download AI-Friendly Template",
                data=template_data,
                file_name="ai_finance_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Clean up the temporary file
            try:
                os.remove(template_path)
            except:
                pass
    
    # Sidebar for navigation
    st.sidebar.title("Navigation")
    page = st.sidebar.radio("Go to", ["Home", "Generate Template", "AI Insights", "AI Template Converter"])
    
    if page == "Home":
        st.markdown("""
        Welcome to your personal Life & Budget Dashboard! This tool helps you track your finances, 
        health, and daily activities all in one place.
        
        ### Features:
        - **Financial Tracking**: Income, expenses, savings, and investments
        - **Health & Wellness**: Weight tracking and habit formation
        - **Life Organization**: Cleaning schedules and meal planning
        
        Get started by generating a new template or upload your existing data for AI-powered insights.
        """)
        
    elif page == "Generate Template":
        st.header("📝 Generate New Template")
        st.write("Create a new Excel template with the sections you need.")
        
        # Month selection
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        selected_month = st.selectbox("Select Month", months, index=datetime.date.today().month - 1)
        
        # Section selection
        st.subheader("Select Sections to Include")
        
        # Create columns for better organization
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("**Financial**")
            income = st.checkbox("Income", value=True, key="income_cb")
            expenses = st.checkbox("Expenses", value=True, key="expenses_cb")
            savings = st.checkbox("Savings", value=True, key="savings_cb")
            stocks = st.checkbox("Stocks", value=True, key="stocks_cb")
        
        with col2:
            st.markdown("**Health**")
            weight = st.checkbox("Weight", value=True, key="weight_cb")
            habits = st.checkbox("Habits", value=True, key="habits_cb")
        
        with col3:
            st.markdown("**Life Organization**")
            cleaning = st.checkbox("Cleaning", value=True, key="cleaning_cb")
            meals = st.checkbox("Meal Planning", value=True, key="meals_cb")
            timetable = st.checkbox("Time Table", value=True, key="timetable_cb")
        
        st.markdown("---")

        # Create a single, correct list of selected sections to be used everywhere
        selected_sections = []
        if income: selected_sections.append("income")
        if expenses: selected_sections.append("expenses")
        if savings: selected_sections.append("savings")
        if stocks: selected_sections.append("stocks")
        if weight: selected_sections.append("weight")
        if habits: selected_sections.append("habits")
        if cleaning: selected_sections.append("cleaning")
        if meals: selected_sections.append("meals")
        if timetable: selected_sections.append("timetable")

        # If no sections are selected, include all
        if not selected_sections:
            selected_sections = [
                'income', 'expenses', 'savings', 'stocks',
                'weight', 'habits', 'cleaning', 'meals', 'timetable'
            ]

        if st.button("🔍 Generate Preview"):
            with st.spinner(f"Creating your {selected_month} template..."):
                try:
                    # Generate the workbook in memory
                    wb = create_excel_template(month=selected_month, sections=selected_sections)
                    
                    st.subheader("Excel Sheet Preview")
                    for sheet in wb.worksheets:
                        with st.expander(f"Sheet: {sheet.title}"):
                            # Convert sheet to DataFrame for preview
                            df = worksheet_to_dataframe(sheet)
                            st.dataframe(df)
                except Exception as e:
                    st.error(f"Error generating preview: {str(e)}")
        
        if st.button("✨ Generate Template"):
            with st.spinner(f"Creating your {selected_month} template..."):
                # Create a temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    file_path = tmp.name
                
                try:
                    # Create the template with the selected month and sections
                    wb = create_excel_template(month=selected_month, sections=selected_sections)
                    
                    # Debug: Print the sheets that were created
                    st.sidebar.write("Sheets created:", wb.sheetnames)
                    
                except Exception as e:
                    st.error(f"Error creating Excel template: {str(e)}")
                    st.stop()
                wb.save(file_path)
                
                # Create download link
                with open(file_path, 'rb') as f:
                    bytes_data = f.read()
                
                st.success(f"{selected_month} template created successfully!")
                st.download_button(
                    label=f"📥 Download {selected_month} Template",
                    data=bytes_data,
                    file_name=f"life_budget_tracker_{selected_month}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Clean up
                os.unlink(file_path)
                
    elif page == "AI Template Converter":
        st.header("🤖 Convert to AI-Friendly Format")
        st.write("Upload your Excel file to convert it to a more AI-friendly format that works better with Ollama.")
        
        uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])
        
        if uploaded_file is not None:
            with st.spinner("Converting your file..."):
                # Create a temporary file for the uploaded file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_in:
                    tmp_in.write(uploaded_file.getvalue())
                    input_path = tmp_in.name
                
                # Create a temporary output file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_out:
                    output_path = tmp_out.name
                
                try:
                    # Convert the file
                    success, message = convert_to_ai_friendly(input_path, output_path)
                    
                    if success:
                        st.success("File converted successfully!")
                        
                        # Create download link
                        with open(output_path, 'rb') as f:
                            bytes_data = f.read()
                        
                        st.download_button(
                            label="💾 Download AI-Friendly File",
                            data=bytes_data,
                            file_name=f"ai_friendly_{uploaded_file.name}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        st.markdown("### What was improved:")
                        st.markdown("""
                        - 🧹 Removed formatting and merged cells
                        - 📊 Standardized data types and formats
                        - 📅 Ensured consistent date formatting
                        - 🏷️ Added clear headers
                        - 🗑️ Removed empty rows and columns
                        - 📝 Added metadata and instructions
                        """)
                    else:
                        st.error(f"Error during conversion: {message}")
                        
                except Exception as e:
                    st.error(f"An error occurred: {str(e)}")
                    
                finally:
                    # Clean up temporary files
                    for path in [input_path, output_path]:
                        try:
                            if os.path.exists(path):
                                os.unlink(path)
                        except:
                            pass
    
    elif page == "AI Insights":
        st.header("📊 Upload Your Data for AI Insights")
        
        # Add category selection
        categories = [
            'Income',
            'Expenses',
            'Savings',
            'Investments',
            'Health',
            'Lifestyle'
        ]
        
        selected_categories = st.multiselect(
            "Select categories to analyze:",
            categories,
            default=categories[:2]  # Default to Income and Expenses
        )
        
        uploaded_file = st.file_uploader("Upload your filled Excel file", type=["xlsx"])
        
        if uploaded_file is not None:
            if st.button("🤖 Generate AI Insights", key="ai_insights_btn"):
                with st.spinner("Analyzing your data with AI..."):
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        tmp.write(uploaded_file.getvalue())
                        tmp_path = tmp.name
                    
                    try:
                        excel_data_str, insights = generate_ai_insights(tmp_path, selected_categories)
                        
                        # --- DEBUG: Show the data sent to the AI --- #
                        with st.expander("View Data Sent to AI (for debugging)"):
                            st.text(excel_data_str)
                        # --- END DEBUG --- #

                        if not insights.startswith(('❌', '⚠️')):
                            st.success("AI Analysis Complete!")
                            st.markdown("### 🎯 Your Personalized Insights")
                            
                            with st.expander("View Insights", expanded=True):
                                st.markdown(insights)
                            
                            # Download buttons
                            col1, col2 = st.columns(2)
                            with col1:
                                st.download_button(
                                    label="📝 Download as Text",
                                    data=insights,
                                    file_name="financial_insights.txt",
                                    mime="text/plain"
                                )
                            with col2:
                                pdf_path = generate_pdf(insights)
                                with open(pdf_path, "rb") as f:
                                    st.download_button(
                                        label="📄 Download as PDF",
                                        data=f,
                                        file_name="financial_insights.pdf",
                                        mime="application/pdf"
                                    )
                                os.unlink(pdf_path)
                        else:
                            st.error(insights)
                    finally:
                        if os.path.exists(tmp_path):
                            os.unlink(tmp_path)
    
    st.markdown("---")
    st.markdown("### 📱 Features at a Glance")
    
    features = st.columns(3)
    
    with features[0]:
        st.markdown("""
        #### 💰 Financial Tracking
        - Income & Expense tracking
        - Budget planning
        - Savings goals
        - Investment portfolio
        """)
    
    with features[1]:
        st.markdown("""
        #### 🏋️ Health & Wellness
        - Weight tracking
        - Habit formation
        - Self-care routines
        - Meal planning
        """)
    
    with features[2]:
        st.markdown("""
        #### 🏠 Life Organization
        - Cleaning checklist
        - Weekly schedule
        - Time management
        - Task tracking
        """)
    
    st.markdown("---")
    st.markdown("""
    ### Getting Started
    1. Download the Excel template above
    2. Fill in your financial and personal data
    3. Upload it back to get AI-powered insights
    4. Use the dashboard to track your progress
    """)

def generate_pdf(insights_text, excel_data_str=None):
    """
    Generate a PDF report from insights and optional Excel data
    
    Args:
        insights_text (str): The insights text to include in the PDF
        excel_data_str (str, optional): Formatted Excel data as string. Defaults to None.
    
    Returns:
        str: Path to the generated PDF file
    """
    # Create PDF object
    pdf = FPDF()
    pdf.add_page()
    
    # Set up the document
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Financial Insights Report', 0, 1, 'C')
    pdf.ln(10)
    
    # Add date
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f'Generated on: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}', 0, 1)
    pdf.ln(5)
    
    # Add a line
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(10)
    
    # Add insights
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'AI-Generated Insights', 0, 1)
    pdf.ln(5)
    
    # Format and add insights text
    pdf.set_font('Arial', '', 12)
    for line in insights_text.split('\n'):
        if line.strip() == '':
            pdf.ln(5)
        else:
            # Handle bullet points and formatting
            if line.strip().startswith(('•', '-', '*')):
                pdf.set_font('', 'B')
                pdf.cell(10, 10, '•', 0, 0)
                pdf.cell(0, 10, line[1:].strip(), 0, 1)
                pdf.set_font('', '')
            elif ':' in line and len(line) < 50:  # Likely a heading
                pdf.set_font('', 'B')
                pdf.cell(0, 10, line.strip(), 0, 1)
                pdf.set_font('', '')
            else:
                pdf.multi_cell(0, 10, line.strip())
    
    # Add Excel data if provided
    if excel_data_str:
        pdf.add_page()
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Supporting Data', 0, 1)
        pdf.ln(5)
        
        pdf.set_font('Courier', '', 8)
        pdf.multi_cell(0, 5, excel_data_str)
    
    # Save the PDF to a temporary file
    temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdf_path = temp_pdf.name
    pdf.output(pdf_path)
    
    return pdf_path

if __name__ == "__main__":
    main()