# app.py
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import tempfile
import os
import datetime
import json
import ollama
from io import BytesIO

# Set page config
st.set_page_config(
    page_title="Life & Budget Dashboard",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for styling
st.markdown("""
<style>
    .main {
        background-color: #f8f9fa;
    }
    .stButton>button {
        background-color: #6f42c1;
        color: white;
        border-radius: 8px;
        padding: 0.5rem 1rem;
    }
    .stButton>button:hover {
        background-color: #5a32a3;
        color: white;
    }
    .stTextInput>div>div>input {
        border: 1px solid #6f42c1;
        border-radius: 4px;
    }
    .css-1d391kg {
        padding-top: 3rem;
    }
</style>
""", unsafe_allow_html=True)

def create_excel_template():
    wb = openpyxl.Workbook()
    # Remove default sheet
    while len(wb.sheetnames) > 0:
        wb.remove(wb[wb.sheetnames[0]])

    # Color scheme
    colors = {
        'pink': 'FFB6C1',
        'purple': 'D8BFD8',
        'blue': 'B0E0E6',
        'green': '98FB98',
        'yellow': 'FFFACD',
        'orange': 'FFDAB9',
        'teal': 'AFEEEE',
        'lavender': 'E6E6FA',
        'peach': 'FFDAB9',
        'mint': '98FF98'
    }

    # Helper function to create headers
    def create_header(sheet, title, color):
        sheet.merge_cells('A1:J1')
        sheet['A1'] = title
        sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sheet['A1'].alignment = Alignment(horizontal="center")
        sheet.row_dimensions[1].height = 30
        
        # Add some spacing
        sheet.row_dimensions[2].height = 10

    # Welcome Guide
    welcome = wb.create_sheet("Welcome Guide")
    create_header(welcome, "🌟 Welcome to Your Life & Budget Dashboard 2025 🌟", '6f42c1')
    
    welcome_text = [
        "Thank you for choosing the Life & Budget Dashboard! This comprehensive tool will help you manage:",
        "• Personal Finances (Income, Expenses, Savings, Investments)",
        "• Health & Wellness (Weight, Self-Care, Habits)",
        "• Daily Life (Meal Planning, Cleaning, Schedules)",
        "",
        "📊 HOW TO USE:",
        "1. Start by entering your monthly budget in the 'Dashboard' tab",
        "2. Track your income and expenses in their respective tabs",
        "3. Monitor your investments in the 'Stock Tracker'",
        "4. Use the health and lifestyle trackers to maintain balance",
        "5. Check the 'Dashboard' for insights and progress",
        "",
        "💡 TIP: Use the 'AI Insights' button to get personalized recommendations!",
        "",
        "🔒 Your data stays on your device. For cloud sync, save this file to your preferred cloud storage."
    ]
    
    for i, line in enumerate(welcome_text, start=3):
        cell = welcome.cell(row=i, column=1, value=line)
        if line.startswith("•"):
            cell.font = Font(bold=True)
        elif ":" in line:
            cell.font = Font(bold=True, color='6f42c1')
    
    # Dashboard
    dashboard = wb.create_sheet("Dashboard")
    create_header(dashboard, "📊 Dashboard Overview", '6f42c1')
    
    # Add sample data and charts
    dashboard['A3'] = "Financial Overview"
    dashboard['A3'].font = Font(bold=True, size=12, color='6f42c1')
    
    # Add sample data
    dashboard_data = [
        ["Category", "Planned", "Actual", "Difference", "% of Budget"],
        ["Income", 5000, 5200, "=C4-B4", "=C4/SUM(B4:B8)"],
        ["Expenses", 3500, 3200, "=C5-B5", "=C5/SUM(B4:B8)"],
        ["Savings", 1000, 1200, "=C6-B6", "=C6/SUM(B4:B8)"],
        ["Investments", 500, 550, "=C7-B7", "=C7/SUM(B4:B8)"],
        ["Total", "=SUM(B4:B7)", "=SUM(C4:C7)", "=C8-B8", "=C8/SUM(B4:B8)"]
    ]
    
    for r, row in enumerate(dashboard_data, start=3):
        for c, value in enumerate(row, start=1):
            cell = dashboard.cell(row=r, column=c, value=value)
            if r == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F2E6FF', end_color='F2E6FF', fill_type="solid")
    
    # Add a sample chart
    chart1 = BarChart()
    data = Reference(dashboard, min_col=2, max_col=3, min_row=3, max_row=7)
    cats = Reference(dashboard, min_col=1, min_row=4, max_row=7)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.title = "Planned vs Actual"
    chart1.style = 10
    dashboard.add_chart(chart1, "A12")
    
    # Add more tabs with similar structure
    tabs = {
        "Income Tracker": {
            "headers": ["Date", "Source", "Amount", "Category", "Notes"],
            "color": "B0E0E6",  # Blue
            "sample_data": [
                [datetime.date.today(), "Salary", 3000, "Primary Income", "Monthly salary"],
                [datetime.date.today().replace(day=15), "Freelance", 1000, "Side Hustle", "Project X"]
            ]
        },
        "Expense Tracker": {
            "headers": ["Date", "Description", "Amount", "Category", "Payment Method", "Notes"],
            "color": "FFB6C1",  # Pink
            "sample_data": [
                [datetime.date.today(), "Groceries", 150, "Food", "Credit Card", "Weekly shopping"],
                [datetime.date.today(), "Electricity", 80, "Utilities", "Direct Debit", "Monthly bill"]
            ]
        },
        "Savings Tracker": {
            "headers": ["Date", "Goal", "Target Amount", "Current Amount", "% Complete"],
            "color": "98FB98",  # Green
            "sample_data": [
                [datetime.date.today(), "Emergency Fund", 10000, 3500, "=D2/C2"]
            ]
        },
        "Stock Tracker": {
            "headers": ["Symbol", "Company", "Shares", "Avg Price", "Current Price", "Total Value", "Gain/Loss", "% Change"],
            "color": "D8BFD8",  # Purple
            "sample_data": [
                ["AAPL", "Apple Inc.", 10, 150, 175, "=C2*E2", "=F2-(C2*D2)", "=(E2-D2)/D2"]
            ]
        },
        "Weight Tracker": {
            "headers": ["Date", "Weight (kg)", "Body Fat %", "Notes"],
            "color": "FFDAB9",  # Orange
            "sample_data": [
                [datetime.date.today() - datetime.timedelta(days=7), 75.5, 22.0, "Started new diet"],
                [datetime.date.today(), 74.8, 21.5, "Feeling good!"]
            ]
        },
        "Habit Tracker": {
            "headers": ["Date", "Exercise", "Water (glasses)", "Sleep (hours)", "Meditation", "Reading", "Notes"],
            "color": "E6E6FA",  # Lavender
            "sample_data": [
                [datetime.date.today() - datetime.timedelta(days=1), "✓", 8, 7.5, "✓", "30 min", "Good day!"],
                [datetime.date.today(), "✓", 7, 8.0, "✓", "45 min", "Focused"]
            ]
        },
        "Cleaning Checklist": {
            "headers": ["Task", "Frequency", "Last Done", "Next Due", "Status", "Notes"],
            "color": "AFEEEE",  # Teal
            "sample_data": [
                ["Vacuum", "Weekly", datetime.date.today() - datetime.timedelta(days=3), 
                 datetime.date.today() + datetime.timedelta(days=4), "Pending", "Living room"],
                ["Laundry", "Bi-weekly", datetime.date.today() - datetime.timedelta(days=6), 
                 datetime.date.today() + datetime.timedelta(days=1), "Due", "Bedding"]
            ]
        },
        "Meal Planner": {
            "headers": ["Day", "Breakfast", "Lunch", "Dinner", "Snacks", "Grocery Items Needed"],
            "color": "FFD700",  # Gold
            "sample_data": [
                ["Monday", "Oatmeal", "Salad", "Grilled Chicken", "Fruits", "Oats, chicken, greens"],
                ["Tuesday", "Smoothie", "Sandwich", "Pasta", "Nuts", "Bread, pasta, vegetables"]
            ]
        },
        "Time Table": {
            "headers": ["Time", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
            "color": "FFA07A",  # Light Salmon
            "sample_data": [
                ["09:00 - 10:00", "Work", "Work", "Work", "Work", "Work", "Sleep in", "Brunch"],
                ["10:00 - 11:00", "Meeting", "Deep Work", "Meeting", "Deep Work", "Meeting", "Exercise", "Family Time"]
            ]
        }
    }

    for sheet_name, data in tabs.items():
        sheet = wb.create_sheet(sheet_name)
        create_header(sheet, f"📋 {sheet_name}", data["color"])
        
        # Add headers
        for col_num, header in enumerate(data["headers"], 1):
            cell = sheet.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=data["color"], end_color=data["color"], fill_type="solid")
            cell.border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))
        
        # Add sample data
        for row_num, row_data in enumerate(data["sample_data"], 4):
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.border = Border(left=Side(style='thin'), 
                                   right=Side(style='thin'), 
                                   top=Side(style='thin'), 
                                   bottom=Side(style='thin'))
        
        # Auto-size columns
        for col in sheet.columns:
            if not col or not col[0]:
                continue
                
            # Skip merged cells
            if hasattr(col[0], 'column_letter'):
                column = col[0].column_letter
                max_length = 0
                for cell in col:
                    try:
                        if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = min(adjusted_width, 30)
    
    # Add AI Insights sheet
    ai_sheet = wb.create_sheet("AI Insights")
    create_header(ai_sheet, "🤖 AI-Powered Financial Insights", '6f42c1')
    
    ai_text = [
        "This section provides AI-generated insights based on your data:",
        "",
        "🔍 To get started:",
        "1. Fill in your financial data in the respective sheets",
        "2. Click the 'Generate AI Insights' button in the app",
        "3. Review personalized recommendations here",
        "",
        "The AI will analyze your spending patterns, saving habits, and overall financial health to provide actionable advice.",
        "",
        "💡 Tip: The more data you provide, the more accurate the insights will be!"
    ]
    
    for i, line in enumerate(ai_text, start=3):
        cell = ai_sheet.cell(row=i, column=1, value=line)
        if line.startswith("🔍") or line.startswith("💡"):
            cell.font = Font(bold=True, color='6f42c1')
    
    return wb

def read_excel_data(file_path):
    """Read relevant data from the Excel file."""
    try:
        # Read the Excel file
        xls = pd.ExcelFile(file_path)
        
        # Initialize data dictionary
        data = {}
        
        # Read each sheet that has data
        for sheet_name in xls.sheet_names:
            if sheet_name == 'AI Insights':
                continue
                
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
            # Only include sheets with data (more than just headers)
            if len(df) > 1:
                data[sheet_name] = df.to_string()
        
        return data
    except Exception as e:
        return {"error": f"Error reading Excel file: {str(e)}"}

def generate_ai_insights(file_path):
    """Generate AI insights using Ollama."""
    try:
        # First, check if Ollama is running
        try:
            models = ollama.list()
            if not models.get('models'):
                raise Exception("No models available. Please install a model with 'ollama pull <model_name>'")
        except Exception as e:
            return f"❌ Error connecting to Ollama: {str(e)}\n\nPlease ensure Ollama is running and you have at least one model installed.\n\nTo install a model, run: ollama pull mistral"
        
        # Read data from Excel
        excel_data = read_excel_data(file_path)
        if isinstance(excel_data, dict) and 'error' in excel_data:
            return excel_data['error']
        
        # Prepare the prompt for the AI
        prompt = """Analyze the following financial and personal data from different sheets of an Excel file. 
        Provide a comprehensive financial analysis including:
        1. Financial health overview
        2. Spending patterns and potential savings opportunities
        3. Progress towards financial goals
        4. Any concerning trends or anomalies
        5. Personalized recommendations
        
        Here's the data from each sheet:
        """
        
        # Add data from each sheet to the prompt
        for sheet_name, content in excel_data.items():
            prompt += f"\n\n--- {sheet_name} ---\n{content[:2000]}..."  # Limit content length to avoid context window issues
        
        # Add instructions for the response format
        prompt += """
        
        Provide your analysis in a clear, structured format with emojis for better readability.
        Focus on actionable insights and specific recommendations.
        """
        
        # Generate response using Ollama
        response = ollama.generate(
            model='mistral',  # You can change this to your preferred model
            prompt=prompt,
            options={
                'temperature': 0.7,
                'max_tokens': 1500
            }
        )
        
        return response['response']
        
    except Exception as e:
        return f"❌ Error generating insights: {str(e)}\n\nPlease ensure Ollama is running and you have a model installed.\n\nTo install a model, run: ollama pull mistral"

def main():
    st.title("💰 Life & Budget Dashboard 2025")
    st.markdown("### Your all-in-one financial and life management solution")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("""
        Welcome to your personal finance and life management dashboard! This tool helps you:
        - Track income, expenses, and savings
        - Monitor investments and net worth
        - Maintain health and wellness goals
        - Organize daily life with meal planning and cleaning checklists
        - Get AI-powered insights on your financial health
        """)
        
        if st.button("📥 Download Excel Template", key="download_btn"):
            with st.spinner("Creating your personalized dashboard..."):
                wb = create_excel_template()
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    wb.save(tmp.name)
                    with open(tmp.name, "rb") as f:
                        st.download_button(
                            label="⬇️ Click to download your template",
                            data=f,
                            file_name="Life_Budget_Dashboard_2025.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                os.unlink(tmp.name)
    
    with col2:
        st.image("https://cdn-icons-png.flaticon.com/512/3132/3132693.png", width=200)
    
    st.markdown("---")
    
    st.header("📊 Upload Your Data for AI Insights")
    uploaded_file = st.file_uploader("Upload your filled Excel file to get AI-powered insights", type=["xlsx"])
    
    if uploaded_file is not None:
        if st.button("🤖 Generate AI Insights", key="ai_insights_btn"):
            with st.spinner("Analyzing your data with AI..."):
                # Save the uploaded file temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(uploaded_file.getvalue())
                    insights = generate_ai_insights(tmp.name)
                    os.unlink(tmp.name)
                
                st.success("AI Analysis Complete!")
                st.markdown("### 🎯 Your Personalized Insights")
                st.markdown(insights)
                
                # Add a button to download the insights
                st.download_button(
                    label="💾 Download Insights as Text",
                    data=insights,
                    file_name="financial_insights.txt",
                    mime="text/plain"
                )
    
    st.markdown("---")
    st.markdown("### 📱 Features at a Glance")
    
    features = st.columns(3)
    
    with features[0]:
        st.markdown("""
        #### 💰 Financial Tracking
        - Income & Expense tracking
        - Budget planning
        - Savings goals
        - Investment portfolio
        """)
    
    with features[1]:
        st.markdown("""
        #### 🏋️ Health & Wellness
        - Weight tracking
        - Habit formation
        - Self-care routines
        - Meal planning
        """)
    
    with features[2]:
        st.markdown("""
        #### 🏠 Life Organization
        - Cleaning checklist
        - Weekly schedule
        - Time management
        - Task tracking
        """)
    
    st.markdown("---")
    st.markdown("""
    ### Getting Started
    1. Download the Excel template above
    2. Fill in your financial and personal data
    3. Upload it back to get AI-powered insights
    4. Use the dashboard to track your progress
    """)

if __name__ == "__main__":
    main()